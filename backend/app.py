# backend/app.py
import os
import sys
import hashlib
import secrets
import re
import logging
import threading
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, Response, make_response
from flask_cors import CORS
from dotenv import load_dotenv
import utils
from database import get_db, init_db, get_all_cached_stats, save_all_cached_stats, USE_POSTGRES

# Configure logging for gunicorn
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# In-memory storage for mobile upload tokens (short-lived, ~5 min)
# Structure: { token: { user_id, created_at, transactions, status } }
_upload_tokens = {}
from auth import (
    verify_google_token, get_or_create_user, create_access_token,
    create_refresh_token, set_auth_cookies, clear_auth_cookies,
    get_current_user, login_required, admin_required
)
from email_utils import send_admin_deletion_alert

# Load environment-specific .env file
env = os.environ.get('FLASK_ENV', 'dev')
env_file = f'.env.{env}'
load_dotenv(env_file)

app = Flask(__name__)
# Allow React (running on localhost:5173) to talk to this API
# Support credentials for cookies
CORS(app, supports_credentials=True)

# Initialize database on startup
init_db() 

@app.route('/api/stats', methods=['GET'])
def get_chess_stats():
    username = request.args.get('username')
    time_class = request.args.get('time_class', 'rapid')  # Default to rapid

    if not username:
        return jsonify({"error": "Username required"}), 400

    if time_class not in ['rapid', 'blitz', 'bullet']:
        return jsonify({"error": "Invalid time_class. Use 'rapid', 'blitz', or 'bullet'"}), 400

    try:
        # 1. Fetch Player data
        player_data = utils.fetch_player_data_and_stats(username)

        # 2. Fetch archives once - reuse for all functions
        archives = utils.fetch_player_games_archives(username)

        # 3. Fetch History (weekly) - filtered by time class
        history = utils.fetch_games_played_per_week(username, time_class=time_class, archives=archives)

        # 4. Fetch Elo history - filtered by time class
        elo_history, total_games = utils.fetch_elo_per_week(username, time_class=time_class, archives=archives)

        # 5. Fetch Openings
        raw_openings = utils.fetch_all_openings(username, archives)
        processed_openings = utils.process_openings_for_json(raw_openings)

        # 6. Fetch win rate by game number per day
        game_number_stats = utils.fetch_win_rate_by_game_number(username, time_class=time_class, archives=archives)

        return jsonify({
            "player": {
                "name": player_data.get("name", username),
                "username": player_data.get("username", username),
                "avatar": player_data.get("avatar"),
                "followers": player_data.get("followers", 0),
                "joined": player_data.get("joined")
            },
            "time_class": time_class,
            "history": history,
            "elo_history": elo_history,
            "total_games": total_games,
            "openings": processed_openings,
            "game_number_stats": game_number_stats
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/stats-stream', methods=['GET'])
def get_chess_stats_stream():
    """SSE endpoint that streams progress while fetching stats. Uses caching for performance.

    Optimization: Fetches ALL time classes (rapid, blitz) in a single pass through archives.
    This makes switching between time classes instant after the first fetch.
    """
    import json as json_module

    username = request.args.get('username')
    time_class = request.args.get('time_class', 'rapid')

    if not username:
        return jsonify({"error": "Username required"}), 400

    if time_class not in ['rapid', 'blitz', 'bullet']:
        return jsonify({"error": "Invalid time_class"}), 400

    def generate():
        try:
            # First, fetch player data (always fresh for profile info)
            player_data = utils.fetch_player_data_and_stats(username)
            player_info = {
                'name': player_data.get('name', username),
                'username': player_data.get('username', username),
                'avatar': player_data.get('avatar'),
                'followers': player_data.get('followers', 0),
                'joined': player_data.get('joined')
            }
            yield f"data: {json_module.dumps({'type': 'player', 'player': player_info})}\n\n"

            # Check cache for ALL time classes
            all_cached = get_all_cached_stats(username)

            # Check if requested time class has fresh cache
            if time_class in all_cached:
                cached_stats, _, is_fresh = all_cached[time_class]
                if is_fresh:
                    # Return cached data immediately
                    yield f"data: {json_module.dumps({'type': 'start', 'total_archives': 0, 'incremental': False, 'cached': True})}\n\n"
                    yield f"data: {json_module.dumps({'type': 'complete', 'data': cached_stats})}\n\n"
                    return

            # Need to fetch - prepare cached stats map for incremental update
            cached_stats_map = None
            last_archive = None
            if all_cached:
                # Use any existing cache for incremental update
                cached_stats_map = {}
                for tc, (stats, archive, _) in all_cached.items():
                    cached_stats_map[tc] = stats
                    if archive:
                        last_archive = archive  # They should all be the same

            # Fetch stats for ALL time classes
            all_time_classes_data = None
            final_stats = None
            for chunk in utils.fetch_all_time_classes_streaming(username, time_class, cached_stats_map, last_archive):
                yield chunk
                # Parse the chunk to capture final data for caching
                if chunk.startswith('data: '):
                    try:
                        msg = json_module.loads(chunk[6:].strip())
                        if msg.get('type') == 'complete':
                            final_stats = msg.get('data')
                            all_time_classes_data = msg.get('all_time_classes')
                    except:
                        pass

            # Save ALL time classes to cache after successful fetch
            if all_time_classes_data:
                last_archive_new = None
                # Clean up last_archive from each time class data before caching
                for tc, tc_data in all_time_classes_data.items():
                    if 'last_archive' in tc_data:
                        last_archive_new = tc_data.pop('last_archive')
                save_all_cached_stats(username, player_info, all_time_classes_data, last_archive_new)

        except Exception as e:
            yield f"data: {json_module.dumps({'type': 'error', 'error': str(e)})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'  # Disable nginx buffering
        }
    )

@app.route('/api/fatigue-analysis', methods=['GET'])
def get_fatigue_analysis():
    username = request.args.get('username')
    time_class = request.args.get('time_class', 'rapid')

    if not username:
        return jsonify({"error": "Username required"}), 400

    if time_class not in ['rapid', 'blitz', 'bullet']:
        return jsonify({"error": "Invalid time_class"}), 400

    try:
        fatigue_analysis = utils.compute_fatigue_analysis(username, time_class=time_class)
        return jsonify(fatigue_analysis)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/win-prediction-stream', methods=['GET'])
def get_win_prediction_stream():
    """SSE endpoint that streams progress while analyzing win prediction patterns."""
    import json as json_module

    username = request.args.get('username')
    time_class = request.args.get('time_class', 'rapid')

    if not username:
        return jsonify({"error": "Username required"}), 400

    if time_class not in ['rapid', 'blitz', 'bullet']:
        return jsonify({"error": "Invalid time_class"}), 400

    def generate():
        try:
            for chunk in utils.compute_win_prediction_analysis_streaming(username, time_class=time_class):
                yield chunk
        except Exception as e:
            yield f"data: {json_module.dumps({'type': 'error', 'error': str(e)})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )

@app.route('/api/youtube-videos', methods=['GET'])
def get_youtube_videos():
    opening = request.args.get('opening')
    side = request.args.get('side')  # 'White' or 'Black' or 'tips'

    if not opening:
        return jsonify({"error": "Opening name required"}), 400

    api_key = os.environ.get('YOUTUBE_API_KEY')
    if not api_key:
        return jsonify({"error": "YouTube API key not configured"}), 500

    try:
        # Enable transcript scoring for pro tips videos
        use_transcript_scoring = (side == 'tips')
        videos = utils.fetch_youtube_videos(opening, side, api_key, use_transcript_scoring=use_transcript_scoring)
        return jsonify({"videos": videos})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ============= AUTH ROUTES =============

@app.route('/api/auth/google', methods=['POST'])
def google_auth():
    """Handle Google OAuth login with ID token."""
    data = request.get_json()
    google_token = data.get('credential')

    if not google_token:
        return jsonify({'error': 'No credential provided'}), 400

    # Verify Google token
    google_user = verify_google_token(google_token)
    if not google_user:
        return jsonify({'error': 'Invalid Google token'}), 401

    # Get or create user
    user_id = get_or_create_user(google_user)

    # Create tokens
    access_token = create_access_token(user_id)
    refresh_token, _ = create_refresh_token(user_id)

    # Get user data for response
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.*, up.chess_username, up.preferred_time_class
            FROM users u
            LEFT JOIN user_preferences up ON u.id = up.user_id
            WHERE u.id = ?
        ''', (user_id,))
        user = dict(cursor.fetchone())

    response = make_response(jsonify({
        'user': {
            'id': user['id'],
            'email': user['email'],
            'name': user['name'],
            'picture': user['picture'],
            'is_admin': bool(user.get('is_admin')),
            'cookie_consent': user.get('cookie_consent'),
            'preferences': {
                'chess_username': user['chess_username'],
                'preferred_time_class': user['preferred_time_class']
            }
        },
        'is_new_user': user.get('sign_in_count') == 1
    }))

    set_auth_cookies(response, access_token, refresh_token)
    return response


@app.route('/api/auth/refresh', methods=['POST'])
def refresh_auth():
    """Refresh access token using refresh token."""
    refresh_token = request.cookies.get('refresh_token')
    if not refresh_token:
        return jsonify({'error': 'No refresh token'}), 401

    token_hash = hashlib.sha256(refresh_token.encode()).hexdigest()

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT user_id, expires_at FROM refresh_tokens
            WHERE token_hash = ?
        ''', (token_hash,))
        row = cursor.fetchone()

        if not row:
            return jsonify({'error': 'Invalid refresh token'}), 401

        expires_at = row['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        if expires_at.replace(tzinfo=None) < datetime.utcnow():
            # Delete expired token
            conn.execute('DELETE FROM refresh_tokens WHERE token_hash = ?', (token_hash,))
            return jsonify({'error': 'Refresh token expired'}), 401

        user_id = row['user_id']

        # Rotate refresh token (delete old, create new)
        conn.execute('DELETE FROM refresh_tokens WHERE token_hash = ?', (token_hash,))

    # Create new tokens
    access_token = create_access_token(user_id)
    new_refresh_token, _ = create_refresh_token(user_id)

    response = make_response(jsonify({'success': True}))
    set_auth_cookies(response, access_token, new_refresh_token)
    return response


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Clear auth cookies and invalidate refresh token."""
    refresh_token = request.cookies.get('refresh_token')

    if refresh_token:
        token_hash = hashlib.sha256(refresh_token.encode()).hexdigest()
        with get_db() as conn:
            conn.execute('DELETE FROM refresh_tokens WHERE token_hash = ?', (token_hash,))

    response = make_response(jsonify({'success': True}))
    clear_auth_cookies(response)
    return response


@app.route('/api/auth/account', methods=['DELETE'])
@login_required
def delete_user_account():
    """Delete user account and all associated data."""
    user_id = request.user_id

    with get_db() as conn:
        # Get user info and stats before deletion
        cursor = conn.execute('SELECT name, email, created_at FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()

        cursor = conn.execute('SELECT COUNT(*) as count FROM portfolio_transactions WHERE user_id = ?', (user_id,))
        tx_count = cursor.fetchone()['count']

        cursor = conn.execute('SELECT COUNT(*) as count FROM investment_accounts WHERE user_id = ?', (user_id,))
        account_count = cursor.fetchone()['count']

        if user:
            # Send admin notification before deletion
            try:
                send_admin_deletion_alert(
                    user_name=user['name'],
                    user_email=user['email'],
                    deletion_type='user_account',
                    details={
                        'User ID': user_id,
                        'Account created': user['created_at'],
                        'Transactions deleted': tx_count,
                        'Investment accounts deleted': account_count,
                    }
                )
            except Exception as e:
                logger.error(f"Failed to send deletion alert: {e}")

        # Delete user (cascades to all related tables)
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))

    response = make_response(jsonify({'success': True}))
    clear_auth_cookies(response)
    return response


@app.route('/api/auth/me', methods=['GET'])
def get_current_user_info():
    """Get current user info if authenticated."""
    user_id = get_current_user()

    if not user_id:
        return jsonify({'user': None})

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.*, up.chess_username, up.preferred_time_class
            FROM users u
            LEFT JOIN user_preferences up ON u.id = up.user_id
            WHERE u.id = ?
        ''', (user_id,))
        row = cursor.fetchone()

        if not row:
            return jsonify({'user': None})

        user = dict(row)

    return jsonify({
        'user': {
            'id': user['id'],
            'email': user['email'],
            'name': user['name'],
            'picture': user['picture'],
            'is_admin': bool(user.get('is_admin')),
            'cookie_consent': user.get('cookie_consent'),
            'preferences': {
                'chess_username': user['chess_username'],
                'preferred_time_class': user['preferred_time_class']
            }
        }
    })


# ============= PREFERENCES ROUTES =============

@app.route('/api/preferences', methods=['GET'])
@login_required
def get_preferences():
    """Get user preferences."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT chess_username, preferred_time_class
            FROM user_preferences WHERE user_id = ?
        ''', (request.user_id,))
        row = cursor.fetchone()

    return jsonify(dict(row) if row else {})


@app.route('/api/preferences', methods=['PUT'])
@login_required
def update_preferences():
    """Update user preferences."""
    data = request.get_json()

    allowed_fields = ['chess_username', 'preferred_time_class']
    updates = {k: v for k, v in data.items() if k in allowed_fields}

    if not updates:
        return jsonify({'error': 'No valid fields to update'}), 400

    # Validate time_class
    if 'preferred_time_class' in updates:
        if updates['preferred_time_class'] not in ['rapid', 'blitz', 'bullet']:
            return jsonify({'error': 'Invalid time_class'}), 400

    set_clause = ', '.join(f'{k} = ?' for k in updates.keys())
    values = list(updates.values()) + [request.user_id]

    with get_db() as conn:
        conn.execute(f'''
            UPDATE user_preferences
            SET {set_clause}, updated_at = CURRENT_TIMESTAMP
            WHERE user_id = ?
        ''', values)

    return jsonify({'success': True, 'preferences': updates})


@app.route('/api/preferences/dashboard-card-order', methods=['GET'])
@login_required
def get_dashboard_card_order():
    """Get user's dashboard card order."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT dashboard_card_order
            FROM user_preferences WHERE user_id = ?
        ''', (request.user_id,))
        row = cursor.fetchone()

    if row and row['dashboard_card_order']:
        import json
        try:
            return jsonify({'order': json.loads(row['dashboard_card_order'])})
        except json.JSONDecodeError:
            return jsonify({'order': None})
    return jsonify({'order': None})


@app.route('/api/preferences/dashboard-card-order', methods=['PUT'])
@login_required
def update_dashboard_card_order():
    """Update user's dashboard card order."""
    import json
    data = request.get_json()
    order = data.get('order')

    if not isinstance(order, list):
        return jsonify({'error': 'Order must be a list'}), 400

    order_json = json.dumps(order)

    with get_db() as conn:
        # Check if user_preferences row exists
        cursor = conn.execute(
            'SELECT id FROM user_preferences WHERE user_id = ?',
            (request.user_id,)
        )
        if cursor.fetchone():
            conn.execute('''
                UPDATE user_preferences
                SET dashboard_card_order = ?, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = ?
            ''', (order_json, request.user_id))
        else:
            conn.execute('''
                INSERT INTO user_preferences (user_id, dashboard_card_order)
                VALUES (?, ?)
            ''', (request.user_id, order_json))

    return jsonify({'success': True, 'order': order})


@app.route('/api/preferences/demo-card-order', methods=['GET'])
def get_demo_card_order():
    """Get card order for demo user (public endpoint for unauthenticated preview)."""
    demo_email = 'rose.louis.mail@gmail.com'

    with get_db() as conn:
        # Find demo user's ID
        if USE_POSTGRES:
            cursor = conn.execute('SELECT id FROM users WHERE email = %s', (demo_email,))
        else:
            cursor = conn.execute('SELECT id FROM users WHERE email = ?', (demo_email,))
        user_row = cursor.fetchone()

        if not user_row:
            return jsonify({'order': None})

        # Get their card order
        if USE_POSTGRES:
            cursor = conn.execute('''
                SELECT dashboard_card_order
                FROM user_preferences WHERE user_id = %s
            ''', (user_row['id'],))
        else:
            cursor = conn.execute('''
                SELECT dashboard_card_order
                FROM user_preferences WHERE user_id = ?
            ''', (user_row['id'],))
        row = cursor.fetchone()

    if row and row['dashboard_card_order']:
        import json
        try:
            return jsonify({'order': json.loads(row['dashboard_card_order'])})
        except json.JSONDecodeError:
            return jsonify({'order': None})
    return jsonify({'order': None})


@app.route('/api/demo-dashboard', methods=['GET'])
def get_demo_dashboard():
    """Get complete dashboard data for demo user (public endpoint for unauthenticated preview).
    Returns all the data needed to render the dashboard cards with real demo user data.
    """
    import json
    from investing_utils import fetch_current_stock_prices_batch, fetch_historical_prices_batch, get_fx_rate_to_eur, get_stock_currency

    demo_email = 'rose.louis.mail@gmail.com'

    with get_db() as conn:
        # Find demo user's ID
        if USE_POSTGRES:
            cursor = conn.execute('SELECT id FROM users WHERE email = %s', (demo_email,))
        else:
            cursor = conn.execute('SELECT id FROM users WHERE email = ?', (demo_email,))
        user_row = cursor.fetchone()

        if not user_row:
            return jsonify({'error': 'Demo user not found'}), 404

        demo_user_id = user_row['id']

        # Get card order
        if USE_POSTGRES:
            cursor = conn.execute('SELECT dashboard_card_order FROM user_preferences WHERE user_id = %s', (demo_user_id,))
        else:
            cursor = conn.execute('SELECT dashboard_card_order FROM user_preferences WHERE user_id = ?', (demo_user_id,))
        pref_row = cursor.fetchone()
        card_order = None
        if pref_row and pref_row['dashboard_card_order']:
            try:
                card_order = json.loads(pref_row['dashboard_card_order'])
            except json.JSONDecodeError:
                pass

        # Get all account IDs for demo user
        if USE_POSTGRES:
            cursor = conn.execute('SELECT id FROM investment_accounts WHERE user_id = %s', (demo_user_id,))
        else:
            cursor = conn.execute('SELECT id FROM investment_accounts WHERE user_id = ?', (demo_user_id,))
        account_ids = [row['id'] for row in cursor.fetchall()]

        # Get watchlist tickers
        if USE_POSTGRES:
            cursor = conn.execute('SELECT stock_ticker FROM watchlist WHERE user_id = %s', (demo_user_id,))
        else:
            cursor = conn.execute('SELECT stock_ticker FROM watchlist WHERE user_id = ?', (demo_user_id,))
        watchlist_tickers = [row['stock_ticker'] for row in cursor.fetchall()]

    # Get portfolio composition
    holdings = compute_holdings_from_transactions(demo_user_id, account_ids if account_ids else None)
    composition = None
    if holdings:
        try:
            composition = compute_portfolio_composition(holdings)
        except Exception:
            pass

    # Get performance data (1, 7, and 30 days)
    perf_1 = None
    perf_7 = None
    perf_30 = None
    if holdings:
        today = datetime.now()
        all_tickers = [h['stock_ticker'] for h in holdings]
        prices = fetch_current_stock_prices_batch(all_tickers)

        for days in [1, 7, 30]:
            past_date = (today - timedelta(days=days)).strftime('%Y-%m-%d')
            current_value = 0
            past_value = 0

            for h in holdings:
                ticker = h['stock_ticker']
                qty = h['quantity']
                current_price = prices.get(ticker, 0) or 0
                try:
                    past_price = fetch_stock_price(ticker, past_date)
                    if past_price is None:
                        past_price = current_price
                except Exception:
                    past_price = current_price

                currency = get_stock_currency(ticker)
                fx_rate = get_fx_rate_to_eur(currency)
                current_value += (current_price or 0) * qty * fx_rate
                past_value += (past_price or 0) * qty * fx_rate

            if past_value > 0:
                perf_pct = ((current_value - past_value) / past_value) * 100
                if days == 1:
                    perf_1 = round(perf_pct, 2)
                elif days == 7:
                    perf_7 = round(perf_pct, 2)
                else:
                    perf_30 = round(perf_pct, 2)

    # Get portfolio top movers (30 days)
    portfolio_movers = []
    if holdings:
        today = datetime.now()
        past_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')
        all_tickers = [h['stock_ticker'] for h in holdings]

        # Batch fetch current, period, and 1-day historical prices
        current_prices = fetch_current_stock_prices_batch(all_tickers)
        past_prices = fetch_historical_prices_batch(all_tickers, past_date)
        yesterday_prices = fetch_historical_prices_batch(all_tickers, yesterday)

        for h in holdings:
            ticker = h['stock_ticker']
            current_price = current_prices.get(ticker, 0) or 0
            past_price = past_prices.get(ticker, 0) or 0
            yesterday_price = yesterday_prices.get(ticker, 0) or 0

            if current_price <= 0 or past_price <= 0:
                continue

            change_pct = ((current_price - past_price) / past_price) * 100
            change_1d = None
            if yesterday_price > 0:
                change_1d = round(((current_price - yesterday_price) / yesterday_price) * 100, 1)

            portfolio_movers.append({
                'ticker': ticker,
                'change_pct': round(change_pct, 1),
                'change_1d': change_1d,
                'current_price': round(current_price, 2),
                'past_price': round(past_price, 2)
            })
        portfolio_movers.sort(key=lambda x: x['change_pct'], reverse=True)
        portfolio_movers = portfolio_movers[:5]

    # Get watchlist top movers (30 days)
    watchlist_movers = []
    if watchlist_tickers:
        today = datetime.now()
        past_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')

        # Batch fetch current, period, and 1-day historical prices
        current_prices = fetch_current_stock_prices_batch(watchlist_tickers)
        past_prices = fetch_historical_prices_batch(watchlist_tickers, past_date)
        yesterday_prices = fetch_historical_prices_batch(watchlist_tickers, yesterday)

        for ticker in watchlist_tickers:
            current_price = current_prices.get(ticker, 0) or 0
            past_price = past_prices.get(ticker, 0) or 0
            yesterday_price = yesterday_prices.get(ticker, 0) or 0

            if current_price <= 0 or past_price <= 0:
                continue

            change_pct = ((current_price - past_price) / past_price) * 100
            change_1d = None
            if yesterday_price > 0:
                change_1d = round(((current_price - yesterday_price) / yesterday_price) * 100, 1)

            watchlist_movers.append({
                'ticker': ticker,
                'change_pct': round(change_pct, 1),
                'change_1d': change_1d,
                'current_price': round(current_price, 2),
                'past_price': round(past_price, 2)
            })
        watchlist_movers.sort(key=lambda x: x['change_pct'], reverse=True)
        watchlist_movers = watchlist_movers[:5]

    # Get earnings data (portfolio only)
    earnings_data = []
    if holdings:
        today_date = datetime.now().date()
        portfolio_tickers = {h['stock_ticker'] for h in holdings if h['quantity'] > 0}

        for ticker in portfolio_tickers:
            cached_date, cached_confirmed, _, is_fresh = get_cached_earnings(ticker)
            if is_fresh and cached_date:
                try:
                    earnings_date = datetime.strptime(cached_date, '%Y-%m-%d').date()
                    remaining_days = (earnings_date - today_date).days
                    if remaining_days >= 0:
                        earnings_data.append({
                            'ticker': ticker,
                            'next_earnings_date': cached_date,
                            'remaining_days': remaining_days,
                            'date_confirmed': cached_confirmed,
                            'source': 'portfolio'
                        })
                except Exception:
                    pass
        earnings_data.sort(key=lambda x: (x['remaining_days'] is None, x['remaining_days'] or 9999))
        earnings_data = earnings_data[:15]

    # Get dividends data from cache (7 day TTL)
    dividends_data = []
    if holdings:
        today_date = datetime.now().date()

        for h in holdings:
            ticker = h['stock_ticker']
            qty = h['quantity']
            if qty <= 0:
                continue

            # Try to get cached dividend data
            ex_date_str, div_amount, pays_dividends, is_fresh = get_cached_dividend(ticker)
            if is_fresh:
                if not pays_dividends:
                    dividends_data.append({
                        'ticker': ticker,
                        'ex_dividend_date': None,
                        'remaining_days': None,
                        'dividend_amount': None,
                        'pays_dividends': False,
                        'quantity': qty,
                        'total_dividend': None
                    })
                elif ex_date_str:
                    try:
                        ex_date = datetime.strptime(ex_date_str, '%Y-%m-%d').date()
                        remaining_days = (ex_date - today_date).days
                        if remaining_days >= 0:
                            dividends_data.append({
                                'ticker': ticker,
                                'ex_dividend_date': ex_date_str,
                                'remaining_days': remaining_days,
                                'dividend_amount': div_amount,
                                'pays_dividends': True,
                                'quantity': qty,
                                'total_dividend': round(div_amount * qty, 2) if div_amount else None
                            })
                    except Exception:
                        pass

        # Sort: dividend payers first, then by remaining days
        dividends_data.sort(key=lambda x: (
            x.get('pays_dividends') == False,
            x['remaining_days'] is None,
            x['remaining_days'] or 9999
        ))
        dividends_data = dividends_data[:10]

    return jsonify({
        'card_order': card_order,
        'composition': composition,
        'performance_1': perf_1,
        'performance_7': perf_7,
        'performance_30': perf_30,
        'portfolio_movers': portfolio_movers,
        'watchlist_movers': watchlist_movers,
        'earnings': earnings_data,
        'dividends': dividends_data
    })


@app.route('/api/preferences/financial-card-order', methods=['GET'])
@login_required
def get_financial_card_order():
    """Get user's financial charts card order."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT financial_card_order
            FROM user_preferences WHERE user_id = ?
        ''', (request.user_id,))
        row = cursor.fetchone()

    if row and row['financial_card_order']:
        import json
        try:
            return jsonify({'order': json.loads(row['financial_card_order'])})
        except json.JSONDecodeError:
            return jsonify({'order': None})
    return jsonify({'order': None})


@app.route('/api/preferences/financial-card-order', methods=['PUT'])
@login_required
def update_financial_card_order():
    """Update user's financial charts card order."""
    import json
    data = request.get_json()
    order = data.get('order')

    if not isinstance(order, list):
        return jsonify({'error': 'Order must be a list'}), 400

    order_json = json.dumps(order)

    with get_db() as conn:
        # Check if user_preferences row exists
        cursor = conn.execute(
            'SELECT id FROM user_preferences WHERE user_id = ?',
            (request.user_id,)
        )
        if cursor.fetchone():
            conn.execute('''
                UPDATE user_preferences
                SET financial_card_order = ?, updated_at = CURRENT_TIMESTAMP
                WHERE user_id = ?
            ''', (order_json, request.user_id))
        else:
            conn.execute('''
                INSERT INTO user_preferences (user_id, financial_card_order)
                VALUES (?, ?)
            ''', (request.user_id, order_json))

    return jsonify({'success': True, 'order': order})


# ============= ACTIVITY TRACKING =============

@app.route('/api/activity/heartbeat', methods=['POST'])
@login_required
def activity_heartbeat():
    """Record a heartbeat for activity tracking (called every 60s by frontend)."""
    today = datetime.now().strftime('%Y-%m-%d')
    data = request.get_json() or {}
    page = data.get('page', 'other')

    # Settings data (optional, sent with heartbeat)
    theme = data.get('theme')
    resolved_theme = data.get('resolved_theme')
    language = data.get('language')
    device_type = data.get('device_type')

    # Normalize page names to categories
    if page.startswith('stock/'):
        page = 'stock'  # Aggregate all company pages
    elif page not in ('portfolio', 'watchlist', 'earnings', 'financials', 'admin'):
        page = 'other'

    with get_db() as conn:
        # Check if this is a new session (30+ min since last ping)
        cursor = conn.execute(
            'SELECT last_session_ping FROM users WHERE id = ?',
            (request.user_id,)
        )
        row = cursor.fetchone()
        last_ping = row['last_session_ping'] if row else None

        is_new_session = True
        if last_ping:
            try:
                last_ping_time = datetime.fromisoformat(last_ping.replace('Z', '+00:00')) if isinstance(last_ping, str) else last_ping
                minutes_since_last = (datetime.utcnow() - last_ping_time.replace(tzinfo=None)).total_seconds() / 60
                is_new_session = minutes_since_last > 30
            except:
                is_new_session = True

        # Update session tracking
        if is_new_session:
            conn.execute('''
                UPDATE users SET session_count = COALESCE(session_count, 0) + 1, last_session_ping = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (request.user_id,))
        else:
            conn.execute('''
                UPDATE users SET last_session_ping = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (request.user_id,))

        # Track daily activity
        conn.execute('''
            INSERT INTO user_activity (user_id, activity_date, minutes, last_ping)
            VALUES (?, ?, 1, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id, activity_date) DO UPDATE SET
                minutes = user_activity.minutes + 1,
                last_ping = CURRENT_TIMESTAMP
        ''', (request.user_id, today))

        # Track page-level activity
        conn.execute('''
            INSERT INTO page_activity (user_id, page, minutes)
            VALUES (?, ?, 1)
            ON CONFLICT(user_id, page) DO UPDATE SET
                minutes = page_activity.minutes + 1
        ''', (request.user_id, page))

        # Track theme preference (if provided)
        if theme and resolved_theme:
            conn.execute('''
                INSERT INTO theme_usage (user_id, theme, resolved_theme, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(user_id) DO UPDATE SET
                    theme = excluded.theme,
                    resolved_theme = excluded.resolved_theme,
                    updated_at = CURRENT_TIMESTAMP
            ''', (request.user_id, theme, resolved_theme))

        # Track language preference (if provided)
        if language:
            conn.execute('''
                INSERT INTO language_usage (user_id, language, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(user_id) DO UPDATE SET
                    language = excluded.language,
                    updated_at = CURRENT_TIMESTAMP
            ''', (request.user_id, language))

        # Track device usage minutes (if provided)
        if device_type in ('mobile', 'desktop'):
            conn.execute('''
                INSERT INTO device_usage (user_id, device_type, minutes, updated_at)
                VALUES (?, ?, 1, CURRENT_TIMESTAMP)
                ON CONFLICT(user_id, device_type) DO UPDATE SET
                    minutes = device_usage.minutes + 1,
                    updated_at = CURRENT_TIMESTAMP
            ''', (request.user_id, device_type))

    return jsonify({'success': True})


@app.route('/api/theme', methods=['POST'])
@login_required
def record_theme():
    """Record user's theme preference for analytics."""
    data = request.get_json()
    theme = data.get('theme')  # 'light', 'dark', 'system'
    resolved_theme = data.get('resolved_theme')  # 'light' or 'dark'

    if not theme or not resolved_theme:
        return jsonify({'error': 'theme and resolved_theme required'}), 400

    with get_db() as conn:
        conn.execute('''
            INSERT INTO theme_usage (user_id, theme, resolved_theme, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id) DO UPDATE SET
                theme = excluded.theme,
                resolved_theme = excluded.resolved_theme,
                updated_at = CURRENT_TIMESTAMP
        ''', (request.user_id, theme, resolved_theme))

    return jsonify({'success': True})


@app.route('/api/language', methods=['POST'])
@login_required
def record_language():
    """Record user's language preference for analytics."""
    data = request.get_json()
    language = data.get('language')  # 'en' or 'fr'

    if not language:
        return jsonify({'error': 'language required'}), 400

    with get_db() as conn:
        conn.execute('''
            INSERT INTO language_usage (user_id, language, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id) DO UPDATE SET
                language = excluded.language,
                updated_at = CURRENT_TIMESTAMP
        ''', (request.user_id, language))

    return jsonify({'success': True})


@app.route('/api/cookie-consent', methods=['POST'])
@login_required
def save_cookie_consent():
    """Save user's cookie consent. Only 'accepted' is persisted; refusals increment a counter."""
    data = request.get_json()
    consent = data.get('consent')  # 'accepted' or 'refused'

    if consent not in ('accepted', 'refused'):
        return jsonify({'error': 'consent must be "accepted" or "refused"'}), 400

    with get_db() as conn:
        if consent == 'accepted':
            # Store acceptance permanently
            conn.execute('''
                UPDATE users SET cookie_consent = 'accepted', cookie_consent_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (request.user_id,))
        else:
            # Increment refusal count (but don't set cookie_consent, so they'll be asked again)
            conn.execute('''
                UPDATE users SET cookie_refusal_count = COALESCE(cookie_refusal_count, 0) + 1
                WHERE id = ?
            ''', (request.user_id,))

    return jsonify({'success': True, 'consent': consent if consent == 'accepted' else None})


@app.route('/api/device', methods=['POST'])
@login_required
def record_device():
    """Record user's device type for analytics."""
    data = request.get_json()
    device_type = data.get('device_type')  # 'mobile' or 'desktop'

    if device_type not in ('mobile', 'desktop'):
        return jsonify({'error': 'device_type must be mobile or desktop'}), 400

    with get_db() as conn:
        conn.execute('''
            INSERT INTO device_usage (user_id, device_type, minutes, updated_at)
            VALUES (?, ?, 0, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id, device_type) DO UPDATE SET
                updated_at = CURRENT_TIMESTAMP
        ''', (request.user_id, device_type))

    return jsonify({'success': True})


# ============= ADMIN ROUTES =============

@app.route('/api/admin/theme-stats', methods=['GET'])
@admin_required
def get_theme_stats():
    """Get theme usage statistics (admin only). Excludes admin users."""
    with get_db() as conn:
        # Get counts by resolved theme (actual display) - exclude admins
        cursor = conn.execute('''
            SELECT t.resolved_theme, COUNT(*) as count
            FROM theme_usage t
            INNER JOIN users u ON t.user_id = u.id
            WHERE u.is_admin = 0
            GROUP BY t.resolved_theme
        ''')
        by_resolved = {row['resolved_theme']: row['count'] for row in cursor.fetchall()}

        # Get counts by theme setting (includes 'system') - exclude admins
        cursor = conn.execute('''
            SELECT t.theme, COUNT(*) as count
            FROM theme_usage t
            INNER JOIN users u ON t.user_id = u.id
            WHERE u.is_admin = 0
            GROUP BY t.theme
        ''')
        by_setting = {row['theme']: row['count'] for row in cursor.fetchall()}

        # Get total users with theme data - exclude admins
        cursor = conn.execute('''
            SELECT COUNT(*) as total
            FROM theme_usage t
            INNER JOIN users u ON t.user_id = u.id
            WHERE u.is_admin = 0
        ''')
        total = cursor.fetchone()['total']

    return jsonify({
        'total': total,
        'by_resolved': by_resolved,
        'by_setting': by_setting
    })


@app.route('/api/admin/language-stats', methods=['GET'])
@admin_required
def get_language_stats():
    """Get language usage statistics (admin only). Excludes admin users."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT l.language, COUNT(*) as count
            FROM language_usage l
            INNER JOIN users u ON l.user_id = u.id
            WHERE u.is_admin = 0
            GROUP BY l.language
        ''')
        by_language = {row['language']: row['count'] for row in cursor.fetchall()}

        cursor = conn.execute('''
            SELECT COUNT(*) as total
            FROM language_usage l
            INNER JOIN users u ON l.user_id = u.id
            WHERE u.is_admin = 0
        ''')
        total = cursor.fetchone()['total']

    return jsonify({
        'total': total,
        'by_language': by_language
    })


@app.route('/api/admin/device-stats', methods=['GET'])
@admin_required
def get_device_stats():
    """Get device type usage statistics (admin only). Excludes admin users."""
    with get_db() as conn:
        # Get total minutes per device type - exclude admins
        cursor = conn.execute('''
            SELECT d.device_type, SUM(d.minutes) as total_minutes
            FROM device_usage d
            INNER JOIN users u ON d.user_id = u.id
            WHERE u.is_admin = 0
            GROUP BY d.device_type
        ''')
        by_device = {row['device_type']: row['total_minutes'] for row in cursor.fetchall()}

        # Get total users with device data - exclude admins
        cursor = conn.execute('''
            SELECT COUNT(DISTINCT d.user_id) as total
            FROM device_usage d
            INNER JOIN users u ON d.user_id = u.id
            WHERE u.is_admin = 0
        ''')
        total = cursor.fetchone()['total']

        # Calculate total minutes for percentage
        total_minutes = sum(by_device.values())

    return jsonify({
        'total': total,
        'total_minutes': total_minutes,
        'by_device': by_device
    })


@app.route('/api/admin/users-by-theme/<theme>', methods=['GET'])
@admin_required
def get_users_by_theme(theme):
    """Get list of users with a specific resolved theme (admin only). Excludes admin users."""
    if theme not in ('dark', 'light'):
        return jsonify({'error': 'Invalid theme'}), 400

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.id, u.name, u.picture
            FROM users u
            INNER JOIN theme_usage t ON u.id = t.user_id
            WHERE t.resolved_theme = ? AND u.is_admin = 0
            ORDER BY u.name
        ''', (theme,))
        users = [{'id': row['id'], 'name': row['name'], 'picture': row['picture']} for row in cursor.fetchall()]

    return jsonify({'users': users})


@app.route('/api/admin/users-by-language/<lang>', methods=['GET'])
@admin_required
def get_users_by_language(lang):
    """Get list of users with a specific language setting (admin only). Excludes admin users."""
    if lang not in ('en', 'fr'):
        return jsonify({'error': 'Invalid language'}), 400

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.id, u.name, u.picture
            FROM users u
            INNER JOIN language_usage l ON u.id = l.user_id
            WHERE l.language = ? AND u.is_admin = 0
            ORDER BY u.name
        ''', (lang,))
        users = [{'id': row['id'], 'name': row['name'], 'picture': row['picture']} for row in cursor.fetchall()]

    return jsonify({'users': users})


@app.route('/api/admin/users-by-device/<device>', methods=['GET'])
@admin_required
def get_users_by_device(device):
    """Get list of users with a specific device type (admin only). Excludes admin users."""
    if device not in ('mobile', 'desktop'):
        return jsonify({'error': 'Invalid device type'}), 400

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.id, u.name, u.picture, d.minutes
            FROM users u
            INNER JOIN device_usage d ON u.id = d.user_id
            WHERE d.device_type = ? AND u.is_admin = 0
            ORDER BY d.minutes DESC
        ''', (device,))
        users = [{'id': row['id'], 'name': row['name'], 'picture': row['picture'], 'minutes': row['minutes']} for row in cursor.fetchall()]

    return jsonify({'users': users})


@app.route('/api/admin/settings-crosstab', methods=['GET'])
@admin_required
def get_settings_crosstab():
    """Get cross-tabulation of theme x language, weighted by time spent."""
    with get_db() as conn:
        # Join theme_usage, language_usage, and user activity to get weighted crosstab
        cursor = conn.execute('''
            SELECT
                t.resolved_theme,
                l.language,
                COUNT(*) as user_count,
                COALESCE(SUM(u.total_minutes), 0) as total_minutes
            FROM theme_usage t
            INNER JOIN language_usage l ON t.user_id = l.user_id
            LEFT JOIN (
                SELECT user_id, SUM(duration_minutes) as total_minutes
                FROM user_activity
                GROUP BY user_id
            ) u ON t.user_id = u.user_id
            GROUP BY t.resolved_theme, l.language
        ''')

        results = cursor.fetchall()

        # Build crosstab data
        crosstab = {}
        total_minutes = 0
        total_users = 0

        for row in results:
            theme = row['resolved_theme']
            lang = row['language']
            minutes = row['total_minutes']
            users = row['user_count']

            key = f"{theme}_{lang}"
            crosstab[key] = {
                'users': users,
                'minutes': minutes
            }
            total_minutes += minutes
            total_users += users

    return jsonify({
        'crosstab': crosstab,
        'total_minutes': total_minutes,
        'total_users': total_users
    })


@app.route('/api/admin/users', methods=['GET'])
@admin_required
def list_users():
    """List all registered users (admin only)."""
    # Hidden accounts (still functional, just not displayed)
    hidden_emails = []  # No hidden accounts

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.id, u.email, u.name, u.picture, u.is_admin, u.created_at, u.updated_at, u.sign_in_count, u.session_count,
                   COALESCE(SUM(a.minutes), 0) as total_minutes,
                   MAX(a.last_ping) as last_active,
                   (SELECT COUNT(*) FROM graph_downloads g WHERE g.user_id = u.id) as graph_downloads,
                   (SELECT COUNT(*) FROM investment_accounts ia WHERE ia.user_id = u.id) as account_count,
                   (SELECT COUNT(DISTINCT stock_ticker) FROM portfolio_transactions pt WHERE pt.user_id = u.id) as portfolio_companies,
                   (SELECT COUNT(DISTINCT stock_ticker) FROM watchlist w WHERE w.user_id = u.id) as watchlist_companies
            FROM users u
            LEFT JOIN user_activity a ON u.id = a.user_id
            GROUP BY u.id
            ORDER BY u.created_at DESC
        ''')
        users = []
        for row in cursor.fetchall():
            if row['email'] in hidden_emails:
                continue
            user = dict(row)
            # Convert datetime objects to ISO strings for consistent JSON serialization
            if user.get('created_at') and hasattr(user['created_at'], 'isoformat'):
                user['created_at'] = user['created_at'].isoformat()
            if user.get('updated_at') and hasattr(user['updated_at'], 'isoformat'):
                user['updated_at'] = user['updated_at'].isoformat()
            if user.get('last_active') and hasattr(user['last_active'], 'isoformat'):
                user['last_active'] = user['last_active'].isoformat()
            users.append(user)

    return jsonify({'users': users, 'total': len(users)})


@app.route('/api/admin/users/<int:user_id>/activity', methods=['GET'])
@admin_required
def get_user_activity(user_id):
    """Get daily activity breakdown for a user (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT activity_date, minutes
            FROM user_activity
            WHERE user_id = ?
            ORDER BY activity_date DESC
        ''', (user_id,))
        activity = [dict(row) for row in cursor.fetchall()]

    return jsonify({'activity': activity})


@app.route('/api/admin/users/<int:user_id>/accounts', methods=['GET'])
@admin_required
def get_user_accounts(user_id):
    """Get investment accounts for a user (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT id, name, account_type, bank, created_at
            FROM investment_accounts
            WHERE user_id = ?
            ORDER BY created_at ASC
        ''', (user_id,))
        accounts = [dict(row) for row in cursor.fetchall()]

    return jsonify({'accounts': accounts})


@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@admin_required
def get_user_detail(user_id):
    """Get detailed info for a specific user (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT u.id, u.email, u.name, u.picture, u.is_admin, u.created_at, u.updated_at,
                   COALESCE(SUM(a.minutes), 0) as total_minutes,
                   MAX(a.last_ping) as last_active
            FROM users u
            LEFT JOIN user_activity a ON u.id = a.user_id
            WHERE u.id = ?
            GROUP BY u.id
        ''', (user_id,))
        user = cursor.fetchone()

    if not user:
        return jsonify({'error': 'User not found'}), 404

    user_dict = dict(user)
    # Convert datetime objects to ISO strings for JSON serialization
    if user_dict.get('created_at') and hasattr(user_dict['created_at'], 'isoformat'):
        user_dict['created_at'] = user_dict['created_at'].isoformat()
    if user_dict.get('updated_at') and hasattr(user_dict['updated_at'], 'isoformat'):
        user_dict['updated_at'] = user_dict['updated_at'].isoformat()
    if user_dict.get('last_active') and hasattr(user_dict['last_active'], 'isoformat'):
        user_dict['last_active'] = user_dict['last_active'].isoformat()

    return jsonify({'user': user_dict})


@app.route('/api/admin/users/<int:user_id>/watchlist', methods=['GET'])
@admin_required
def get_user_watchlist(user_id):
    """Get a user's watchlist (admin only)."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT stock_ticker FROM watchlist WHERE user_id = ? ORDER BY created_at ASC',
            (user_id,)
        )
        rows = cursor.fetchall()

    symbols = [row['stock_ticker'] for row in rows]
    return jsonify({'symbols': symbols})


@app.route('/api/admin/users/<int:user_id>/portfolio', methods=['GET'])
@admin_required
def get_user_portfolio(user_id):
    """Get a user's portfolio composition grouped by account (admin only)."""
    # Get all accounts for this user
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT id, name, account_type, bank, created_at
            FROM investment_accounts
            WHERE user_id = ?
            ORDER BY created_at ASC
        ''', (user_id,))
        accounts = [dict(row) for row in cursor.fetchall()]

    result = []
    total_value_eur = 0

    for account in accounts:
        holdings = compute_holdings_from_transactions(user_id, [account['id']])
        if holdings:
            try:
                composition = compute_portfolio_composition(holdings)
                account_data = {
                    'account': account,
                    'holdings': composition.get('holdings', []),
                    'total_value_eur': composition.get('total_value_eur', 0)
                }
                total_value_eur += composition.get('total_value_eur', 0)
            except:
                account_data = {
                    'account': account,
                    'holdings': [],
                    'total_value_eur': 0
                }
        else:
            account_data = {
                'account': account,
                'holdings': [],
                'total_value_eur': 0
            }
        result.append(account_data)

    return jsonify({
        'accounts': result,
        'total_value_eur': total_value_eur
    })


@app.route('/api/admin/users/<int:user_id>/graph-downloads', methods=['GET'])
@admin_required
def get_user_graph_downloads(user_id):
    """Get a user's graph download history (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT id, graph_type, downloaded_at
            FROM graph_downloads
            WHERE user_id = ?
            ORDER BY downloaded_at DESC
        ''', (user_id,))
        downloads = [dict(row) for row in cursor.fetchall()]

    return jsonify({'downloads': downloads})


# ============= INVESTING ROUTES =============

from investing_utils import (
    compute_portfolio_composition, compute_portfolio_performance_from_transactions,
    fetch_stock_price, get_previous_weekday, set_db_getter, get_news_feed_videos,
    get_stock_currency
)

# Initialize database getter for caching in investing_utils
set_db_getter(get_db)

@app.route('/api/investing/transactions', methods=['GET'])
@login_required
def get_transactions():
    """Get user's transaction history."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT pt.id, pt.stock_ticker, pt.transaction_type, pt.quantity,
                      pt.transaction_date, pt.price_per_share, pt.price_currency, pt.account_id,
                      ia.name as account_name, ia.account_type, ia.bank
               FROM portfolio_transactions pt
               LEFT JOIN investment_accounts ia ON pt.account_id = ia.id
               WHERE pt.user_id = ?
               ORDER BY pt.transaction_date DESC, pt.id DESC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    transactions = [{
        'id': row['id'],
        'stock_ticker': row['stock_ticker'],
        'transaction_type': row['transaction_type'],
        'quantity': row['quantity'],
        'transaction_date': row['transaction_date'],
        'price_per_share': row['price_per_share'],
        'price_currency': row['price_currency'] or 'EUR',  # Default to EUR for legacy data
        'account_id': row['account_id'],
        'account_name': row['account_name'],
        'account_type': row['account_type'],
        'bank': row['bank'],
    } for row in rows]
    return jsonify({'transactions': transactions})


@app.route('/api/investing/transactions', methods=['POST'])
@login_required
def add_transaction():
    """Add a new transaction (BUY or SELL) with auto-fetched price."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    stock_ticker = data.get('stock_ticker', '').upper().strip()
    transaction_type = data.get('transaction_type', '').upper().strip()
    quantity = data.get('quantity')
    transaction_date = data.get('transaction_date')  # YYYY-MM-DD format
    account_id = data.get('account_id')  # Optional: link to investment account
    provided_price = data.get('price_per_share')  # Optional: price from import
    provided_currency = data.get('price_currency')  # Optional: currency of provided price

    if not stock_ticker:
        return jsonify({'error': 'Stock ticker required'}), 400
    if transaction_type not in ['BUY', 'SELL']:
        return jsonify({'error': 'Transaction type must be BUY or SELL'}), 400
    if quantity is None or not isinstance(quantity, (int, float)) or quantity <= 0:
        return jsonify({'error': 'Valid quantity required (must be > 0)'}), 400
    if not transaction_date:
        return jsonify({'error': 'Transaction date required (YYYY-MM-DD)'}), 400

    quantity = round(float(quantity), 2)  # Keep 2 decimal places for fractional shares

    # Validate account_id if provided
    if account_id is not None:
        with get_db() as conn:
            cursor = conn.execute(
                'SELECT id FROM investment_accounts WHERE id = ? AND user_id = ?',
                (account_id, request.user_id)
            )
            if not cursor.fetchone():
                return jsonify({'error': 'Invalid account_id'}), 400

    # Validate date format
    try:
        datetime.strptime(transaction_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # Adjust for weekends (markets closed)
    transaction_date = get_previous_weekday(transaction_date)

    # Use provided price if available, otherwise fetch from Yahoo Finance
    if provided_price is not None:
        price_per_share = float(provided_price)
        # Use provided currency or default to EUR (CM import is in EUR)
        price_currency = provided_currency or 'EUR'
    else:
        # Fetch historical price at transaction date (skip cache for fresh data)
        try:
            price_per_share = fetch_stock_price(stock_ticker, transaction_date, use_cache=False)
            # Convert numpy types to native Python types for PostgreSQL compatibility
            if price_per_share is not None:
                price_per_share = float(price_per_share)
            # Price from Yahoo is in stock's native currency
            price_currency = get_stock_currency(stock_ticker)
        except Exception as e:
            return jsonify({'error': f'Could not fetch price for {stock_ticker} on {transaction_date}: {str(e)}'}), 400

    try:
        with get_db() as conn:
            cursor = conn.execute('''
                INSERT INTO portfolio_transactions (user_id, account_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
            ''', (request.user_id, account_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency))
            transaction_id = cursor.fetchone()['id']
    except Exception as e:
        print(f"[Transaction Error] {stock_ticker}: {str(e)}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    # Refresh videos for this ticker in the background
    refresh_videos_for_ticker_async(stock_ticker)

    return jsonify({
        'success': True,
        'id': transaction_id,
        'account_id': account_id,
        'stock_ticker': stock_ticker,
        'transaction_type': transaction_type,
        'quantity': quantity,
        'transaction_date': transaction_date,
        'price_per_share': price_per_share,
        'price_currency': price_currency
    })


@app.route('/api/investing/transactions/bulk', methods=['POST'])
@login_required
def bulk_add_transactions():
    """Bulk add transactions (for initial setup). Fetches historical prices for each."""
    data = request.get_json()
    if not data or 'transactions' not in data:
        return jsonify({'error': 'Transactions array required'}), 400

    transactions = data['transactions']
    if not isinstance(transactions, list):
        return jsonify({'error': 'Transactions must be an array'}), 400

    processed = []
    errors = []

    for t in transactions:
        stock_ticker = t.get('stock_ticker', '').upper().strip()
        transaction_type = t.get('transaction_type', 'BUY').upper().strip()
        quantity = t.get('quantity', 0)
        transaction_date = t.get('transaction_date')

        if not stock_ticker or quantity <= 0 or not transaction_date:
            continue
        if transaction_type not in ['BUY', 'SELL']:
            transaction_type = 'BUY'

        try:
            adjusted_date = get_previous_weekday(transaction_date)
            price_per_share = fetch_stock_price(stock_ticker, adjusted_date, use_cache=False)
            # Convert numpy types to native Python types for PostgreSQL compatibility
            if price_per_share is not None:
                price_per_share = float(price_per_share)
            processed.append({
                'stock_ticker': stock_ticker,
                'transaction_type': transaction_type,
                'quantity': int(quantity),
                'transaction_date': adjusted_date,
                'price_per_share': price_per_share
            })
        except Exception as e:
            errors.append(f'{stock_ticker}: {str(e)}')

    if not processed:
        return jsonify({'error': 'No valid transactions to save', 'details': errors}), 400

    with get_db() as conn:
        # Get user info and count existing transactions before clearing
        cursor = conn.execute('SELECT name, email FROM users WHERE id = ?', (request.user_id,))
        user = cursor.fetchone()

        cursor = conn.execute('SELECT COUNT(*) as count FROM portfolio_transactions WHERE user_id = ?', (request.user_id,))
        old_count = cursor.fetchone()['count']

        # Clear existing transactions
        conn.execute('DELETE FROM portfolio_transactions WHERE user_id = ?', (request.user_id,))

        # Insert new transactions
        for t in processed:
            conn.execute('''
                INSERT INTO portfolio_transactions (user_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (request.user_id, t['stock_ticker'], t['transaction_type'], t['quantity'], t['transaction_date'], t['price_per_share']))

        # Send admin notification if there were existing transactions
        if user and old_count > 0:
            try:
                send_admin_deletion_alert(
                    user_name=user['name'],
                    user_email=user['email'],
                    deletion_type='bulk_replace',
                    details={
                        'Previous transactions': old_count,
                        'New transactions': len(processed),
                        'Tickers': ', '.join(set(t['stock_ticker'] for t in processed[:10])) + ('...' if len(processed) > 10 else ''),
                    }
                )
            except Exception as e:
                logger.error(f"Failed to send deletion alert: {e}")

    return jsonify({
        'success': True,
        'count': len(processed),
        'transactions': processed,
        'errors': errors if errors else None
    })


@app.route('/api/investing/transactions/cleanup-duplicates', methods=['POST'])
@login_required
def cleanup_duplicate_transactions():
    """Remove duplicate transactions, keeping only one of each unique transaction."""
    with get_db() as conn:
        # Find all duplicates (same ticker, type, quantity, date, account)
        # Keep the one with the lowest id
        cursor = conn.execute('''
            SELECT stock_ticker, transaction_type, quantity, transaction_date, account_id, COUNT(*) as cnt, MIN(id) as keep_id
            FROM portfolio_transactions
            WHERE user_id = ?
            GROUP BY stock_ticker, transaction_type, quantity, transaction_date, account_id
            HAVING COUNT(*) > 1
        ''', (request.user_id,))
        duplicates = cursor.fetchall()

        if not duplicates:
            return jsonify({'success': True, 'removed': 0, 'message': 'No duplicates found'})

        total_removed = 0
        for dup in duplicates:
            # Delete all duplicates except the one with the lowest id
            cursor = conn.execute('''
                DELETE FROM portfolio_transactions
                WHERE user_id = ? AND stock_ticker = ? AND transaction_type = ?
                AND quantity = ? AND transaction_date = ?
                AND (account_id = ? OR (account_id IS NULL AND ? IS NULL))
                AND id != ?
            ''', (request.user_id, dup['stock_ticker'], dup['transaction_type'],
                  dup['quantity'], dup['transaction_date'], dup['account_id'], dup['account_id'], dup['keep_id']))
            total_removed += cursor.rowcount

    return jsonify({
        'success': True,
        'removed': total_removed,
        'duplicate_groups': len(duplicates),
        'message': f'Removed {total_removed} duplicate transactions'
    })


def _parse_revolut_pdf_with_gemini(pdf_bytes):
    """Parse Revolut PDF using Gemini Vision for high accuracy. Returns (transactions, errors)."""
    import google.generativeai as genai
    from pdf2image import convert_from_bytes
    import json

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return None, ['GEMINI_API_KEY not configured']

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-3-flash-preview')

        # Convert PDF pages to images
        images = convert_from_bytes(pdf_bytes, dpi=150)

        prompt = """Extract ALL stock transactions from this Revolut account statement.
Be thorough - check every row carefully, including the LAST row of each table.

Return ONLY a valid JSON array with this exact format:
[{"ticker": "AAPL", "type": "BUY", "quantity": 10, "date": "2024-03-15"}, ...]

Rules:
- Include ONLY stock trades (Buy/Sell market orders), NOT dividends, fees, cash top-ups, or custody fees
- ticker: The stock symbol (e.g., "AAPL", "META", "GOOGL")
- type: "BUY" or "SELL" only
- quantity: Number of shares (integer)
- date: Format YYYY-MM-DD
- If there are no transactions, return an empty array: []

Return ONLY the JSON array, no other text."""

        # Send all pages to Gemini
        response = model.generate_content([prompt] + images)

        # Parse JSON response
        response_text = response.text.strip()
        # Remove markdown code blocks if present
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1]
            if response_text.endswith('```'):
                response_text = response_text.rsplit('```', 1)[0]
            response_text = response_text.strip()

        transactions_raw = json.loads(response_text)

        # Normalize to our format
        transactions = []
        for tx in transactions_raw:
            transactions.append({
                'stock_ticker': tx['ticker'].upper(),
                'transaction_type': tx['type'].upper(),
                'quantity': int(tx['quantity']),
                'transaction_date': tx['date'],
                'price_per_share': None,  # Will be fetched later
            })

        return transactions, []

    except Exception as e:
        return None, [f'Gemini parsing failed: {str(e)}']


def _parse_revolut_pdf_bytes(pdf_bytes):
    """Helper function to parse Revolut PDF bytes. Returns (transactions, errors).
    Uses Gemini Vision if available, falls back to pdfplumber."""

    # Try Gemini first for better accuracy
    transactions, errors = _parse_revolut_pdf_with_gemini(pdf_bytes)
    if transactions is not None:
        return transactions, errors

    # Fallback to pdfplumber
    print(f"[Revolut Import] Gemini unavailable ({errors}), falling back to pdfplumber")
    import pdfplumber
    import io
    import re

    pdf_file = io.BytesIO(pdf_bytes)
    transactions = []
    errors = []

    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()

            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Find header row
                header = None
                header_idx = 0
                for i, row in enumerate(table):
                    row_lower = [str(cell).lower() if cell else '' for cell in row]
                    # Look for Revolut trading statement headers
                    if any('date' in cell for cell in row_lower) and \
                       any('symbol' in cell or 'ticker' in cell or 'instrument' in cell for cell in row_lower):
                        header = row_lower
                        header_idx = i
                        break

                if not header:
                    continue

                # Map column indices
                date_col = next((i for i, h in enumerate(header) if 'date' in h and 'settle' not in h), None)
                ticker_col = next((i for i, h in enumerate(header) if 'symbol' in h or 'ticker' in h), None)
                instrument_col = next((i for i, h in enumerate(header) if 'instrument' in h or 'name' in h), None)
                type_col = next((i for i, h in enumerate(header) if h == 'type' or 'activity' in h), None)
                side_col = next((i for i, h in enumerate(header) if 'side' in h), None)
                qty_col = next((i for i, h in enumerate(header) if 'quantity' in h or 'qty' in h or 'shares' in h), None)
                price_col = next((i for i, h in enumerate(header) if 'price' in h and 'total' not in h), None)

                # Process data rows
                for row in table[header_idx + 1:]:
                    if not row or all(not cell for cell in row):
                        continue

                    try:
                        # Extract date
                        date_str = row[date_col] if date_col is not None and date_col < len(row) else None
                        if not date_str:
                            continue

                        # Parse date (Revolut uses various formats)
                        parsed_date = None
                        date_str = str(date_str).strip()

                        # Try formats with time first (e.g., "17 Nov 2022 19:00:16 GMT")
                        for fmt in ['%d %b %Y %H:%M:%S GMT', '%d %b %Y %H:%M:%S', '%d %B %Y %H:%M:%S GMT', '%d %B %Y %H:%M:%S']:
                            try:
                                parsed_date = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
                                break
                            except:
                                pass

                        # Try date-only formats
                        if not parsed_date:
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d %b %Y', '%d %B %Y']:
                                try:
                                    parsed_date = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
                                    break
                                except:
                                    pass

                        if not parsed_date:
                            # Try to extract date with regex
                            date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
                            if date_match:
                                parsed_date = date_match.group(0)
                            else:
                                # Try "DD Mon YYYY" pattern from longer string
                                date_match = re.search(r'(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})', date_str)
                                if date_match:
                                    try:
                                        parsed_date = datetime.strptime(f"{date_match.group(1)} {date_match.group(2)} {date_match.group(3)}", '%d %b %Y').strftime('%Y-%m-%d')
                                    except:
                                        pass

                        if not parsed_date:
                            continue

                        # Extract ticker
                        ticker = None
                        if ticker_col is not None and ticker_col < len(row) and row[ticker_col]:
                            ticker_cell = str(row[ticker_col]).strip().upper()
                            # Handle merged columns - extract first word if it looks like a ticker
                            # Ticker should be 1-5 uppercase letters
                            ticker_match = re.match(r'^([A-Z]{1,5})\b', ticker_cell)
                            if ticker_match:
                                ticker = ticker_match.group(1)
                            else:
                                # Also try to find ticker anywhere in the cell
                                ticker_search = re.search(r'\b([A-Z]{1,5})\b', ticker_cell)
                                if ticker_search:
                                    ticker = ticker_search.group(1)

                        if not ticker and instrument_col is not None and instrument_col < len(row) and row[instrument_col]:
                            # Try to extract ticker from instrument name
                            inst = str(row[instrument_col]).strip()
                            # Common pattern: "AAPL - Apple Inc" or "Apple Inc (AAPL)"
                            ticker_match = re.search(r'\b([A-Z]{1,5})\b', inst)
                            if ticker_match:
                                ticker = ticker_match.group(1)

                        if not ticker:
                            continue

                        # Skip non-trade rows (Cash top-up, Custody fee, Dividend, etc.)
                        if type_col is not None and type_col < len(row) and row[type_col]:
                            type_str = str(row[type_col]).strip().upper()
                            if 'TRADE' not in type_str and 'MARKET' not in type_str and 'LIMIT' not in type_str:
                                # Check if it's a non-trade transaction
                                if any(skip in type_str for skip in ['CASH', 'FEE', 'DIVIDEND', 'TRANSFER', 'TOP-UP', 'TOP UP', 'CUSTODY']):
                                    continue

                        # Extract transaction type (Buy/Sell) - prefer Side column
                        tx_type = 'BUY'
                        if side_col is not None and side_col < len(row) and row[side_col]:
                            side_str = str(row[side_col]).strip().upper()
                            if 'SELL' in side_str or 'SOLD' in side_str:
                                tx_type = 'SELL'
                            elif 'BUY' in side_str or 'BOUGHT' in side_str:
                                tx_type = 'BUY'
                        elif type_col is not None and type_col < len(row) and row[type_col]:
                            # Fallback to type column if no side column
                            type_str = str(row[type_col]).strip().upper()
                            if 'SELL' in type_str or 'SOLD' in type_str:
                                tx_type = 'SELL'

                        # Extract quantity
                        qty = None
                        if qty_col is not None and qty_col < len(row) and row[qty_col]:
                            qty_str = str(row[qty_col]).replace(',', '.').strip()
                            qty_match = re.search(r'[\d.]+', qty_str)
                            if qty_match:
                                qty = float(qty_match.group())

                        if not qty or qty <= 0:
                            continue

                        # Extract price (optional - will be fetched if not provided)
                        price = None
                        if price_col is not None and price_col < len(row) and row[price_col]:
                            price_str = str(row[price_col]).replace(',', '.').replace('$', '').replace('€', '').strip()
                            price_match = re.search(r'[\d.]+', price_str)
                            if price_match:
                                price = float(price_match.group())

                        transactions.append({
                            'stock_ticker': ticker,
                            'transaction_type': tx_type,
                            'quantity': int(qty) if qty == int(qty) else qty,
                            'transaction_date': parsed_date,
                            'price_per_share': price,
                        })

                    except Exception as e:
                        errors.append(f"Row parse error: {str(e)}")

    # If table extraction failed, try text-based extraction
    if not transactions:
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() or ""

            # Pattern: date followed by ticker, Trade - Market, quantity, price, Buy/Sell
            # Example: "17 Nov 2022 19:00:16 GMT META Trade - Market 2 US$112.12 Buy US$224.24"
            pattern = r'(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})\s+\d{2}:\d{2}:\d{2}\s+GMT\s+([A-Z]{1,5})\s+Trade\s*-\s*Market\s+(\d+(?:\.\d+)?)\s+US\$[\d.]+\s+(Buy|Sell)'

            for match in re.finditer(pattern, full_text):
                try:
                    date_part = match.group(1)  # "17 Nov 2022"
                    ticker = match.group(2)     # "META"
                    qty = float(match.group(3)) # "2"
                    side = match.group(4)       # "Buy" or "Sell"

                    # Parse date
                    parsed_date = datetime.strptime(date_part, '%d %b %Y').strftime('%Y-%m-%d')
                    tx_type = 'BUY' if side.upper() == 'BUY' else 'SELL'

                    transactions.append({
                        'stock_ticker': ticker,
                        'transaction_type': tx_type,
                        'quantity': int(qty) if qty == int(qty) else qty,
                        'transaction_date': parsed_date,
                        'price_per_share': None,
                    })
                except Exception as e:
                    errors.append(f"Text parse error: {str(e)}")

    return transactions, errors


def _cleanup_expired_tokens():
    """Remove tokens older than 5 minutes."""
    now = datetime.now()
    expired = [t for t, data in _upload_tokens.items()
               if now - data['created_at'] > timedelta(minutes=5)]
    for t in expired:
        del _upload_tokens[t]


@app.route('/api/investing/import/revolut', methods=['POST'])
@login_required
def parse_revolut_pdf():
    """Parse a Revolut trading statement PDF and return extracted transactions."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF'}), 400

    try:
        pdf_bytes = file.read()
        transactions, errors = _parse_revolut_pdf_bytes(pdf_bytes)

        return jsonify({
            'success': True,
            'transactions': transactions,
            'count': len(transactions),
            'warnings': errors if errors else None
        })

    except Exception as e:
        return jsonify({'error': f'Failed to parse PDF: {str(e)}'}), 400


def _map_stock_names_to_tickers_with_gemini(stock_names: list[str]) -> dict[str, str]:
    """Use Gemini to map French stock names to tickers."""
    import google.generativeai as genai
    import json

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        # Fallback: return names as-is
        return {name: name for name in stock_names}

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')

        prompt = f"""Map these French stock names to their stock tickers. Return ONLY a JSON object mapping each name to its ticker.

Stock names:
{json.dumps(stock_names, ensure_ascii=False)}

Common mappings:
- "META PLATFORMS CLA" or "META PLATFORMS" → "META"
- "ALPHABET CL.A" or "ALPHABET" → "GOOGL"
- "NVIDIA" → "NVDA"
- "MICROSOFT" → "MSFT"
- "VISA CL.A" or "VISA" → "V"
- "LVMH MOET HENNESSY VUITTON" → "MC.PA"
- "NU HOLDINGS LIMITED" → "NU"
- "AMAZON COM" or "AMAZON" → "AMZN"
- "APPLE" → "AAPL"
- "TESLA" → "TSLA"
- "NETFLIX" → "NFLX"

Return JSON like: {{"META PLATFORMS CLA": "META", "NVIDIA": "NVDA"}}
Only return the JSON object, no other text."""

        response = model.generate_content(prompt)
        text = response.text.strip()

        # Extract JSON from response
        if text.startswith('```'):
            text = text.split('```')[1]
            if text.startswith('json'):
                text = text[4:]
            text = text.strip()

        mapping = json.loads(text)
        # Filter out null values - use original name as fallback
        return {name: (ticker if ticker else name) for name, ticker in mapping.items()}

    except Exception as e:
        print(f"[Gemini mapping error] {e}")
        # Fallback: return names as-is
        return {name: name for name in stock_names}


def _get_split_adjustments(tickers: list[str], transactions: list[dict]) -> dict:
    """
    Get cumulative split adjustment factors for each transaction.

    For each transaction, calculates how many shares one original share
    has become due to splits that occurred AFTER the transaction date.

    Example: NVDA 10-for-1 split in June 2024
    - Transaction from May 2024: factor = 10 (20 shares → 200 shares)
    - Transaction from July 2024: factor = 1 (no adjustment needed)

    Returns: dict mapping (ticker, date) -> { 'quantity_factor': float, 'price_factor': float }
    """
    import yfinance as yf
    from datetime import datetime

    adjustments = {}

    # Group transactions by ticker to minimize API calls
    ticker_dates = {}
    for tx in transactions:
        ticker = tx.get('stock_ticker')
        date_str = tx.get('transaction_date')
        if ticker and date_str:
            if ticker not in ticker_dates:
                ticker_dates[ticker] = set()
            ticker_dates[ticker].add(date_str)

    # Fetch split data for each ticker
    for ticker in ticker_dates:
        try:
            stock = yf.Ticker(ticker)
            splits = stock.splits

            if splits is None or splits.empty:
                # No splits for this ticker
                for date_str in ticker_dates[ticker]:
                    adjustments[(ticker, date_str)] = {'quantity_factor': 1.0, 'price_factor': 1.0}
                continue

            # For each transaction date, calculate cumulative split factor
            for date_str in ticker_dates[ticker]:
                tx_date = datetime.strptime(date_str, '%Y-%m-%d')

                # Get splits that occurred AFTER the transaction date
                cumulative_factor = 1.0
                for split_date, split_ratio in splits.items():
                    # split_date is a pandas Timestamp
                    if split_date.to_pydatetime().replace(tzinfo=None) > tx_date:
                        cumulative_factor *= split_ratio

                adjustments[(ticker, date_str)] = {
                    'quantity_factor': cumulative_factor,
                    'price_factor': 1.0 / cumulative_factor if cumulative_factor != 0 else 1.0
                }

                if cumulative_factor != 1.0:
                    print(f"[Split adjustment] {ticker} on {date_str}: factor={cumulative_factor}")

        except Exception as e:
            print(f"[Split adjustment error] {ticker}: {e}")
            # Default to no adjustment on error
            for date_str in ticker_dates[ticker]:
                adjustments[(ticker, date_str)] = {'quantity_factor': 1.0, 'price_factor': 1.0}

    return adjustments


def _parse_credit_mutuel_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse Crédit Mutuel Excel export and return transactions."""
    import openpyxl
    from io import BytesIO
    from datetime import datetime

    transactions = []
    errors = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Find header row
        header_row = None
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
            row_values = [cell.value for cell in row]
            if 'Exécution' in row_values or 'Opération' in row_values:
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if cell.value:
                        headers[cell.value] = col_idx
                break

        if not header_row:
            return [], ['Could not find header row in Excel file']

        # Extract data rows
        raw_transactions = []
        stock_names_set = set()

        for row in ws.iter_rows(min_row=header_row + 1):
            row_values = [cell.value for cell in row]

            # Get operation type
            op_col = headers.get('Opération', headers.get('Operation', -1))
            operation = row_values[op_col] if op_col >= 0 and op_col < len(row_values) else None

            if not operation:
                continue

            # Only process Achat (buy) and Vente (sell)
            operation_lower = str(operation).lower()
            if 'achat' in operation_lower:
                tx_type = 'BUY'
            elif 'vente' in operation_lower:
                tx_type = 'SELL'
            else:
                # Skip dividends and other operations
                continue

            # Get date
            date_col = headers.get('Exécution', headers.get('Execution', -1))
            date_val = row_values[date_col] if date_col >= 0 and date_col < len(row_values) else None

            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif date_val:
                # Try to parse DD/MM/YYYY format
                try:
                    date_str = datetime.strptime(str(date_val), '%d/%m/%Y').strftime('%Y-%m-%d')
                except:
                    try:
                        date_str = datetime.strptime(str(date_val), '%Y-%m-%d').strftime('%Y-%m-%d')
                    except:
                        errors.append(f"Could not parse date: {date_val}")
                        continue
            else:
                continue

            # Get quantity
            qty_col = headers.get('Quantité / Montant nominal', headers.get('Quantité', -1))
            quantity = row_values[qty_col] if qty_col >= 0 and qty_col < len(row_values) else None

            if quantity is None:
                continue

            try:
                quantity = int(float(str(quantity).replace(',', '.').replace(' ', '')))
            except:
                errors.append(f"Could not parse quantity: {quantity}")
                continue

            # Get stock name
            stock_col = headers.get('Valeur', -1)
            stock_name = row_values[stock_col] if stock_col >= 0 and stock_col < len(row_values) else None

            if not stock_name:
                continue

            stock_name = str(stock_name).strip()

            # Skip OAT (French government bonds) - pattern like "OAT 8,50%92-25042023"
            if re.match(r'^OAT.*%', stock_name, re.IGNORECASE):
                continue

            stock_names_set.add(stock_name)

            # Get net amount (Montant net) to calculate price per share
            montant_col = headers.get('Montant net', headers.get('Montant Net', -1))
            montant_net = row_values[montant_col] if montant_col >= 0 and montant_col < len(row_values) else None

            price_per_share = None
            if montant_net is not None and quantity > 0:
                try:
                    montant_value = float(str(montant_net).replace(',', '.').replace(' ', '').replace('€', ''))
                    # Montant net is negative for buys, positive for sells - take absolute value
                    price_per_share = round(abs(montant_value) / quantity, 4)
                except:
                    pass  # Will be None, backend will fetch from Yahoo

            raw_transactions.append({
                'date': date_str,
                'type': tx_type,
                'quantity': quantity,
                'stock_name': stock_name,
                'price_per_share': price_per_share
            })

        # Map stock names to tickers using Gemini
        if stock_names_set:
            name_to_ticker = _map_stock_names_to_tickers_with_gemini(list(stock_names_set))
        else:
            name_to_ticker = {}

        # Build final transactions (before split adjustment)
        for raw_tx in raw_transactions:
            ticker = name_to_ticker.get(raw_tx['stock_name']) or raw_tx['stock_name']
            transactions.append({
                'stock_ticker': ticker,
                'transaction_type': raw_tx['type'],
                'quantity': raw_tx['quantity'],
                'transaction_date': raw_tx['date'],
                'price_per_share': raw_tx['price_per_share']  # Calculated from Montant net
            })

        # Apply stock split adjustments
        if transactions:
            tickers = list(set(tx['stock_ticker'] for tx in transactions))
            split_adjustments = _get_split_adjustments(tickers, transactions)

            for tx in transactions:
                key = (tx['stock_ticker'], tx['transaction_date'])
                adj = split_adjustments.get(key, {'quantity_factor': 1.0, 'price_factor': 1.0})

                if adj['quantity_factor'] != 1.0:
                    original_qty = tx['quantity']
                    tx['quantity'] = int(round(tx['quantity'] * adj['quantity_factor']))
                    if tx['price_per_share'] is not None:
                        tx['price_per_share'] = round(tx['price_per_share'] * adj['price_factor'], 4)
                    print(f"[Split applied] {tx['stock_ticker']}: {original_qty} -> {tx['quantity']} shares")

        return transactions, errors

    except Exception as e:
        return [], [f'Error parsing Excel: {str(e)}']


@app.route('/api/investing/import/credit-mutuel', methods=['POST'])
@login_required
def parse_credit_mutuel_excel():
    """Parse a Crédit Mutuel Excel export and return extracted transactions."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'File must be an Excel file (.xlsx)'}), 400

    try:
        file_bytes = file.read()
        transactions, errors = _parse_credit_mutuel_excel(file_bytes)

        return jsonify({
            'success': True,
            'transactions': transactions,
            'count': len(transactions),
            'warnings': errors if errors else None
        })

    except Exception as e:
        return jsonify({'error': f'Failed to parse Excel: {str(e)}'}), 400


def _parse_ibkr_pdf_with_gemini(pdf_bytes: bytes) -> tuple[list[dict] | None, list[str]]:
    """Parse Interactive Brokers Activity Statement PDF using Gemini Vision."""
    import google.generativeai as genai
    from pdf2image import convert_from_bytes
    import json

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return None, ['GEMINI_API_KEY not configured']

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')

        # Convert PDF pages to images
        images = convert_from_bytes(pdf_bytes, dpi=150)

        prompt = """Extract ALL stock transactions from this Interactive Brokers Activity Statement.
Look for the "Trades" section which contains stock buy/sell transactions.

Return ONLY a valid JSON array with this exact format:
[{"ticker": "AAPL", "type": "BUY", "quantity": 10.25, "date": "2024-03-15", "price": 150.25}, ...]

Rules:
- Include ONLY stock trades (Buy/Sell), NOT dividends, fees, interest, forex, or options
- ticker: The stock symbol (e.g., "AAPL", "META", "GOOGL")
- type: "BUY" or "SELL" only. Proceeds indicate the impact on the cash balance. If negative, the transaction was a Buy.
- quantity: Number of shares (always positive). Can be fractional, keep two decimals precision.
- date: Format YYYY-MM-DD
- price: Price per share (number, can be decimal). "T. Price" is the transaction price.
- If there are no stock transactions, return an empty array: []

Return ONLY the JSON array, no other text."""

        # Send all pages to Gemini
        response = model.generate_content([prompt] + images)

        # Parse JSON response
        response_text = response.text.strip()
        # Remove markdown code blocks if present
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1]
            if response_text.endswith('```'):
                response_text = response_text.rsplit('```', 1)[0]
            response_text = response_text.strip()

        transactions_raw = json.loads(response_text)

        # Normalize to our format
        transactions = []
        for tx in transactions_raw:
            price = tx.get('price')
            transactions.append({
                'stock_ticker': tx['ticker'].upper(),
                'transaction_type': tx['type'].upper(),
                'quantity': round(abs(float(tx['quantity'])), 2),  # Ensure positive, keep 2 decimals
                'transaction_date': tx['date'],
                'price_per_share': float(price) if price is not None else None,
            })

        # Apply stock split adjustments
        if transactions:
            tickers = list(set(tx['stock_ticker'] for tx in transactions))
            split_adjustments = _get_split_adjustments(tickers, transactions)

            for tx in transactions:
                key = (tx['stock_ticker'], tx['transaction_date'])
                adj = split_adjustments.get(key, {'quantity_factor': 1.0, 'price_factor': 1.0})

                if adj['quantity_factor'] != 1.0:
                    tx['quantity'] = round(tx['quantity'] * adj['quantity_factor'], 2)
                    if tx['price_per_share'] is not None:
                        tx['price_per_share'] = round(tx['price_per_share'] * adj['price_factor'], 4)

        return transactions, []

    except Exception as e:
        return None, [f'Gemini parsing failed: {str(e)}']


def _parse_ibkr_html_with_gemini(html_bytes: bytes) -> tuple[list[dict] | None, list[str]]:
    """Parse Interactive Brokers Activity Statement HTML using Gemini."""
    import google.generativeai as genai
    import json

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return None, ['GEMINI_API_KEY not configured']

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')

        # Decode HTML content
        html_content = html_bytes.decode('utf-8', errors='ignore')

        prompt = f"""Extract ALL stock transactions from this Interactive Brokers Activity Statement HTML.
Look for the "Trades" section which contains stock buy/sell transactions.

Return ONLY a valid JSON array with this exact format:
[{{"ticker": "AAPL", "type": "BUY", "quantity": 10, "date": "2024-03-15", "price": 150.25}}, ...]

Rules:
- Include ONLY stock trades (Buy/Sell), NOT dividends, fees, interest, forex, or options
- ticker: The stock symbol (e.g., "AAPL", "META", "GOOGL")
- type: "BUY" or "SELL" only
- quantity: Number of shares (integer, always positive)
- date: Format YYYY-MM-DD
- price: Price per share (number, can be decimal)
- If there are no stock transactions, return an empty array: []

Return ONLY the JSON array, no other text.

HTML Content:
{html_content[:100000]}"""  # Limit to ~100k chars

        response = model.generate_content(prompt)

        # Parse JSON response
        response_text = response.text.strip()
        # Remove markdown code blocks if present
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1]
            if response_text.endswith('```'):
                response_text = response_text.rsplit('```', 1)[0]
            response_text = response_text.strip()

        transactions_raw = json.loads(response_text)

        # Normalize to our format
        transactions = []
        for tx in transactions_raw:
            price = tx.get('price')
            transactions.append({
                'stock_ticker': tx['ticker'].upper(),
                'transaction_type': tx['type'].upper(),
                'quantity': round(abs(float(tx['quantity'])), 2),  # Ensure positive, keep 2 decimals
                'transaction_date': tx['date'],
                'price_per_share': float(price) if price is not None else None,
            })

        # Apply stock split adjustments
        if transactions:
            tickers = list(set(tx['stock_ticker'] for tx in transactions))
            split_adjustments = _get_split_adjustments(tickers, transactions)

            for tx in transactions:
                key = (tx['stock_ticker'], tx['transaction_date'])
                adj = split_adjustments.get(key, {'quantity_factor': 1.0, 'price_factor': 1.0})

                if adj['quantity_factor'] != 1.0:
                    tx['quantity'] = round(tx['quantity'] * adj['quantity_factor'], 2)
                    if tx['price_per_share'] is not None:
                        tx['price_per_share'] = round(tx['price_per_share'] * adj['price_factor'], 4)

        return transactions, []

    except Exception as e:
        logger.error(f"[IBKR HTML Import Error] {type(e).__name__}: {str(e)}")
        return None, [f'Gemini parsing failed: {str(e)}']


def _parse_ibkr_image_with_gemini(image_bytes: bytes) -> tuple[list[dict] | None, list[str]]:
    """Parse Interactive Brokers Activity Statement image using Gemini Vision."""
    import google.generativeai as genai
    from PIL import Image
    from io import BytesIO
    import json

    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return None, ['GEMINI_API_KEY not configured']

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')

        # Load and convert image to RGB (in case of RGBA/PNG with transparency)
        image = Image.open(BytesIO(image_bytes))
        if image.mode in ('RGBA', 'LA', 'P'):
            image = image.convert('RGB')

        prompt = """Extract ALL stock transactions from this Interactive Brokers Activity Statement image.
Look for the "Trades" section which contains stock buy/sell transactions.

Return ONLY a valid JSON array with this exact format:
[{"ticker": "AAPL", "type": "BUY", "quantity": 10, "date": "2024-03-15", "price": 150.25}, ...]

Rules:
- Include ONLY stock trades (Buy/Sell), NOT dividends, fees, interest, forex, or options
- ticker: The stock symbol (e.g., "AAPL", "META", "GOOGL")
- type: "BUY" or "SELL" only
- quantity: Number of shares (integer, always positive)
- date: Format YYYY-MM-DD
- price: Price per share (number, can be decimal)
- If there are no stock transactions, return an empty array: []

Return ONLY the JSON array, no other text."""

        response = model.generate_content([prompt, image])

        # Check if response is valid
        if not response or not response.text:
            return None, ['Gemini returned empty response']

        # Parse JSON response
        response_text = response.text.strip()
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1]
            if response_text.endswith('```'):
                response_text = response_text.rsplit('```', 1)[0]
            response_text = response_text.strip()

        transactions_raw = json.loads(response_text)

        # Normalize to our format
        transactions = []
        for tx in transactions_raw:
            price = tx.get('price')
            transactions.append({
                'stock_ticker': tx['ticker'].upper(),
                'transaction_type': tx['type'].upper(),
                'quantity': round(abs(float(tx['quantity'])), 2),  # Ensure positive, keep 2 decimals
                'transaction_date': tx['date'],
                'price_per_share': float(price) if price is not None else None,
            })

        # Apply stock split adjustments
        if transactions:
            tickers = list(set(tx['stock_ticker'] for tx in transactions))
            split_adjustments = _get_split_adjustments(tickers, transactions)

            for tx in transactions:
                key = (tx['stock_ticker'], tx['transaction_date'])
                adj = split_adjustments.get(key, {'quantity_factor': 1.0, 'price_factor': 1.0})

                if adj['quantity_factor'] != 1.0:
                    tx['quantity'] = round(tx['quantity'] * adj['quantity_factor'], 2)
                    if tx['price_per_share'] is not None:
                        tx['price_per_share'] = round(tx['price_per_share'] * adj['price_factor'], 4)

        return transactions, []

    except Exception as e:
        logger.error(f"[IBKR Image Import Error] {type(e).__name__}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None, [f'Gemini parsing failed: {str(e)}']


@app.route('/api/investing/import/interactive-brokers', methods=['POST'])
@login_required
def parse_ibkr_file():
    """Parse an Interactive Brokers Activity Statement PDF and return extracted transactions."""
    logger.info("[IBKR Import] Starting file parsing...")

    if 'file' not in request.files:
        logger.info("[IBKR Import] No file in request")
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    filename = file.filename.lower()
    logger.info(f"[IBKR Import] Filename: {filename}")

    if not filename.endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF'}), 400

    try:
        file_bytes = file.read()
        logger.info(f"[IBKR Import] File size: {len(file_bytes)} bytes")

        logger.info("[IBKR Import] Parsing as PDF...")
        transactions, errors = _parse_ibkr_pdf_with_gemini(file_bytes)

        logger.info(f"[IBKR Import] Result: transactions={transactions is not None}, errors={errors}")

        if transactions is None:
            logger.error(f"[IBKR Import] Failed: {errors}")
            return jsonify({'error': errors[0] if errors else 'Failed to parse file'}), 400

        return jsonify({
            'success': True,
            'transactions': transactions,
            'count': len(transactions),
            'warnings': errors if errors else None
        })

    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 400


@app.route('/api/investing/import/create-token', methods=['POST'])
@login_required
def create_upload_token():
    """Create a temporary token for mobile PDF upload."""
    _cleanup_expired_tokens()

    # Get user email for PostHog tracking on mobile
    with get_db() as conn:
        user = conn.execute('SELECT email FROM users WHERE id = ?', (request.user_id,)).fetchone()
        user_email = user['email'] if user else None

    token = secrets.token_urlsafe(32)
    _upload_tokens[token] = {
        'user_id': request.user_id,
        'user_email': user_email,
        'created_at': datetime.now(),
        'transactions': None,
        'status': 'pending'  # pending, uploaded, error
    }

    return jsonify({'token': token})


@app.route('/api/investing/import/token-info/<token>', methods=['GET'])
def get_token_info(token):
    """Get user info for a token (for PostHog identification on mobile upload page)."""
    _cleanup_expired_tokens()

    if token not in _upload_tokens:
        return jsonify({'error': 'Invalid or expired token'}), 400

    token_data = _upload_tokens[token]
    return jsonify({'email': token_data.get('user_email')})


@app.route('/api/investing/import/upload/<token>', methods=['POST'])
def upload_with_token(token):
    """Upload PDF using a token (no auth required - token is the auth)."""
    _cleanup_expired_tokens()

    if token not in _upload_tokens:
        return jsonify({'error': 'Invalid or expired token'}), 400

    token_data = _upload_tokens[token]

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF'}), 400

    try:
        pdf_bytes = file.read()
        transactions, errors = _parse_revolut_pdf_bytes(pdf_bytes)

        token_data['transactions'] = transactions
        token_data['warnings'] = errors if errors else None
        token_data['status'] = 'uploaded'
        token_data['count'] = len(transactions)

        return jsonify({
            'success': True,
            'count': len(transactions)
        })

    except Exception as e:
        token_data['status'] = 'error'
        token_data['error'] = str(e)
        return jsonify({'error': f'Failed to parse PDF: {str(e)}'}), 400


@app.route('/api/investing/import/status/<token>', methods=['GET'])
@login_required
def check_upload_status(token):
    """Check the status of a mobile upload."""
    _cleanup_expired_tokens()

    if token not in _upload_tokens:
        return jsonify({'error': 'Invalid or expired token'}), 400

    token_data = _upload_tokens[token]

    # Verify the token belongs to this user
    if token_data['user_id'] != request.user_id:
        return jsonify({'error': 'Invalid token'}), 403

    if token_data['status'] == 'pending':
        return jsonify({'status': 'pending'})
    elif token_data['status'] == 'error':
        return jsonify({'status': 'error', 'error': token_data.get('error')})
    else:
        return jsonify({
            'status': 'uploaded',
            'transactions': token_data['transactions'],
            'count': token_data['count'],
            'warnings': token_data.get('warnings')
        })


@app.route('/api/investing/transactions/<int:transaction_id>', methods=['DELETE'])
@login_required
def delete_transaction(transaction_id):
    """Delete a transaction by ID."""
    with get_db() as conn:
        # Get transaction and user info before deletion
        cursor = conn.execute('''
            SELECT pt.stock_ticker, pt.quantity, pt.transaction_type, pt.transaction_date,
                   u.name, u.email
            FROM portfolio_transactions pt
            JOIN users u ON pt.user_id = u.id
            WHERE pt.user_id = ? AND pt.id = ?
        ''', (request.user_id, transaction_id))
        row = cursor.fetchone()

        if row:
            ticker, qty, tx_type, tx_date, user_name, user_email = row['stock_ticker'], row['quantity'], row['transaction_type'], row['transaction_date'], row['name'], row['email']

            conn.execute(
                'DELETE FROM portfolio_transactions WHERE user_id = ? AND id = ?',
                (request.user_id, transaction_id)
            )


    return jsonify({'success': True, 'id': transaction_id})


@app.route('/api/investing/transactions/<int:transaction_id>', methods=['PUT'])
@login_required
def update_transaction(transaction_id):
    """Update a transaction by ID."""
    from investing_utils import fetch_stock_price, get_stock_currency
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    with get_db() as conn:
        # Verify transaction belongs to user
        cursor = conn.execute(
            'SELECT id, stock_ticker, transaction_date, price_per_share, price_currency FROM portfolio_transactions WHERE user_id = ? AND id = ?',
            (request.user_id, transaction_id)
        )
        existing = cursor.fetchone()
        if not existing:
            return jsonify({'error': 'Transaction not found'}), 404

        # Extract update fields
        new_ticker = data.get('stock_ticker', existing['stock_ticker']).upper().strip()
        new_type = data.get('transaction_type')
        new_quantity = data.get('quantity')
        new_date = data.get('transaction_date', existing['transaction_date'])

        # Validate transaction type
        if new_type and new_type not in ('BUY', 'SELL'):
            return jsonify({'error': 'Invalid transaction type'}), 400

        # Validate quantity
        if new_quantity is not None:
            try:
                new_quantity = float(new_quantity)
                if new_quantity <= 0:
                    return jsonify({'error': 'Quantity must be positive'}), 400
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid quantity'}), 400

        # Re-fetch price if ticker or date changed
        price = existing['price_per_share']
        currency = existing['price_currency']
        if new_ticker != existing['stock_ticker'] or new_date != existing['transaction_date']:
            try:
                new_price = fetch_stock_price(new_ticker, new_date)
                if new_price:
                    price = new_price
                    currency = get_stock_currency(new_ticker) or 'USD'
            except Exception:
                pass  # Keep existing price if fetch fails

        # Build update query
        updates = []
        params = []

        updates.append('stock_ticker = ?')
        params.append(new_ticker)

        if new_type:
            updates.append('transaction_type = ?')
            params.append(new_type)

        if new_quantity is not None:
            updates.append('quantity = ?')
            params.append(new_quantity)

        updates.append('transaction_date = ?')
        params.append(new_date)

        updates.append('price_per_share = ?')
        params.append(price)

        updates.append('price_currency = ?')
        params.append(currency)

        params.extend([request.user_id, transaction_id])

        conn.execute(f'''
            UPDATE portfolio_transactions
            SET {', '.join(updates)}
            WHERE user_id = ? AND id = ?
        ''', params)

    return jsonify({
        'success': True,
        'id': transaction_id,
        'price_per_share': price,
        'price_currency': currency
    })


@app.route('/api/investing/fx-rates', methods=['POST'])
@login_required
def get_fx_rates():
    """Get historical EUR/USD FX rates for a list of dates."""
    from investing_utils import fetch_eurusd_rate
    data = request.get_json()
    if not data or 'dates' not in data:
        return jsonify({'error': 'dates array required'}), 400

    dates = data['dates']
    rates = {}
    for date_str in dates:
        try:
            rates[date_str] = fetch_eurusd_rate(date_str)
        except Exception:
            rates[date_str] = 1.0  # Fallback
    return jsonify({'rates': rates})


@app.route('/api/investing/holdings', methods=['GET'])
@login_required
def get_holdings():
    """Get computed current holdings from transaction history."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
               FROM portfolio_transactions WHERE user_id = ?
               ORDER BY transaction_date ASC, id ASC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    # Compute holdings using FIFO for cost basis
    holdings_map = {}  # ticker -> { quantity, lots: [{qty, price, date}] }

    for row in rows:
        ticker = row['stock_ticker']
        qty = row['quantity']
        price = row['price_per_share']
        date = row['transaction_date']
        tx_type = row['transaction_type']

        if ticker not in holdings_map:
            holdings_map[ticker] = {'quantity': 0, 'lots': []}

        if tx_type == 'BUY':
            holdings_map[ticker]['quantity'] += qty
            holdings_map[ticker]['lots'].append({'qty': qty, 'price': price, 'date': date})
        elif tx_type == 'SELL':
            holdings_map[ticker]['quantity'] -= qty
            # FIFO: remove from oldest lots first
            remaining_sell = qty
            while remaining_sell > 0 and holdings_map[ticker]['lots']:
                lot = holdings_map[ticker]['lots'][0]
                if lot['qty'] <= remaining_sell:
                    remaining_sell -= lot['qty']
                    holdings_map[ticker]['lots'].pop(0)
                else:
                    lot['qty'] -= remaining_sell
                    remaining_sell = 0

    # Build response with cost basis
    holdings = []
    for ticker, data in holdings_map.items():
        if data['quantity'] > 0:
            # Calculate weighted average cost basis from remaining lots
            total_cost = sum(lot['qty'] * lot['price'] for lot in data['lots'])
            total_qty = sum(lot['qty'] for lot in data['lots'])
            avg_cost = total_cost / total_qty if total_qty > 0 else 0

            holdings.append({
                'stock_ticker': ticker,
                'quantity': data['quantity'],
                'cost_basis': round(avg_cost, 2)
            })

    return jsonify({'holdings': holdings})


def compute_holdings_from_transactions(user_id, account_ids=None):
    """Helper to compute current holdings from transactions using FIFO.
    Tracks both USD and EUR cost basis (EUR uses historical rates at transaction time).
    Optionally filters by account_ids (list of account IDs).
    Excludes orphan transactions (null account_id) when no specific accounts are specified.
    """
    from investing_utils import fetch_eurusd_rate

    with get_db() as conn:
        if account_ids and len(account_ids) > 0:
            placeholders = ','.join('?' for _ in account_ids)
            cursor = conn.execute(
                f'''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ? AND account_id IN ({placeholders})
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id, *account_ids)
            )
        else:
            # Exclude orphan transactions (null account_id) - only include transactions with valid accounts
            cursor = conn.execute(
                '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ? AND account_id IS NOT NULL
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id,)
            )
        rows = cursor.fetchall()

    holdings_map = {}

    for row in rows:
        ticker = row['stock_ticker']
        qty = row['quantity']
        price = row['price_per_share']
        date = row['transaction_date']
        tx_type = row['transaction_type']
        price_currency = row['price_currency'] or 'USD'

        if ticker not in holdings_map:
            holdings_map[ticker] = {'quantity': 0, 'lots': []}

        if tx_type == 'BUY':
            # Calculate cost in EUR based on transaction currency
            if price_currency == 'EUR':
                cost_eur = qty * price
                # Fetch EUR/USD rate to get USD equivalent
                try:
                    eurusd_at_tx = fetch_eurusd_rate(date)
                except:
                    eurusd_at_tx = 1.0
                cost_usd = cost_eur * eurusd_at_tx
            else:
                # Price is in USD (or other currency treated as USD)
                cost_usd = qty * price
                try:
                    eurusd_at_tx = fetch_eurusd_rate(date)
                except:
                    eurusd_at_tx = 1.0
                cost_eur = cost_usd / eurusd_at_tx

            holdings_map[ticker]['quantity'] += qty
            holdings_map[ticker]['lots'].append({
                'qty': qty,
                'price': price,
                'date': date,
                'cost_eur': cost_eur
            })
        elif tx_type == 'SELL':
            holdings_map[ticker]['quantity'] -= qty
            remaining_sell = qty
            while remaining_sell > 0 and holdings_map[ticker]['lots']:
                lot = holdings_map[ticker]['lots'][0]
                if lot['qty'] <= remaining_sell:
                    remaining_sell -= lot['qty']
                    holdings_map[ticker]['lots'].pop(0)
                else:
                    # Reduce lot proportionally (including EUR cost)
                    sell_fraction = remaining_sell / lot['qty']
                    lot['cost_eur'] *= (1 - sell_fraction)
                    lot['qty'] -= remaining_sell
                    remaining_sell = 0

    holdings = []
    for ticker, data in holdings_map.items():
        if data['quantity'] > 0:
            total_cost_usd = sum(lot['qty'] * lot['price'] for lot in data['lots'])
            total_cost_eur = sum(lot['cost_eur'] for lot in data['lots'])
            total_qty = sum(lot['qty'] for lot in data['lots'])
            avg_cost = total_cost_usd / total_qty if total_qty > 0 else 0

            holdings.append({
                'stock_ticker': ticker,
                'quantity': data['quantity'],
                'cost_basis': round(avg_cost, 2),
                'total_cost': round(total_cost_usd, 2),
                'total_cost_eur': round(total_cost_eur, 2),
                'first_buy_date': data['lots'][0]['date'] if data['lots'] else None
            })

    return holdings


def compute_realized_gains(user_id, account_ids=None):
    """Calculate realized gains using FIFO with historical EUR rates.
    Returns both USD and EUR realized gains.
    Optionally filters by account_ids (list of account IDs).
    """
    from investing_utils import fetch_eurusd_rate

    with get_db() as conn:
        if account_ids and len(account_ids) > 0:
            placeholders = ','.join('?' for _ in account_ids)
            cursor = conn.execute(
                f'''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ? AND account_id IN ({placeholders})
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id, *account_ids)
            )
        else:
            cursor = conn.execute(
                '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ?
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id,)
            )
        rows = cursor.fetchall()

    # Track inventory per ticker: array of { qty, cost_usd, cost_eur }
    inventory = {}
    total_realized_gain_usd = 0
    total_realized_gain_eur = 0
    total_sold_cost_basis_eur = 0
    sell_count = 0

    for row in rows:
        ticker = row['stock_ticker']
        qty = row['quantity']
        price = row['price_per_share']
        date = row['transaction_date']
        tx_type = row['transaction_type']
        price_currency = row['price_currency'] or 'USD'

        if ticker not in inventory:
            inventory[ticker] = []

        if tx_type == 'BUY':
            # Get historical EUR/USD rate
            try:
                eurusd_at_tx = fetch_eurusd_rate(date)
            except:
                eurusd_at_tx = 1.0

            # Calculate cost based on transaction currency
            if price_currency == 'EUR':
                cost_eur = qty * price
                cost_usd = cost_eur * eurusd_at_tx
                cost_usd_per_share = price * eurusd_at_tx
            else:
                cost_usd = qty * price
                cost_eur = cost_usd / eurusd_at_tx
                cost_usd_per_share = price

            inventory[ticker].append({
                'qty': qty,
                'cost_usd_per_share': cost_usd_per_share,
                'cost_eur': cost_eur
            })
        else:  # SELL
            # Get EUR/USD rate at sell time
            try:
                eurusd_at_sell = fetch_eurusd_rate(date)
            except:
                eurusd_at_sell = 1.0

            remaining_sell = qty
            # Convert sale price to USD if needed
            if price_currency == 'EUR':
                sale_price_usd = price * eurusd_at_sell
                sale_proceeds_eur = qty * price
            else:
                sale_price_usd = price
                sale_proceeds_eur = (qty * sale_price_usd) / eurusd_at_sell
            sell_count += 1

            # FIFO: consume oldest lots first
            cost_basis_usd = 0
            cost_basis_eur = 0
            while remaining_sell > 0 and inventory[ticker]:
                lot = inventory[ticker][0]
                sell_from_lot = min(remaining_sell, lot['qty'])

                # Calculate cost basis for this portion
                portion_cost_usd = sell_from_lot * lot['cost_usd_per_share']
                portion_cost_eur = (sell_from_lot / lot['qty']) * lot['cost_eur']

                cost_basis_usd += portion_cost_usd
                cost_basis_eur += portion_cost_eur

                lot['qty'] -= sell_from_lot
                lot['cost_eur'] -= portion_cost_eur
                remaining_sell -= sell_from_lot

                if lot['qty'] <= 0:
                    inventory[ticker].pop(0)

            # Calculate gains
            gain_usd = (qty * sale_price_usd) - cost_basis_usd
            gain_eur = sale_proceeds_eur - cost_basis_eur

            total_realized_gain_usd += gain_usd
            total_realized_gain_eur += gain_eur
            total_sold_cost_basis_eur += cost_basis_eur

    return {
        'total_usd': round(total_realized_gain_usd, 2),
        'total_eur': round(total_realized_gain_eur, 2),
        'sold_cost_basis_eur': round(total_sold_cost_basis_eur, 2),
        'count': sell_count
    }


@app.route('/api/investing/portfolio/composition', methods=['GET'])
@login_required
def get_portfolio_composition():
    """Get portfolio composition with current values, weights, cost basis and gains."""
    # Support both single account_id and comma-separated account_ids
    account_ids_str = request.args.get('account_ids')
    if account_ids_str:
        account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()]
    else:
        account_id = request.args.get('account_id', type=int)
        account_ids = [account_id] if account_id else None

    holdings = compute_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({
            'holdings': [],
            'total_value_usd': 0,
            'total_value_eur': 0,
            'total_cost_basis': 0,
            'total_cost_basis_eur': 0,
            'total_gain_usd': 0,
            'total_gain_pct': 0,
            'realized_gains_usd': 0,
            'realized_gains_eur': 0,
            'sold_cost_basis_eur': 0,
            'eurusd_rate': 1.0
        })

    try:
        composition = compute_portfolio_composition(holdings)
        # Add realized gains
        realized = compute_realized_gains(request.user_id, account_ids)
        composition['realized_gains_usd'] = realized['total_usd']
        composition['realized_gains_eur'] = realized['total_eur']
        composition['sold_cost_basis_eur'] = realized['sold_cost_basis_eur']
        return jsonify(composition)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/investing/portfolio/performance-1m', methods=['GET'])
@app.route('/api/investing/portfolio/performance-period', methods=['GET'])
@login_required
def get_portfolio_performance_period():
    """Get portfolio performance over a period (lightweight). Supports days=1, 7, or 30."""
    from investing_utils import fetch_stock_price, get_fx_rate_to_eur, get_stock_currency, fetch_current_stock_prices_batch

    days = request.args.get('days', 30, type=int)
    if days not in [1, 7, 30]:
        days = 30

    account_ids_str = request.args.get('account_ids')
    if account_ids_str:
        account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()]
    else:
        account_id = request.args.get('account_id', type=int)
        account_ids = [account_id] if account_id else None

    holdings = compute_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({'performance': None, 'performance_1m': None, 'message': 'No holdings'})

    today = datetime.now()
    past_date = (today - timedelta(days=days)).strftime('%Y-%m-%d')

    # Batch fetch current prices
    all_tickers = [h['stock_ticker'] for h in holdings]
    prices = fetch_current_stock_prices_batch(all_tickers)

    current_value = 0
    past_value = 0

    for h in holdings:
        ticker = h['stock_ticker']
        qty = h['quantity']

        current_price = prices.get(ticker, 0) or 0

        try:
            past_price = fetch_stock_price(ticker, past_date)
            if past_price is None:
                past_price = current_price
        except:
            past_price = current_price

        currency = get_stock_currency(ticker)
        fx_rate = get_fx_rate_to_eur(currency)

        current_value += (current_price or 0) * qty * fx_rate
        past_value += (past_price or 0) * qty * fx_rate

    if past_value > 0:
        perf_pct = ((current_value - past_value) / past_value) * 100
    else:
        perf_pct = 0

    return jsonify({
        'performance': round(perf_pct, 2),
        'performance_1m': round(perf_pct, 2),  # backward compat
        'days': days,
        'current_value': round(current_value, 2),
        'past_value': round(past_value, 2)
    })


@app.route('/api/investing/portfolio/top-movers', methods=['GET'])
@login_required
def get_portfolio_top_movers():
    """Get top movers in portfolio based on price change over period."""
    from investing_utils import fetch_current_stock_prices_batch, fetch_historical_prices_batch, fetch_todays_open_prices_batch

    days = request.args.get('days', 30, type=int)
    if days not in [1, 7, 30]:
        days = 30

    account_ids_str = request.args.get('account_ids')
    if account_ids_str:
        account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()]
    else:
        account_id = request.args.get('account_id', type=int)
        account_ids = [account_id] if account_id else None

    holdings = compute_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({'movers': [], 'days': days})

    all_tickers = [h['stock_ticker'] for h in holdings]

    # Batch fetch current prices
    current_prices = fetch_current_stock_prices_batch(all_tickers)

    # For 1D: compare to today's market open price
    # For 7D/30D: compare to historical closing price
    if days == 1:
        past_prices = fetch_todays_open_prices_batch(all_tickers)
    else:
        today = datetime.now()
        past_date = (today - timedelta(days=days)).strftime('%Y-%m-%d')
        past_prices = fetch_historical_prices_batch(all_tickers, past_date)

    movers = []
    for h in holdings:
        ticker = h['stock_ticker']
        current_price = current_prices.get(ticker, 0) or 0
        past_price = past_prices.get(ticker, 0) or 0

        if current_price <= 0 or past_price <= 0:
            continue

        change_pct = ((current_price - past_price) / past_price) * 100

        movers.append({
            'ticker': ticker,
            'change_pct': round(change_pct, 1),
            'current_price': round(current_price, 2),
            'past_price': round(past_price, 2)
        })

    # Sort by change (best to worst)
    movers.sort(key=lambda x: x['change_pct'], reverse=True)

    return jsonify({
        'movers': movers[:5],
        'days': days
    })


@app.route('/api/investing/portfolio/stock-performance-3m', methods=['GET'])
@login_required
def get_stock_performance_3m():
    """Get 3-month stock price performance for all holdings in selected accounts."""
    from investing_utils import fetch_current_stock_prices_batch, fetch_historical_prices_batch

    account_ids_str = request.args.get('account_ids')
    if account_ids_str:
        account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()]
    else:
        account_id = request.args.get('account_id', type=int)
        account_ids = [account_id] if account_id else None

    holdings = compute_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({'stocks': []})

    all_tickers = [h['stock_ticker'] for h in holdings]

    # 3 months ago from today (same day, 3 months back)
    import calendar
    today = datetime.now()
    year = today.year
    month = today.month - 3
    if month <= 0:
        month += 12
        year -= 1
    day = min(today.day, calendar.monthrange(year, month)[1])
    past_date_str = f'{year:04d}-{month:02d}-{day:02d}'

    current_prices = fetch_current_stock_prices_batch(all_tickers)
    past_prices = fetch_historical_prices_batch(all_tickers, past_date_str)

    stocks = []
    for h in holdings:
        ticker = h['stock_ticker']
        current_price = current_prices.get(ticker, 0) or 0
        past_price = past_prices.get(ticker, 0) or 0

        if current_price <= 0 or past_price <= 0:
            continue

        change_pct = ((current_price - past_price) / past_price) * 100

        stocks.append({
            'ticker': ticker,
            'change_pct': round(change_pct, 1),
        })

    # Sort by change (best to worst)
    stocks.sort(key=lambda x: x['change_pct'], reverse=True)

    return jsonify({
        'stocks': stocks,
        'start_date': past_date_str,
    })


@app.route('/api/investing/watchlist/top-movers', methods=['GET'])
@login_required
def get_watchlist_top_movers():
    """Get top movers in watchlist based on price change over period."""
    from investing_utils import fetch_current_stock_prices_batch, fetch_historical_prices_batch, fetch_todays_open_prices_batch

    days = request.args.get('days', 30, type=int)
    if days not in [1, 7, 30]:
        days = 30

    # Get watchlist tickers
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT stock_ticker FROM watchlist WHERE user_id = ?',
            (request.user_id,)
        )
        watchlist_tickers = [row['stock_ticker'] for row in cursor.fetchall()]

    if not watchlist_tickers:
        return jsonify({'movers': [], 'days': days})

    # Batch fetch current prices
    current_prices = fetch_current_stock_prices_batch(watchlist_tickers)

    # For 1D: compare to today's market open price
    # For 7D/30D: compare to historical closing price
    if days == 1:
        past_prices = fetch_todays_open_prices_batch(watchlist_tickers)
    else:
        today = datetime.now()
        past_date = (today - timedelta(days=days)).strftime('%Y-%m-%d')
        past_prices = fetch_historical_prices_batch(watchlist_tickers, past_date)

    movers = []
    for ticker in watchlist_tickers:
        current_price = current_prices.get(ticker, 0) or 0
        past_price = past_prices.get(ticker, 0) or 0

        if current_price <= 0 or past_price <= 0:
            continue

        change_pct = ((current_price - past_price) / past_price) * 100

        movers.append({
            'ticker': ticker,
            'change_pct': round(change_pct, 1),
            'current_price': round(current_price, 2),
            'past_price': round(past_price, 2)
        })

    # Sort by change (best to worst)
    movers.sort(key=lambda x: x['change_pct'], reverse=True)

    return jsonify({
        'movers': movers[:5],
        'days': days
    })


@app.route('/api/investing/portfolio/performance', methods=['GET'])
@login_required
def get_portfolio_performance():
    """Get portfolio performance vs benchmark, tracking actual holdings over time."""
    benchmark = request.args.get('benchmark', 'NASDAQ')
    currency = request.args.get('currency', 'EUR')

    # Support both single account_id and comma-separated account_ids
    account_ids_str = request.args.get('account_ids')
    if account_ids_str:
        account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()]
    else:
        account_id = request.args.get('account_id', type=int)
        account_ids = [account_id] if account_id else None

    # Map benchmark name + currency to actual ticker
    benchmark_tickers = {
        'NASDAQ': {'USD': 'QQQ', 'EUR': 'EQQQ.DE'},
        'SP500': {'USD': 'SPY', 'EUR': 'CSPX.L'},
        # Legacy support
        'QQQ': {'USD': 'QQQ', 'EUR': 'EQQQ.DE'},
    }

    if benchmark not in benchmark_tickers:
        return jsonify({'error': f'Invalid benchmark. Use NASDAQ or SP500'}), 400

    if currency not in ['USD', 'EUR']:
        return jsonify({'error': 'Invalid currency. Use USD or EUR'}), 400

    benchmark_ticker = benchmark_tickers[benchmark].get(currency, 'QQQ')

    # Get all transactions (filtered by accounts if specified)
    with get_db() as conn:
        if account_ids and len(account_ids) > 0:
            placeholders = ','.join('?' for _ in account_ids)
            cursor = conn.execute(
                f'''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ? AND account_id IN ({placeholders})
                   ORDER BY transaction_date ASC''',
                (request.user_id, *account_ids)
            )
        else:
            cursor = conn.execute(
                '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
                   FROM portfolio_transactions WHERE user_id = ?
                   ORDER BY transaction_date ASC''',
                (request.user_id,)
            )
        rows = cursor.fetchall()

    if not rows:
        return jsonify({'error': 'No transactions found', 'data': [], 'summary': None})

    transactions = [dict(row) for row in rows]

    try:
        performance = compute_portfolio_performance_from_transactions(transactions, benchmark_ticker=benchmark_ticker)
        return jsonify(performance)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/investing/watchlist', methods=['GET'])
@login_required
def get_watchlist():
    """Get user's watchlist."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT stock_ticker FROM watchlist WHERE user_id = ? ORDER BY created_at ASC',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    symbols = [row['stock_ticker'] for row in rows]
    return jsonify({'symbols': symbols})


@app.route('/api/investing/watchlist', methods=['POST'])
@login_required
def add_to_watchlist():
    """Add symbol to watchlist."""
    data = request.get_json()
    symbol = data.get('symbol') if data else None
    if not symbol:
        return jsonify({'error': 'Symbol required'}), 400

    symbol = symbol.upper().strip()
    with get_db() as conn:
        try:
            conn.execute(
                'INSERT INTO watchlist (user_id, stock_ticker) VALUES (?, ?)',
                (request.user_id, symbol)
            )
        except Exception:
            # Already exists, ignore
            pass

    # Refresh videos for this ticker in the background
    refresh_videos_for_ticker_async(symbol)

    return jsonify({'success': True, 'symbol': symbol})


@app.route('/api/investing/watchlist/<symbol>', methods=['DELETE'])
@login_required
def remove_from_watchlist(symbol):
    """Remove symbol from watchlist."""
    symbol = symbol.upper().strip()
    with get_db() as conn:
        conn.execute(
            'DELETE FROM watchlist WHERE user_id = ? AND stock_ticker = ?',
            (request.user_id, symbol)
        )
    return jsonify({'success': True, 'symbol': symbol})


@app.route('/api/investing/market-cap', methods=['GET'])
def get_market_cap():
    """Get market cap for one or more tickers."""
    import yfinance as yf
    from investing_utils import EUROPEAN_TICKER_MAP, get_stock_currency

    tickers_param = request.args.get('tickers', '')
    if not tickers_param:
        return jsonify({'error': 'No tickers provided'}), 400

    tickers = [t.strip().upper() for t in tickers_param.split(',') if t.strip()]
    if not tickers:
        return jsonify({'error': 'No valid tickers provided'}), 400

    results = {}
    for ticker in tickers:
        try:
            # Map European tickers to Yahoo Finance format
            yf_ticker = EUROPEAN_TICKER_MAP.get(ticker, ticker)
            stock = yf.Ticker(yf_ticker)
            info = stock.info
            market_cap = info.get('marketCap')
            name = info.get('shortName') or info.get('longName') or ticker
            # Get currency from Yahoo Finance, fallback to exchange-based detection
            currency = info.get('currency') or get_stock_currency(ticker)

            results[ticker] = {
                'ticker': ticker,
                'name': name,
                'market_cap': market_cap,
                'currency': currency,
                'trailing_pe': info.get('trailingPE'),
                'forward_pe': info.get('forwardPE'),
                'dividend_yield': info.get('trailingAnnualDividendYield') or info.get('dividendYield'),
                'beta': info.get('beta'),
                'price_to_book': info.get('priceToBook'),
                'trailing_eps': info.get('trailingEps'),
                'profit_margin': info.get('profitMargins'),
                'return_on_equity': info.get('returnOnEquity'),
                'fifty_two_week_high': info.get('fiftyTwoWeekHigh'),
                'fifty_two_week_low': info.get('fiftyTwoWeekLow'),
                'revenue_growth': info.get('revenueGrowth'),
            }
        except Exception as e:
            results[ticker] = {
                'ticker': ticker,
                'error': str(e)
            }

    return jsonify({'stocks': results})


@app.route('/api/investing/stock-history/<ticker>', methods=['GET'])
def get_stock_history(ticker):
    """Get historical price data for a stock."""
    import yfinance as yf
    from investing_utils import EUROPEAN_TICKER_MAP, get_stock_currency

    ticker = ticker.upper()
    # Map European tickers to Yahoo Finance format
    yf_ticker = EUROPEAN_TICKER_MAP.get(ticker, ticker)
    period = request.args.get('period', '1M')  # Default to 1 month
    display_currency = request.args.get('currency', None)  # Optional: EUR or USD

    # Map period to yfinance parameters
    period_config = {
        '1D': {'period': '1d', 'interval': '5m'},
        '5D': {'period': '5d', 'interval': '15m'},
        '1M': {'period': '1mo', 'interval': '1d'},
        '6M': {'period': '6mo', 'interval': '1d'},
        'YTD': {'period': 'ytd', 'interval': '1d'},
        '1Y': {'period': '1y', 'interval': '1d'},
        '5Y': {'period': '5y', 'interval': '1wk'},
        'MAX': {'period': 'max', 'interval': '1mo'},
    }

    try:
        stock = yf.Ticker(yf_ticker)

        # Handle year-specific periods like 'Y2024', 'Y2023', etc.
        if period.upper().startswith('Y') and len(period) == 5:
            year = int(period[1:])
            start_date = f"{year}-01-01"
            end_date = f"{year}-12-31"
            hist = stock.history(start=start_date, end=end_date, interval='1d')
        else:
            config = period_config.get(period.upper(), period_config['1M'])
            hist = stock.history(period=config['period'], interval=config['interval'])

        if hist.empty:
            return jsonify({'error': 'No data available'}), 404

        # Get previous close and currency for reference
        info = stock.info
        previous_close = info.get('previousClose') or info.get('regularMarketPreviousClose')
        native_currency = info.get('currency') or get_stock_currency(ticker)

        # Determine output currency
        output_currency = display_currency.upper() if display_currency else native_currency

        # Fetch EUR/USD historical rates if currency conversion needed
        fx_rates = {}
        if output_currency != native_currency and output_currency in ('EUR', 'USD') and native_currency in ('EUR', 'USD'):
            # Fetch EUR/USD rates for all dates
            eurusd = yf.Ticker('EURUSD=X')
            from datetime import timedelta
            end_date_fx = (hist.index[-1].to_pydatetime() + timedelta(days=1)).strftime('%Y-%m-%d')
            fx_hist = eurusd.history(start=hist.index[0].strftime('%Y-%m-%d'),
                                      end=end_date_fx,
                                      interval='1d')
            for ts, row in fx_hist.iterrows():
                date_str = ts.strftime('%Y-%m-%d')
                fx_rates[date_str] = (row['Open'] + row['Close']) / 2

        # Format data for frontend with optional currency conversion
        data = []
        for timestamp, row in hist.iterrows():
            price = row['Close']

            # Convert currency if needed
            if output_currency != native_currency and native_currency in ('EUR', 'USD') and output_currency in ('EUR', 'USD'):
                date_str = timestamp.strftime('%Y-%m-%d')
                rate = fx_rates.get(date_str)
                if rate:
                    if native_currency == 'USD' and output_currency == 'EUR':
                        price = price / rate  # USD to EUR
                    elif native_currency == 'EUR' and output_currency == 'USD':
                        price = price * rate  # EUR to USD

            data.append({
                'timestamp': timestamp.isoformat(),
                'price': round(price, 2),
            })

        # Convert previous_close too
        if previous_close and output_currency != native_currency and fx_rates:
            # Use most recent rate for previous_close
            latest_rate = list(fx_rates.values())[-1] if fx_rates else None
            if latest_rate:
                if native_currency == 'USD' and output_currency == 'EUR':
                    previous_close = previous_close / latest_rate
                elif native_currency == 'EUR' and output_currency == 'USD':
                    previous_close = previous_close * latest_rate

        return jsonify({
            'ticker': ticker,
            'period': period.upper(),
            'previous_close': round(previous_close, 2) if previous_close else None,
            'currency': output_currency,
            'native_currency': native_currency,
            'data': data,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/investing/financials-history/<ticker>', methods=['GET'])
def get_financials_history(ticker):
    """Get historical financial data (quarterly income statement) for a stock."""
    import yfinance as yf
    import pandas as pd
    from investing_utils import EUROPEAN_TICKER_MAP

    ticker = ticker.upper()
    yf_ticker = EUROPEAN_TICKER_MAP.get(ticker, ticker)
    metric = request.args.get('metric', 'NetIncome')  # Default metric

    # Map frontend metric names to possible yfinance column names (try multiple variations)
    # Note: yfinance uses specific names like 'Net Income From Continuing Operation Net Minority Interest'
    METRIC_MAP = {
        'NetIncome': ['Net Income', 'Net Income Common Stockholders', 'Net Income From Continuing Operations',
                      'Net Income From Continuing Operation Net Minority Interest', 'Net Income Including Noncontrolling Interests'],
        'Revenue': ['Total Revenue', 'Revenue', 'Operating Revenue', 'Total Operating Income As Reported'],
        'GrossProfit': ['Gross Profit', 'Gross Margin'],
        'OperatingIncome': ['Operating Income', 'Operating Income Loss', 'EBIT', 'Operating Income/Loss'],
        'EBITDA': ['EBITDA', 'Normalized EBITDA'],
        'EPS': ['Basic EPS', 'Diluted EPS', 'Basic Earnings Per Share', 'Diluted Earnings Per Share'],
    }

    metric_variations = METRIC_MAP.get(metric, [metric])

    try:
        stock = yf.Ticker(yf_ticker)
        info = stock.info
        company_name = info.get('shortName') or info.get('longName') or ticker
        currency = info.get('financialCurrency') or info.get('currency') or 'USD'

        # Get quarterly income statement (most detailed)
        quarterly = stock.quarterly_income_stmt
        annual = stock.income_stmt

        # Debug logging
        print(f"[Financials] Ticker: {yf_ticker}, Metric: {metric}, Looking for: {metric_variations}")
        print(f"[Financials] Quarterly data available: {quarterly is not None and not quarterly.empty}")
        print(f"[Financials] Annual data available: {annual is not None and not annual.empty}")
        if quarterly is not None and not quarterly.empty:
            print(f"[Financials] Quarterly index (ALL): {list(quarterly.index)}")
        if annual is not None and not annual.empty:
            print(f"[Financials] Annual index (ALL): {list(annual.index)}")

        data_points = []

        # Find the actual metric name in the data
        def find_metric_in_df(df, variations):
            if df is None or df.empty:
                return None
            for var in variations:
                if var in df.index:
                    return var
            # Also try case-insensitive matching
            for var in variations:
                for idx in df.index:
                    if var.lower() == idx.lower():
                        return idx
            return None

        # Process quarterly data
        if quarterly is not None and not quarterly.empty:
            actual_metric = find_metric_in_df(quarterly, metric_variations)
            print(f"[Financials] Found metric: {actual_metric}")
            if actual_metric:
                for col in quarterly.columns:
                    try:
                        value = quarterly.loc[actual_metric, col]
                        if pd.notna(value):
                            # col is a Timestamp
                            date = col.to_pydatetime()
                            quarter = f"Q{((date.month - 1) // 3) + 1} {date.year}"
                            data_points.append({
                                'date': date.strftime('%Y-%m-%d'),
                                'quarter': quarter,
                                'value': float(value),
                                'type': 'quarterly'
                            })
                    except Exception as e:
                        print(f"[Financials] Error extracting value: {e}")
                        continue
                print(f"[Financials] Extracted {len(data_points)} quarterly data points")

        # Also get annual data for longer history
        if annual is not None and not annual.empty:
            actual_metric = find_metric_in_df(annual, metric_variations)
            if actual_metric:
                for col in annual.columns:
                    try:
                        value = annual.loc[actual_metric, col]
                        if pd.notna(value):
                            date = col.to_pydatetime()
                            # Only add if we don't have quarterly data for this year
                            year_str = f"FY {date.year}"
                            data_points.append({
                                'date': date.strftime('%Y-%m-%d'),
                                'quarter': year_str,
                                'value': float(value),
                                'type': 'annual'
                            })
                    except:
                        continue

        # Sort by date (oldest first for charting)
        data_points.sort(key=lambda x: x['date'])

        # Calculate growth rates
        def calc_growth(data, years):
            if len(data) < 2:
                return None
            # Filter to quarterly data for more accurate comparison
            quarterly_data = [d for d in data if d['type'] == 'quarterly']
            if len(quarterly_data) < 4:  # Need at least 1 year of quarters
                quarterly_data = data  # Fall back to all data

            if len(quarterly_data) < 2:
                return None

            # Get TTM (trailing twelve months) values
            current_value = sum(d['value'] for d in quarterly_data[-4:]) if len(quarterly_data) >= 4 else quarterly_data[-1]['value']

            # Find value from N years ago
            target_quarters = years * 4
            if len(quarterly_data) > target_quarters:
                past_value = sum(d['value'] for d in quarterly_data[-(target_quarters + 4):-target_quarters]) if len(quarterly_data) >= target_quarters + 4 else quarterly_data[-target_quarters - 1]['value']
            elif len(quarterly_data) > 4:
                # Use oldest available data
                past_value = sum(d['value'] for d in quarterly_data[:4]) if len(quarterly_data) >= 4 else quarterly_data[0]['value']
            else:
                return None

            if past_value == 0 or past_value is None:
                return None

            return round(((current_value / past_value) - 1) * 100, 2)

        growth_rates = {
            '1Y': calc_growth(data_points, 1),
            '2Y': calc_growth(data_points, 2),
            '5Y': calc_growth(data_points, 5),
            '10Y': calc_growth(data_points, 10),
        }

        return jsonify({
            'ticker': ticker,
            'company_name': company_name,
            'metric': metric,
            'metric_label': metric_variations[0] if metric_variations else metric,
            'currency': currency,
            'data': data_points,
            'growth_rates': growth_rates,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# =============================================================================
# Investment Accounts API (for fee tracking)
# =============================================================================

# French banks/brokers with their fee structures
BANKS = {
    'CREDIT_AGRICOLE': {
        'name': 'Crédit Agricole',
        'order_fee_pct': 1.30,  # variable selon caisse régionale
        'order_fee_min': 16,
        'custody_fee_pct_year': 0.30,  # + frais par ligne, variable selon caisse
        'custody_fee_pct_year_pea': 0.40,  # plafonné légalement + 5€/ligne
        'fx_fee_info_fr': '0.50% (min 20€)',
        'fx_fee_info_en': '0.50% (min €20)',
    },
    'BNP_PARIBAS': {
        'name': 'BNP Paribas',
        'order_fee_pct': 0.50,
        'order_fee_min': 5,
        'custody_fee_pct_year': 0.40,  # <50k, dégressif pour gros montants
        'custody_fee_pct_year_pea': 0.40,  # plafonné légalement
        'fx_fee_info_fr': '0.50% (min 15-25€)',
        'fx_fee_info_en': '0.50% (min €15-25)',
    },
    'SOCIETE_GENERALE': {
        'name': 'Société Générale',
        'order_fee_pct': 0.50,  # 0.50% ≤2k€, 0.45% 2-8k€, 0.35% >8k€
        'order_fee_min': 6,  # CTO uniquement, PEA pas de minimum
        'custody_fee_pct_year': 0.30,  # + 4.50€/ligne, dégressif selon montant
        'custody_fee_pct_year_pea': 0.40,  # plafonné légalement
        'account_fee_year': 17.50,  # offert si 1 achat/an
        'fx_fee_info_fr': '0.50% (min 16€ USA, 40€ autres)',
        'fx_fee_info_en': '0.50% (min €16 USA, €40 others)',
    },
    'CREDIT_MUTUEL': {
        'name': 'Crédit Mutuel',
        'order_fee_pct': 0.50,
        'order_fee_min': 5,
        'custody_fee_pct_year': 0.25,  # 0.125%/semestre
        'custody_fee_pct_year_pea': 0.40,  # plafonné légalement
        'fx_fee_info_fr': '0.50% (min 30€)',
        'fx_fee_info_en': '0.50% (min €30)',
    },
    'REVOLUT': {
        'name': 'Revolut',
        'order_fee_pct': 0,
        'order_fee_min': 0,  # Free trades (limited), then €1/trade for Standard
        'custody_fee_pct_year': 0,
        'custody_fee_pct_year_pea': None,  # No PEA available
        'fx_fee_info_fr': '0% (taux interbancaire)',
        'fx_fee_info_en': '0% (interbank rate)',
    },
    'FORTUNEO': {
        'name': 'Fortuneo',
        'order_fee_pct': 0.35,  # 0.35% for orders <2k€, then 0.20% for 2-10k€
        'order_fee_min': 1.95,  # Optimum plan
        'custody_fee_pct_year': 0,
        'custody_fee_pct_year_pea': 0,  # No custody fees on PEA
        'fx_fee_info_fr': '0.10%',
        'fx_fee_info_en': '0.10%',
    },
    'INTERACTIVE_BROKERS': {
        'name': 'Interactive Brokers',
        'order_fee_pct': 0.05,  # Tiered: 0.05% (min $1, max 1% of trade)
        'order_fee_min': 1,  # $1 minimum per order
        'custody_fee_pct_year': 0,
        'custody_fee_pct_year_pea': None,  # No PEA available
        'fx_fee_info_fr': '0.002% (min $2)',
        'fx_fee_info_en': '0.002% (min $2)',
    },
}

# Account types with tax implications
ACCOUNT_TYPES = {
    'CTO': {
        'name': 'CTO',
        'description_en': '30% flat tax on capital gains (PFU)',
        'description_fr': '30% d\'imposition forfaitaire sur les plus-values (PFU)',
        'tax_rate': 30.0
    },
    'PEA': {
        'name': 'PEA',
        'description_en': '17.2% social contributions only (after 5 years)',
        'description_fr': '17.2% de prélèvements sociaux uniquement (après 5 ans)',
        'tax_rate': 17.2
    },
}


@app.route('/api/investing/banks', methods=['GET'])
def get_banks():
    """Get list of available banks with their fee structures."""
    return jsonify({'banks': BANKS})


@app.route('/api/investing/account-types', methods=['GET'])
def get_account_types():
    """Get list of available account types with tax info."""
    return jsonify({'account_types': ACCOUNT_TYPES})


@app.route('/api/investing/accounts', methods=['GET'])
@login_required
def get_accounts():
    """Get user's investment accounts."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT id, name, account_type, bank, display_order, created_at
               FROM investment_accounts WHERE user_id = ?
               ORDER BY display_order ASC, created_at ASC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    accounts = [{
        'id': row['id'],
        'name': row['name'],
        'account_type': row['account_type'],
        'bank': row['bank'],
        'bank_info': BANKS.get(row['bank'], {}),
        'type_info': ACCOUNT_TYPES.get(row['account_type'], {}),
    } for row in rows]
    return jsonify({'accounts': accounts})


@app.route('/api/investing/accounts', methods=['POST'])
@login_required
def create_account():
    """Create a new investment account."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    name = data.get('name', '').strip()
    account_type = data.get('account_type', '').upper().strip()
    bank = data.get('bank', '').upper().strip()

    if not name:
        return jsonify({'error': 'Account name required'}), 400
    if account_type not in ACCOUNT_TYPES:
        return jsonify({'error': f'Invalid account type. Valid: {list(ACCOUNT_TYPES.keys())}'}), 400
    if bank not in BANKS:
        return jsonify({'error': f'Invalid bank. Valid: {list(BANKS.keys())}'}), 400

    with get_db() as conn:
        # Get max display_order for this user
        cursor = conn.execute(
            'SELECT COALESCE(MAX(display_order), -1) + 1 as next_order FROM investment_accounts WHERE user_id = ?',
            (request.user_id,)
        )
        next_order = cursor.fetchone()['next_order']

        cursor = conn.execute('''
            INSERT INTO investment_accounts (user_id, name, account_type, bank, display_order)
            VALUES (?, ?, ?, ?, ?)
            RETURNING id
        ''', (request.user_id, name, account_type, bank, next_order))
        account_id = cursor.fetchone()['id']

    return jsonify({
        'success': True,
        'id': account_id,
        'name': name,
        'account_type': account_type,
        'bank': bank,
        'bank_info': BANKS[bank],
        'type_info': ACCOUNT_TYPES[account_type],
    })


@app.route('/api/investing/accounts/<int:account_id>', methods=['DELETE'])
@login_required
def delete_account(account_id):
    """Delete an investment account."""
    with get_db() as conn:
        # Get account and user info before deletion
        cursor = conn.execute('''
            SELECT ia.name, ia.account_type, ia.bank, u.name as user_name, u.email
            FROM investment_accounts ia
            JOIN users u ON ia.user_id = u.id
            WHERE ia.user_id = ? AND ia.id = ?
        ''', (request.user_id, account_id))
        row = cursor.fetchone()

        if row:
            # Send admin notification
            try:
                send_admin_deletion_alert(
                    user_name=row['user_name'],
                    user_email=row['email'],
                    deletion_type='account',
                    details={
                        'Account ID': account_id,
                        'Account name': row['name'],
                        'Type': row['account_type'],
                        'Bank': row['bank'],
                    }
                )
            except Exception as e:
                logger.error(f"Failed to send deletion alert: {e}")

        # First delete associated transactions
        conn.execute(
            'DELETE FROM portfolio_transactions WHERE user_id = ? AND account_id = ?',
            (request.user_id, account_id)
        )
        # Then delete the account
        conn.execute(
            'DELETE FROM investment_accounts WHERE user_id = ? AND id = ?',
            (request.user_id, account_id)
        )
    return jsonify({'success': True, 'id': account_id})


@app.route('/api/investing/accounts/<int:account_id>/duplicate', methods=['POST'])
@login_required
def duplicate_account(account_id):
    """Duplicate an investment account with all its transactions."""
    with get_db() as conn:
        # Get original account
        cursor = conn.execute('''
            SELECT name, account_type, bank
            FROM investment_accounts
            WHERE user_id = ? AND id = ?
        ''', (request.user_id, account_id))
        original = cursor.fetchone()

        if not original:
            return jsonify({'error': 'Account not found'}), 404

        # Create new account with "Copy of" prefix
        new_name = f"Copy of {original['name']}"

        # Get max display_order for this user
        cursor = conn.execute(
            'SELECT COALESCE(MAX(display_order), -1) + 1 as next_order FROM investment_accounts WHERE user_id = ?',
            (request.user_id,)
        )
        next_order = cursor.fetchone()['next_order']

        # Create the new account
        cursor = conn.execute('''
            INSERT INTO investment_accounts (user_id, name, account_type, bank, display_order)
            VALUES (?, ?, ?, ?, ?)
            RETURNING id
        ''', (request.user_id, new_name, original['account_type'], original['bank'], next_order))
        new_account_id = cursor.fetchone()['id']

        # Copy all transactions from original account to new account
        cursor = conn.execute('''
            SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency
            FROM portfolio_transactions
            WHERE user_id = ? AND account_id = ?
        ''', (request.user_id, account_id))
        transactions = cursor.fetchall()

        for tx in transactions:
            conn.execute('''
                INSERT INTO portfolio_transactions (user_id, account_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share, price_currency)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (request.user_id, new_account_id, tx['stock_ticker'], tx['transaction_type'], tx['quantity'], tx['transaction_date'], tx['price_per_share'], tx['price_currency']))

    return jsonify({
        'success': True,
        'id': new_account_id,
        'name': new_name,
        'account_type': original['account_type'],
        'bank': original['bank'],
        'bank_info': BANKS[original['bank']],
        'type_info': ACCOUNT_TYPES[original['account_type']],
        'transactions_copied': len(transactions),
    })


@app.route('/api/investing/accounts/<int:account_id>/rename', methods=['PUT'])
@login_required
def rename_account(account_id):
    """Rename an investment account."""
    data = request.get_json()
    new_name = data.get('name', '').strip()

    if not new_name:
        return jsonify({'error': 'Name is required'}), 400

    with get_db() as conn:
        # Verify account belongs to user
        cursor = conn.execute('''
            SELECT id, account_type, bank
            FROM investment_accounts
            WHERE user_id = ? AND id = ?
        ''', (request.user_id, account_id))
        account = cursor.fetchone()

        if not account:
            return jsonify({'error': 'Account not found'}), 404

        # Update the name
        conn.execute('''
            UPDATE investment_accounts
            SET name = ?
            WHERE user_id = ? AND id = ?
        ''', (new_name, request.user_id, account_id))

    return jsonify({
        'success': True,
        'id': account_id,
        'name': new_name,
        'account_type': account['account_type'],
        'bank': account['bank'],
        'bank_info': BANKS[account['bank']],
        'type_info': ACCOUNT_TYPES[account['account_type']],
    })


@app.route('/api/investing/accounts/reorder', methods=['PUT'])
@login_required
def reorder_accounts():
    """Reorder investment accounts (drag & drop)."""
    data = request.get_json()
    if not data or 'account_ids' not in data:
        return jsonify({'error': 'account_ids array required'}), 400

    account_ids = data['account_ids']
    if not isinstance(account_ids, list):
        return jsonify({'error': 'account_ids must be an array'}), 400

    with get_db() as conn:
        # Update display_order for each account
        for order, account_id in enumerate(account_ids):
            conn.execute(
                'UPDATE investment_accounts SET display_order = ? WHERE user_id = ? AND id = ?',
                (order, request.user_id, account_id)
            )

    return jsonify({'success': True})


@app.route('/api/investing/earnings-watchlist', methods=['GET'])
@login_required
def get_earnings_watchlist():
    """Get user's earnings watchlist."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT stock_ticker FROM earnings_watchlist WHERE user_id = ? ORDER BY created_at ASC',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    symbols = [row['stock_ticker'] for row in rows]
    return jsonify({'symbols': symbols})


@app.route('/api/investing/earnings-watchlist', methods=['POST'])
@login_required
def add_to_earnings_watchlist():
    """Add symbol to earnings watchlist."""
    data = request.get_json()
    symbol = data.get('symbol') if data else None
    if not symbol:
        return jsonify({'error': 'Symbol required'}), 400

    symbol = symbol.upper().strip()
    with get_db() as conn:
        try:
            conn.execute(
                'INSERT INTO earnings_watchlist (user_id, stock_ticker) VALUES (?, ?)',
                (request.user_id, symbol)
            )
        except Exception:
            # Already exists, ignore
            pass

    return jsonify({'success': True, 'symbol': symbol})


@app.route('/api/investing/earnings-watchlist/<symbol>', methods=['DELETE'])
@login_required
def remove_from_earnings_watchlist(symbol):
    """Remove symbol from earnings watchlist."""
    symbol = symbol.upper().strip()
    with get_db() as conn:
        conn.execute(
            'DELETE FROM earnings_watchlist WHERE user_id = ? AND stock_ticker = ?',
            (request.user_id, symbol)
        )
    return jsonify({'success': True, 'symbol': symbol})


def fetch_earnings_from_yfinance(ticker):
    """Fetch earnings date from yfinance for a single ticker.
    Returns (date_string, date_confirmed, is_estimated).
    If no future date found, estimates based on last earnings + 3 months.
    """
    import yfinance as yf
    from investing_utils import get_yfinance_ticker
    from dateutil.relativedelta import relativedelta

    try:
        # Convert to yfinance ticker (add exchange suffix for European stocks)
        yf_ticker = get_yfinance_ticker(ticker)
        stock = yf.Ticker(yf_ticker)
        calendar = stock.calendar

        next_earnings_date = None
        date_confirmed = False

        if calendar is not None:
            # Handle different yfinance return formats
            if hasattr(calendar, 'to_dict'):
                # DataFrame format
                cal_dict = calendar.to_dict()
                if 'Earnings Date' in cal_dict:
                    dates = cal_dict['Earnings Date']
                    if dates:
                        first_key = list(dates.keys())[0]
                        next_earnings_date = dates[first_key]
                        date_confirmed = len(dates) == 1
            elif isinstance(calendar, dict):
                # Dict format (newer yfinance)
                if 'Earnings Date' in calendar:
                    earnings_dates = calendar['Earnings Date']
                    if isinstance(earnings_dates, list) and len(earnings_dates) > 0:
                        next_earnings_date = earnings_dates[0]
                        date_confirmed = len(earnings_dates) == 1
                    elif earnings_dates:
                        next_earnings_date = earnings_dates
                        date_confirmed = True

        # Convert to date and check if it's in the future
        today = datetime.now().date()
        if next_earnings_date is not None:
            if hasattr(next_earnings_date, 'date'):
                next_earnings_date = next_earnings_date.date()
            elif isinstance(next_earnings_date, str):
                next_earnings_date = datetime.strptime(next_earnings_date, '%Y-%m-%d').date()
            # Only return if the date is in the future
            if next_earnings_date > today:
                return next_earnings_date.strftime('%Y-%m-%d'), date_confirmed, False

        # No future earnings date found - try to estimate from historical data
        last_earnings = None

        # Method 0: Use the past calendar date if available (most accurate - actual earnings date)
        if next_earnings_date is not None and next_earnings_date <= today:
            last_earnings = next_earnings_date

        # Method 1: Try earnings_dates (most accurate)
        try:
            earnings_dates = stock.earnings_dates
            if earnings_dates is not None and len(earnings_dates) > 0:
                past_dates = [d.date() if hasattr(d, 'date') else d for d in earnings_dates.index]
                past_dates = [d for d in past_dates if d <= today]
                if past_dates:
                    last_earnings = max(past_dates)
        except Exception:
            pass

        # Method 2: Try quarterly_income_stmt columns (fallback for stocks without earnings_dates)
        if last_earnings is None:
            try:
                quarterly_income = stock.quarterly_income_stmt
                if quarterly_income is not None and len(quarterly_income.columns) > 0:
                    # Columns are the report dates
                    col_dates = []
                    for col in quarterly_income.columns:
                        if hasattr(col, 'date'):
                            col_dates.append(col.date())
                        elif isinstance(col, str):
                            try:
                                col_dates.append(datetime.strptime(col, '%Y-%m-%d').date())
                            except ValueError:
                                pass
                    past_dates = [d for d in col_dates if d <= today]
                    if past_dates:
                        last_earnings = max(past_dates)
            except Exception:
                pass

        # Method 3: Try quarterly_financials (another fallback)
        if last_earnings is None:
            try:
                quarterly_fin = stock.quarterly_financials
                if quarterly_fin is not None and len(quarterly_fin.columns) > 0:
                    col_dates = []
                    for col in quarterly_fin.columns:
                        if hasattr(col, 'date'):
                            col_dates.append(col.date())
                        elif isinstance(col, str):
                            try:
                                col_dates.append(datetime.strptime(col, '%Y-%m-%d').date())
                            except ValueError:
                                pass
                    past_dates = [d for d in col_dates if d <= today]
                    if past_dates:
                        last_earnings = max(past_dates)
            except Exception:
                pass

        # Estimate next earnings as last + 3 months
        if last_earnings:
            estimated_date = last_earnings + relativedelta(months=3)
            # If estimated date is in the past, add another 3 months
            while estimated_date <= today:
                estimated_date = estimated_date + relativedelta(months=3)
            return estimated_date.strftime('%Y-%m-%d'), False, True

        return None, False, False

    except Exception as e:
        print(f"Error fetching earnings for {ticker}: {e}")
        return None, False, False


def get_cached_earnings(ticker):
    """Get cached earnings data if fresh.
    Successful lookups (with date) are cached for 48 hours.
    Failed lookups (null date) are cached for only 2 hours to retry sooner.
    Returns (date, confirmed, earnings_time, is_fresh).
    """
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT next_earnings_date, date_confirmed, earnings_time, updated_at
               FROM earnings_cache WHERE ticker = ?''',
            (ticker,)
        )
        row = cursor.fetchone()

        if not row:
            return None, None, None, False  # Not in cache

        updated_at = row['updated_at']
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at)
        age_hours = (datetime.now() - updated_at.replace(tzinfo=None)).total_seconds() / 3600

        # Null results have shorter TTL (2 hours) to retry failed lookups sooner
        max_age = 48 if row['next_earnings_date'] else 2

        if age_hours < max_age:
            # Cache is fresh
            return row['next_earnings_date'], bool(row['date_confirmed']), row['earnings_time'], True

        return None, None, None, False  # Cache is stale


def save_earnings_cache(ticker, next_earnings_date, date_confirmed, earnings_time=None):
    """Save earnings data to cache."""
    with get_db() as conn:
        conn.execute(
            '''INSERT INTO earnings_cache (ticker, next_earnings_date, date_confirmed, earnings_time, updated_at)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(ticker) DO UPDATE SET
                   next_earnings_date = excluded.next_earnings_date,
                   date_confirmed = excluded.date_confirmed,
                   earnings_time = excluded.earnings_time,
                   updated_at = excluded.updated_at''',
            (ticker, next_earnings_date, 1 if date_confirmed else 0, earnings_time, datetime.now().isoformat())
        )


def get_cached_dividend(ticker):
    """Get cached dividend data if fresh.
    Data is cached for 7 days (168 hours) since dividends change infrequently.
    Returns (ex_date, amount, pays_dividends, is_fresh) or (None, None, None, False).
    """
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT ex_dividend_date, dividend_amount, pays_dividends, updated_at
               FROM dividends_cache WHERE ticker = ?''',
            (ticker,)
        )
        row = cursor.fetchone()

        if not row:
            return None, None, None, False  # Not in cache

        updated_at = row['updated_at']
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at)
        age_hours = (datetime.now() - updated_at.replace(tzinfo=None)).total_seconds() / 3600

        # 7 days TTL for dividend data
        if age_hours < 168:
            return row['ex_dividend_date'], row['dividend_amount'], bool(row['pays_dividends']), True

        return None, None, None, False  # Cache is stale


def save_dividend_cache(ticker, ex_dividend_date, dividend_amount, pays_dividends=True):
    """Save dividend data to cache."""
    with get_db() as conn:
        conn.execute(
            '''INSERT INTO dividends_cache (ticker, ex_dividend_date, dividend_amount, pays_dividends, updated_at)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(ticker) DO UPDATE SET
                   ex_dividend_date = excluded.ex_dividend_date,
                   dividend_amount = excluded.dividend_amount,
                   pays_dividends = excluded.pays_dividends,
                   updated_at = excluded.updated_at''',
            (ticker, ex_dividend_date, dividend_amount, 1 if pays_dividends else 0, datetime.now().isoformat())
        )


@app.route('/api/investing/earnings-calendar', methods=['GET'])
@login_required
def get_earnings_calendar():
    """Get upcoming earnings dates for portfolio holdings and watchlist.
    Uses yfinance with database cache (48 hour TTL).
    Optionally filters portfolio by account_id.
    """
    include_portfolio = request.args.get('include_portfolio', 'true').lower() == 'true'
    include_watchlist = request.args.get('include_watchlist', 'true').lower() == 'true'
    account_id = request.args.get('account_id', type=int)

    tickers = set()
    portfolio_tickers = set()
    watchlist_tickers = set()

    # Get portfolio holdings if include_portfolio is true
    if include_portfolio:
        account_ids = [account_id] if account_id else None
        holdings = compute_holdings_from_transactions(request.user_id, account_ids)
        portfolio_tickers = {h['stock_ticker'] for h in holdings if h['quantity'] > 0}
        tickers.update(portfolio_tickers)

    # Get watchlist tickers if include_watchlist is true
    if include_watchlist:
        with get_db() as conn:
            cursor = conn.execute(
                'SELECT stock_ticker FROM watchlist WHERE user_id = ?',
                (request.user_id,)
            )
            watchlist_tickers = {row['stock_ticker'] for row in cursor.fetchall()}
            tickers.update(watchlist_tickers)

    if not tickers:
        return jsonify({'earnings': [], 'watchlist': [], 'message': 'No tickers to track'})

    today = datetime.now().date()
    earnings_data = []

    for ticker in tickers:
        # Try to get from cache first
        cached_date, cached_confirmed, _, is_fresh = get_cached_earnings(ticker)

        is_estimated = False
        if is_fresh:
            # Use cached data
            next_earnings_date = cached_date
            date_confirmed = cached_confirmed
        else:
            # Fetch fresh data from yfinance
            next_earnings_date, date_confirmed, is_estimated = fetch_earnings_from_yfinance(ticker)
            # Save to cache (even if None, to avoid repeated fetches) - don't cache estimated dates
            if not is_estimated:
                save_earnings_cache(ticker, next_earnings_date, date_confirmed)

        if next_earnings_date:
            earnings_date = datetime.strptime(next_earnings_date, '%Y-%m-%d').date()
            remaining_days = (earnings_date - today).days

            # Handle past earnings dates
            if remaining_days < 0:
                # Cached date is stale, try to refresh
                next_earnings_date, date_confirmed, is_estimated = fetch_earnings_from_yfinance(ticker)
                if not is_estimated:
                    save_earnings_cache(ticker, next_earnings_date, date_confirmed)
                if next_earnings_date:
                    earnings_date = datetime.strptime(next_earnings_date, '%Y-%m-%d').date()
                    remaining_days = (earnings_date - today).days
                    if remaining_days < 0:
                        # Still in the past after refresh, show as unavailable
                        earnings_data.append({
                            'ticker': ticker,
                            'next_earnings_date': None,
                            'remaining_days': None,
                            'date_confirmed': False,
                            'is_estimated': False,
                            'source': 'portfolio' if ticker in portfolio_tickers else 'watchlist'
                        })
                        continue
                else:
                    # No future date found, show as unavailable
                    earnings_data.append({
                        'ticker': ticker,
                        'next_earnings_date': None,
                        'remaining_days': None,
                        'date_confirmed': False,
                        'is_estimated': False,
                        'source': 'portfolio' if ticker in portfolio_tickers else 'watchlist'
                    })
                    continue

            earnings_data.append({
                'ticker': ticker,
                'next_earnings_date': next_earnings_date,
                'remaining_days': remaining_days,
                'date_confirmed': date_confirmed,
                'is_estimated': is_estimated,
                'source': 'portfolio' if ticker in portfolio_tickers else 'watchlist'
            })
        else:
            earnings_data.append({
                'ticker': ticker,
                'next_earnings_date': None,
                'remaining_days': None,
                'date_confirmed': False,
                'is_estimated': False,
                'source': 'portfolio' if ticker in portfolio_tickers else 'watchlist'
            })

    # Sort by remaining days (nulls at the end)
    earnings_data.sort(key=lambda x: (x['remaining_days'] is None, x['remaining_days'] or 9999))

    return jsonify({
        'earnings': earnings_data,
        'watchlist': list(watchlist_tickers)
    })


@app.route('/api/investing/dividends-calendar', methods=['GET'])
@login_required
def get_dividends_calendar():
    """Get upcoming dividend dates for portfolio holdings using FMP API with caching."""
    import yfinance as yf
    import requests
    from investing_utils import get_yfinance_ticker

    account_ids_str = request.args.get('account_ids', '')
    account_ids = [int(x) for x in account_ids_str.split(',') if x.strip()] if account_ids_str else None

    # Get portfolio holdings
    holdings = compute_holdings_from_transactions(request.user_id, account_ids)
    holdings_by_ticker = {h['stock_ticker']: h['quantity'] for h in holdings if h['quantity'] > 0}
    portfolio_tickers = list(holdings_by_ticker.keys())

    if not portfolio_tickers:
        return jsonify({'dividends': []})

    today = datetime.now().date()
    dividends_data = []

    # Check which tickers need fresh data (not in cache or stale)
    tickers_needing_fetch = []
    cached_data = {}
    for ticker in portfolio_tickers:
        ex_date, amount, pays_div, is_fresh = get_cached_dividend(ticker)
        if is_fresh:
            cached_data[ticker] = (ex_date, amount, pays_div)
        else:
            tickers_needing_fetch.append(ticker)

    # Only fetch FMP calendar if we have tickers that need fresh data
    fmp_dividends = {}
    if tickers_needing_fetch:
        fmp_api_key = os.environ.get('FMP_API_KEY')
        if fmp_api_key:
            try:
                from_date = today.strftime('%Y-%m-%d')
                to_date = (today + timedelta(days=180)).strftime('%Y-%m-%d')
                fmp_url = f"https://financialmodelingprep.com/api/v3/stock_dividend_calendar?from={from_date}&to={to_date}&apikey={fmp_api_key}"
                fmp_response = requests.get(fmp_url, timeout=10)

                if fmp_response.status_code == 200:
                    fmp_calendar = fmp_response.json()
                    for event in fmp_calendar:
                        symbol = event.get('symbol', '')
                        if symbol and symbol not in fmp_dividends:
                            fmp_dividends[symbol] = event
            except Exception as e:
                app.logger.warning(f"Failed to fetch FMP dividend calendar: {e}")

    for ticker in portfolio_tickers:
        quantity = holdings_by_ticker.get(ticker, 0)

        # Use cached data if available
        if ticker in cached_data:
            ex_date_str, dividend_amount, pays_dividends = cached_data[ticker]
            remaining_days = None
            if ex_date_str:
                try:
                    ex_date = datetime.strptime(ex_date_str, '%Y-%m-%d').date()
                    remaining_days = (ex_date - today).days
                except Exception:
                    pass

            dividend_per_share = round(dividend_amount, 4) if dividend_amount else None
            total_dividend = round(quantity * dividend_amount, 2) if dividend_amount and quantity else None

            dividends_data.append({
                'ticker': ticker,
                'ex_dividend_date': ex_date_str,
                'payment_date': None,
                'remaining_days': remaining_days,
                'dividend_amount': dividend_per_share,
                'dividend_yield': None,
                'frequency': None,
                'confirmed': True,
                'pays_dividends': pays_dividends,
                'quantity': quantity,
                'total_dividend': total_dividend,
                'amount_source': 'cache',
            })
            continue

        # Fetch fresh data for this ticker
        try:
            yf_ticker = get_yfinance_ticker(ticker)
            stock = yf.Ticker(yf_ticker)
            info = stock.info

            dividend_rate = info.get('dividendRate')
            dividend_yield = info.get('dividendYield')
            last_dividend = info.get('lastDividendValue')

            if dividend_yield and dividend_yield < 1:
                dividend_yield = dividend_yield * 100

            ex_date_str = None
            remaining_days = None
            dividend_amount = None
            payment_date = None
            confirmed = False
            amount_source = None
            frequency = None

            freq_days = None
            if dividend_rate and last_dividend and last_dividend > 0:
                ratio = dividend_rate / last_dividend
                if ratio > 3.5:
                    frequency = 'Quarterly'
                    freq_days = 91
                elif ratio > 1.8:
                    frequency = 'Semi-Annual'
                    freq_days = 182
                else:
                    frequency = 'Annual'
                    freq_days = 365

            yf_ex_date = info.get('exDividendDate')
            if yf_ex_date:
                yf_date = datetime.fromtimestamp(yf_ex_date).date()
                if yf_date >= today:
                    ex_date_str = yf_date.strftime('%Y-%m-%d')
                    remaining_days = (yf_date - today).days
                    confirmed = True
                    dividend_amount = last_dividend
                    amount_source = 'yfinance'

            if not ex_date_str:
                fmp_data = fmp_dividends.get(ticker) or fmp_dividends.get(yf_ticker)
                if fmp_data:
                    ex_date_str = fmp_data.get('date')
                    payment_date = fmp_data.get('paymentDate')
                    dividend_amount = fmp_data.get('adjDividend') or fmp_data.get('dividend')
                    if ex_date_str:
                        ex_date = datetime.strptime(ex_date_str, '%Y-%m-%d').date()
                        remaining_days = (ex_date - today).days
                        confirmed = True
                        amount_source = 'fmp'

            if not ex_date_str and yf_ex_date and freq_days:
                last_ex_date = datetime.fromtimestamp(yf_ex_date).date()
                estimated_date = last_ex_date
                while estimated_date < today:
                    estimated_date = estimated_date + timedelta(days=freq_days)
                ex_date_str = estimated_date.strftime('%Y-%m-%d')
                remaining_days = (estimated_date - today).days
                confirmed = False
                dividend_amount = last_dividend
                amount_source = 'estimate'

            if frequency is None and dividend_rate and last_dividend and last_dividend > 0:
                ratio = dividend_rate / last_dividend
                if ratio > 3.5:
                    frequency = 'Quarterly'
                elif ratio > 1.8:
                    frequency = 'Semi-Annual'
                else:
                    frequency = 'Annual'

            pays_dividends = bool(dividend_rate or dividend_amount or ex_date_str)
            dividend_per_share = round(dividend_amount, 4) if dividend_amount else None
            total_dividend = round(quantity * dividend_amount, 2) if dividend_amount and quantity else None

            # Save to cache
            save_dividend_cache(ticker, ex_date_str, dividend_amount, pays_dividends)

            dividends_data.append({
                'ticker': ticker,
                'ex_dividend_date': ex_date_str,
                'payment_date': payment_date,
                'remaining_days': remaining_days,
                'dividend_amount': dividend_per_share,
                'dividend_yield': round(dividend_yield, 2) if dividend_yield else None,
                'frequency': frequency,
                'confirmed': confirmed,
                'pays_dividends': pays_dividends,
                'quantity': quantity,
                'total_dividend': total_dividend,
                'amount_source': amount_source,
            })
        except Exception as e:
            app.logger.warning(f"Failed to fetch dividend data for {ticker}: {e}")
            # Save to cache as non-payer to avoid repeated failures
            save_dividend_cache(ticker, None, None, False)
            continue

    # Sort: dividend payers with future dates first, then past dates, then non-payers
    def sort_key(x):
        days = x['remaining_days']
        pays = x.get('pays_dividends', True)
        if not pays:
            return (3, x['ticker'])
        elif days is None:
            return (1, 9999)
        elif days < 0:
            return (2, -days)
        else:
            return (0, days)

    dividends_data.sort(key=sort_key)

    return jsonify({'dividends': dividends_data})


@app.route('/api/investing/dividend-history/<ticker>', methods=['GET'])
@login_required
def get_dividend_history(ticker):
    """Get dividend payment history for a stock."""
    import yfinance as yf
    from investing_utils import get_yfinance_ticker

    try:
        yf_ticker = get_yfinance_ticker(ticker)
        stock = yf.Ticker(yf_ticker)

        # Get dividend history
        dividends = stock.dividends

        if dividends is None or dividends.empty:
            return jsonify({
                'ticker': ticker,
                'history': [],
                'growth_rates': {'1Y': None, '2Y': None, '5Y': None, '10Y': None}
            })

        # Convert to list of dicts with date and amount
        history = []
        for date, amount in dividends.items():
            history.append({
                'date': date.strftime('%Y-%m-%d'),
                'amount': round(float(amount), 4),
                'quarter': date.strftime('%b %Y')
            })

        # Calculate growth rates (comparing annual totals)
        def calc_annual_growth(years):
            if len(history) < 4:  # Need at least a year of data
                return None

            now = datetime.now()
            current_year_start = datetime(now.year - 1, now.month, now.day)
            past_year_start = datetime(now.year - 1 - years, now.month, now.day)
            past_year_end = datetime(now.year - years, now.month, now.day)

            current_total = sum(
                h['amount'] for h in history
                if datetime.strptime(h['date'], '%Y-%m-%d') >= current_year_start
            )
            past_total = sum(
                h['amount'] for h in history
                if past_year_start <= datetime.strptime(h['date'], '%Y-%m-%d') < past_year_end
            )

            if past_total == 0:
                return None
            return round(((current_total - past_total) / past_total) * 100, 2)

        growth_rates = {
            '1Y': calc_annual_growth(1),
            '2Y': calc_annual_growth(2),
            '5Y': calc_annual_growth(5),
            '10Y': calc_annual_growth(10),
        }

        return jsonify({
            'ticker': ticker,
            'history': history,
            'growth_rates': growth_rates
        })

    except Exception as e:
        app.logger.warning(f"Failed to fetch dividend history for {ticker}: {e}")
        return jsonify({
            'ticker': ticker,
            'history': [],
            'growth_rates': {'1Y': None, '2Y': None, '5Y': None, '10Y': None}
        })


@app.route('/api/investing/earnings-alerts', methods=['GET'])
@login_required
def get_earnings_alert_preferences():
    """Get current user's earnings alert preferences."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT weekly_enabled, days_before_enabled, days_before FROM earnings_alert_preferences WHERE user_id = ?',
            (request.user_id,)
        )
        row = cursor.fetchone()

        if row:
            return jsonify({
                'weekly_enabled': bool(row['weekly_enabled']),
                'days_before_enabled': bool(row['days_before_enabled']),
                'days_before': row['days_before']
            })
        else:
            # Return default values if no preferences set
            return jsonify({
                'weekly_enabled': False,
                'days_before_enabled': False,
                'days_before': 7
            })


@app.route('/api/investing/earnings-alerts', methods=['POST'])
@login_required
def save_earnings_alert_preferences():
    """Save or update earnings alert preferences."""
    data = request.get_json()

    weekly_enabled = data.get('weekly_enabled', False)
    days_before_enabled = data.get('days_before_enabled', False)
    days_before = data.get('days_before', 7)

    if days_before_enabled and (not isinstance(days_before, int) or days_before < 1 or days_before > 30):
        return jsonify({'error': 'days_before must be an integer between 1 and 30'}), 400

    with get_db() as conn:
        conn.execute('''
            INSERT INTO earnings_alert_preferences (user_id, weekly_enabled, days_before_enabled, days_before, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(user_id) DO UPDATE SET
                weekly_enabled = excluded.weekly_enabled,
                days_before_enabled = excluded.days_before_enabled,
                days_before = excluded.days_before,
                updated_at = CURRENT_TIMESTAMP
        ''', (request.user_id, 1 if weekly_enabled else 0, 1 if days_before_enabled else 0, days_before))

    return jsonify({'success': True, 'message': 'Alert preferences saved'})


@app.route('/api/investing/earnings-alerts', methods=['DELETE'])
@login_required
def disable_earnings_alerts():
    """Disable all earnings alerts for the current user."""
    with get_db() as conn:
        conn.execute(
            'UPDATE earnings_alert_preferences SET weekly_enabled = 0, days_before_enabled = 0, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?',
            (request.user_id,)
        )

    return jsonify({'success': True, 'message': 'Alerts disabled'})


@app.route('/api/investing/earnings-alerts/send-now', methods=['POST'])
@login_required
def send_earnings_alert_now():
    """Send an immediate test email with upcoming earnings."""
    from email_utils import send_earnings_alert_email
    import yfinance as yf
    from datetime import datetime

    # Get user info and alert preferences
    with get_db() as conn:
        cursor = conn.execute('SELECT email, name FROM users WHERE id = ?', (request.user_id,))
        user = cursor.fetchone()

        cursor = conn.execute('SELECT weekly_enabled, days_before_enabled, days_before FROM earnings_alert_preferences WHERE user_id = ?', (request.user_id,))
        prefs = cursor.fetchone()

    if not user:
        return jsonify({'error': 'User not found'}), 404

    email = user['email']
    name = user['name'] or 'Investor'

    # Build schedule info from preferences
    schedule_info = None
    if prefs:
        schedule_info = {
            'weekly_enabled': bool(prefs['weekly_enabled']),
            'days_before_enabled': bool(prefs['days_before_enabled']),
            'days_before': prefs['days_before']
        }

    # Get tickers from earnings_watchlist (bell-activated stocks only)
    portfolio_tickers = set()
    watchlist_tickers = set()
    alert_tickers = set()
    with get_db() as conn:
        # Get all bell-activated tickers
        cursor = conn.execute('SELECT stock_ticker FROM earnings_watchlist WHERE user_id = ?', (request.user_id,))
        for row in cursor.fetchall():
            alert_tickers.add(row['stock_ticker'])

        # Get portfolio tickers (for source info)
        cursor = conn.execute('SELECT DISTINCT stock_ticker FROM portfolio_transactions WHERE user_id = ?', (request.user_id,))
        for row in cursor.fetchall():
            portfolio_tickers.add(row['stock_ticker'])

        # Get watchlist tickers (for source info)
        cursor = conn.execute('SELECT stock_ticker FROM watchlist WHERE user_id = ?', (request.user_id,))
        for row in cursor.fetchall():
            watchlist_tickers.add(row['stock_ticker'])

    if not alert_tickers:
        return jsonify({'error': 'No stocks with alerts enabled. Click the bell icon next to stocks to enable alerts.'}), 400

    # Get earnings data
    today = datetime.now().date()
    earnings_data = []

    for ticker in alert_tickers:
        with get_db() as conn:
            cursor = conn.execute('SELECT next_earnings_date, date_confirmed FROM earnings_cache WHERE ticker = ?', (ticker,))
            row = cursor.fetchone()

            if row and row['next_earnings_date']:
                try:
                    earnings_date = datetime.strptime(row['next_earnings_date'], '%Y-%m-%d').date()
                    remaining_days = (earnings_date - today).days

                    if remaining_days >= 0:
                        # Get company name
                        try:
                            stock = yf.Ticker(ticker)
                            info = stock.info
                            company_name = info.get('longName') or info.get('shortName') or ticker
                        except Exception:
                            company_name = ticker

                        # Determine source
                        if ticker in portfolio_tickers:
                            source = 'portfolio'
                        elif ticker in watchlist_tickers:
                            source = 'watchlist'
                        else:
                            source = 'none'

                        earnings_data.append({
                            'ticker': ticker,
                            'company_name': company_name,
                            'next_earnings_date': row['next_earnings_date'],
                            'remaining_days': remaining_days,
                            'date_confirmed': bool(row['date_confirmed']),
                            'source': source
                        })
                except ValueError:
                    continue

    if not earnings_data:
        return jsonify({'error': 'No upcoming earnings found'}), 400

    # Sort by remaining days
    earnings_data.sort(key=lambda x: x['remaining_days'])

    # Send email with schedule info
    success = send_earnings_alert_email(email, name, earnings_data, 'test', schedule_info)

    if success:
        return jsonify({'success': True, 'message': f'Email sent to {email}'})
    else:
        return jsonify({'error': 'Failed to send email. Check SMTP configuration.'}), 500


def refresh_videos_for_ticker_async(ticker: str):
    """Refresh video selections for a ticker in the background."""
    def _refresh():
        try:
            api_key = os.environ.get('YOUTUBE_API_KEY')
            if api_key:
                from investing_utils import get_news_feed_videos
                get_news_feed_videos(get_db, api_key, ticker=ticker, limit=3)
                print(f"[Videos] Refreshed videos for {ticker}")
        except Exception as e:
            print(f"[Videos] Error refreshing videos for {ticker}: {e}")

    thread = threading.Thread(target=_refresh, daemon=True)
    thread.start()


@app.route('/api/investing/news-feed', methods=['GET'])
def get_news_feed():
    """Get YouTube news feed videos, optionally filtered by ticker and company name."""
    ticker = request.args.get('ticker')
    company_name = request.args.get('company_name')
    limit = request.args.get('limit', 50, type=int)
    force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'

    api_key = os.environ.get('YOUTUBE_API_KEY')

    try:
        result = get_news_feed_videos(get_db, api_key, ticker=ticker, company_name=company_name, limit=limit, force_refresh=force_refresh)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/investing/video-summary/<video_id>', methods=['GET'])
def get_video_summary(video_id):
    """Get transcript and/or summary of a YouTube video.

    Query params:
    - ticker: Required for getting company-specific summary
    - language: 'en' or 'fr' (default 'en')

    Transcripts/summaries are synced by a local cron job (sync_video_summaries.py)
    since YouTube blocks transcript requests from cloud IPs.
    """
    ticker = request.args.get('ticker')
    language = request.args.get('language', 'en')
    if language not in ('en', 'fr'):
        language = 'en'

    with get_db() as conn:
        # Get transcript (one per video)
        cursor = conn.execute(
            'SELECT transcript, has_transcript FROM video_transcripts WHERE video_id = ?',
            (video_id,)
        )
        transcript_row = cursor.fetchone()

        # Get summary (per video+ticker pair)
        summary_row = None
        if ticker:
            cursor = conn.execute(
                'SELECT summary_en, summary_fr FROM video_summaries WHERE video_id = ? AND ticker = ?',
                (video_id, ticker)
            )
            summary_row = cursor.fetchone()

        result = {}

        if transcript_row:
            result['transcript'] = transcript_row['transcript']
            result['has_transcript'] = bool(transcript_row['has_transcript'])

        if summary_row:
            # Return summary in requested language, fall back to English
            summary_col = f'summary_{language}'
            result['summary'] = summary_row.get(summary_col) or summary_row.get('summary_en')

        if result:
            return jsonify(result)

    # No data yet - check if sync is running
    with get_db() as conn:
        if USE_POSTGRES:
            cursor = conn.execute('''
                SELECT current_step FROM video_sync_runs
                WHERE status = 'running'
                AND updated_at > CURRENT_TIMESTAMP - INTERVAL '2 minutes'
                ORDER BY started_at DESC LIMIT 1
            ''')
        else:
            cursor = conn.execute('''
                SELECT current_step FROM video_sync_runs
                WHERE status = 'running'
                AND updated_at > datetime('now', '-2 minutes')
                ORDER BY started_at DESC LIMIT 1
            ''')
        running = cursor.fetchone()

    if running:
        step = running.get('current_step', 'downloading')
        return jsonify({'pending': True, 'sync_running': True, 'sync_step': step}), 202

    return jsonify({'pending': True, 'sync_running': False}), 202


@app.route('/api/investing/video-summaries/pending', methods=['GET'])
def get_videos_pending_sync():
    """Get videos that need transcript/summary sync. Used by local sync script.

    Only returns videos that are in a company's current selection (company_video_selections table).
    This ensures we only sync transcripts for videos currently being displayed.

    Returns (video_id, ticker) pairs that either:
    - Have no transcript yet (and hasn't been marked as unavailable)
    - Have transcript but no summary for that specific ticker

    Each video can appear multiple times if it's selected for multiple tickers.
    """
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    ticker_filter = request.args.get('ticker')

    with get_db() as conn:
        if ticker_filter:
            # Filter by ticker - return (video_id, ticker) pairs needing sync
            # A video needs sync if: no transcript OR (has transcript but missing EN or FR summary)
            cursor = conn.execute('''
                SELECT v.video_id, sel.ticker, v.title, v.channel_name, v.published_at, v.duration,
                       t.transcript IS NOT NULL as has_transcript,
                       (s.summary_en IS NOT NULL AND s.summary_fr IS NOT NULL) as has_summary
                FROM company_video_selections sel
                JOIN youtube_videos_cache v ON sel.video_id = v.video_id
                LEFT JOIN video_transcripts t ON v.video_id = t.video_id
                LEFT JOIN video_summaries s ON v.video_id = s.video_id AND sel.ticker = s.ticker
                WHERE sel.ticker = ?
                  AND ((t.video_id IS NULL)
                       OR (t.has_transcript = 1 AND t.transcript IS NOT NULL
                           AND (s.video_id IS NULL OR s.summary_en IS NULL OR s.summary_fr IS NULL)))
                ORDER BY v.published_at DESC
            ''', (ticker_filter,))
        else:
            # Get all pending (video_id, ticker) pairs
            cursor = conn.execute('''
                SELECT v.video_id, sel.ticker, v.title, v.channel_name, v.published_at, v.duration,
                       t.transcript IS NOT NULL as has_transcript,
                       (s.summary_en IS NOT NULL AND s.summary_fr IS NOT NULL) as has_summary
                FROM company_video_selections sel
                JOIN youtube_videos_cache v ON sel.video_id = v.video_id
                LEFT JOIN video_transcripts t ON v.video_id = t.video_id
                LEFT JOIN video_summaries s ON v.video_id = s.video_id AND sel.ticker = s.ticker
                WHERE (t.video_id IS NULL)
                   OR (t.has_transcript = 1 AND t.transcript IS NOT NULL
                       AND (s.video_id IS NULL OR s.summary_en IS NULL OR s.summary_fr IS NULL))
                ORDER BY v.published_at DESC
            ''')
        videos = cursor.fetchall()

    return jsonify({'videos': [dict(v) for v in videos]})


@app.route('/api/investing/sync/tickers-to-sync', methods=['GET'])
def get_tickers_to_sync():
    """Get all unique tickers from all users' portfolios and watchlists.

    Used by the sync script to know which companies need video updates.
    This list updates automatically when users add/remove stocks.
    """
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    with get_db() as conn:
        # Get tickers from all portfolios
        cursor = conn.execute('''
            SELECT DISTINCT stock_ticker FROM portfolio_transactions
        ''')
        portfolio_tickers = set(row['stock_ticker'] for row in cursor.fetchall())

        # Get tickers from all watchlists
        cursor = conn.execute('''
            SELECT DISTINCT stock_ticker FROM watchlist
        ''')
        watchlist_tickers = set(row['stock_ticker'] for row in cursor.fetchall())

        # Combine and sort
        all_tickers = sorted(portfolio_tickers | watchlist_tickers)

    return jsonify({
        'tickers': all_tickers,
        'count': len(all_tickers),
        'portfolio_count': len(portfolio_tickers),
        'watchlist_count': len(watchlist_tickers)
    })


@app.route('/api/investing/sync/clear-video-cache', methods=['POST'])
def sync_clear_video_cache():
    """Clear all cached videos, transcripts, summaries, and selections.

    Used to reset the video sync state.
    """
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    with get_db() as conn:
        conn.execute('DELETE FROM company_video_selections')
        conn.execute('DELETE FROM video_transcripts')
        conn.execute('DELETE FROM video_summaries')
        conn.execute('DELETE FROM youtube_videos_cache')
        conn.execute('DELETE FROM youtube_channel_fetch_log')

    return jsonify({'status': 'ok', 'message': 'Video cache cleared'})


@app.route('/api/investing/sync/start', methods=['POST'])
def sync_start():
    """Start a new sync run and return its ID."""
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json() or {}
    tickers_count = data.get('tickers_count', 0)

    with get_db() as conn:
        # Mark any existing "running" runs as interrupted
        conn.execute('''
            UPDATE video_sync_runs
            SET status = 'interrupted', ended_at = CURRENT_TIMESTAMP
            WHERE status = 'running'
        ''')

        if USE_POSTGRES:
            cursor = conn.execute('''
                INSERT INTO video_sync_runs (tickers_count, status, updated_at)
                VALUES (%s, 'running', CURRENT_TIMESTAMP)
                RETURNING id
            ''', (tickers_count,))
        else:
            cursor = conn.execute('''
                INSERT INTO video_sync_runs (tickers_count, status, updated_at)
                VALUES (?, 'running', CURRENT_TIMESTAMP)
            ''', (tickers_count,))
            cursor = conn.execute('SELECT last_insert_rowid() as id')
        run_id = cursor.fetchone()['id']

    return jsonify({'run_id': run_id})


@app.route('/api/investing/sync/update/<int:run_id>', methods=['POST'])
def sync_update(run_id):
    """Update sync run progress."""
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json() or {}

    with get_db() as conn:
        updates = []
        values = []
        for field in ['videos_total', 'videos_processed', 'current_video', 'current_step', 'videos_list']:
            if field in data:
                if USE_POSTGRES:
                    updates.append(f"{field} = %s")
                else:
                    updates.append(f"{field} = ?")
                values.append(data[field])

        # Always update updated_at for heartbeat tracking
        if USE_POSTGRES:
            updates.append("updated_at = CURRENT_TIMESTAMP")
        else:
            updates.append("updated_at = CURRENT_TIMESTAMP")

        if updates:
            values.append(run_id)
            placeholder = '%s' if USE_POSTGRES else '?'
            conn.execute(
                f"UPDATE video_sync_runs SET {', '.join(updates)} WHERE id = {placeholder}",
                values
            )

    return jsonify({'status': 'ok'})


@app.route('/api/investing/sync/end/<int:run_id>', methods=['POST'])
def sync_end(run_id):
    """Mark sync run as complete."""
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json() or {}

    with get_db() as conn:
        if USE_POSTGRES:
            conn.execute('''
                UPDATE video_sync_runs
                SET status = %s, ended_at = CURRENT_TIMESTAMP,
                    transcripts_fetched = %s, summaries_generated = %s,
                    errors = %s, error_message = %s, current_video = NULL
                WHERE id = %s
            ''', (data.get('status', 'completed'), data.get('transcripts_fetched', 0),
                  data.get('summaries_generated', 0), data.get('errors', 0),
                  data.get('error_message'), run_id))
        else:
            conn.execute('''
                UPDATE video_sync_runs
                SET status = ?, ended_at = CURRENT_TIMESTAMP,
                    transcripts_fetched = ?, summaries_generated = ?,
                    errors = ?, error_message = ?, current_video = NULL
                WHERE id = ?
            ''', (data.get('status', 'completed'), data.get('transcripts_fetched', 0),
                  data.get('summaries_generated', 0), data.get('errors', 0),
                  data.get('error_message'), run_id))

    return jsonify({'status': 'ok'})


@app.route('/api/admin/sync-status', methods=['GET'])
@admin_required
def get_sync_status():
    """Get current and recent sync run status (admin only)."""
    with get_db() as conn:
        # Mark stale runs (running but no heartbeat in 2+ minutes)
        if USE_POSTGRES:
            conn.execute('''
                UPDATE video_sync_runs
                SET status = 'stale', ended_at = CURRENT_TIMESTAMP,
                    error_message = 'No heartbeat received (sync script likely crashed)'
                WHERE status = 'running'
                AND updated_at < CURRENT_TIMESTAMP - INTERVAL '2 minutes'
            ''')
        else:
            conn.execute('''
                UPDATE video_sync_runs
                SET status = 'stale', ended_at = CURRENT_TIMESTAMP,
                    error_message = 'No heartbeat received (sync script likely crashed)'
                WHERE status = 'running'
                AND updated_at < datetime('now', '-2 minutes')
            ''')

        # Get current/most recent runs
        cursor = conn.execute('''
            SELECT * FROM video_sync_runs
            ORDER BY started_at DESC
            LIMIT 5
        ''')
        runs = [dict(row) for row in cursor.fetchall()]

    return jsonify({'runs': runs})


@app.route('/api/admin/sync-run/<int:run_id>', methods=['DELETE'])
@admin_required
def delete_sync_run(run_id):
    """Delete or stop a sync run (admin only)."""
    with get_db() as conn:
        # Check if it's running - if so, mark as interrupted instead of deleting
        if USE_POSTGRES:
            cursor = conn.execute('SELECT status FROM video_sync_runs WHERE id = %s', (run_id,))
        else:
            cursor = conn.execute('SELECT status FROM video_sync_runs WHERE id = ?', (run_id,))
        row = cursor.fetchone()

        if not row:
            return jsonify({'error': 'Run not found'}), 404

        if row['status'] == 'running':
            # Mark as interrupted (the script will check this and stop)
            if USE_POSTGRES:
                conn.execute('''
                    UPDATE video_sync_runs
                    SET status = 'interrupted', ended_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                ''', (run_id,))
            else:
                conn.execute('''
                    UPDATE video_sync_runs
                    SET status = 'interrupted', ended_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (run_id,))
            return jsonify({'success': True, 'action': 'interrupted'})
        else:
            # Delete the record
            if USE_POSTGRES:
                conn.execute('DELETE FROM video_sync_runs WHERE id = %s', (run_id,))
            else:
                conn.execute('DELETE FROM video_sync_runs WHERE id = ?', (run_id,))
            return jsonify({'success': True, 'action': 'deleted'})


@app.route('/api/investing/video-summaries/upload', methods=['POST'])
def upload_video_data():
    """Upload transcript and/or summary. Used by local sync script.

    Summaries are stored per (video_id, ticker) pair since the same video
    can have different summaries focused on different companies.
    """
    sync_key = request.headers.get('X-Sync-Key')
    expected_key = os.environ.get('SYNC_API_KEY', 'lumna-sync-2024')
    if sync_key != expected_key:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    video_id = data.get('video_id')
    ticker = data.get('ticker')  # Required for summaries
    transcript = data.get('transcript')  # Can be None if no transcript available
    has_transcript = data.get('has_transcript', True)  # False if video has no transcript
    summary_en = data.get('summary_en')
    summary_fr = data.get('summary_fr')

    if not video_id:
        return jsonify({'error': 'video_id required'}), 400

    if (summary_en or summary_fr) and not ticker:
        return jsonify({'error': 'ticker required when uploading summary'}), 400

    with get_db() as conn:
        # Save transcript if provided (or mark as unavailable) - one per video
        if transcript is not None or not has_transcript:
            if USE_POSTGRES:
                conn.execute(
                    '''INSERT INTO video_transcripts (video_id, transcript, has_transcript, created_at)
                       VALUES (?, ?, ?, NOW())
                       ON CONFLICT (video_id) DO UPDATE SET transcript = EXCLUDED.transcript, has_transcript = EXCLUDED.has_transcript''',
                    (video_id, transcript, 1 if has_transcript else 0)
                )
            else:
                conn.execute(
                    'INSERT OR REPLACE INTO video_transcripts (video_id, transcript, has_transcript) VALUES (?, ?, ?)',
                    (video_id, transcript, 1 if has_transcript else 0)
                )

        # Save summaries if provided - one per (video_id, ticker) pair
        if summary_en or summary_fr:
            if USE_POSTGRES:
                conn.execute(
                    '''INSERT INTO video_summaries (video_id, ticker, summary_en, summary_fr, created_at)
                       VALUES (?, ?, ?, ?, NOW())
                       ON CONFLICT (video_id, ticker) DO UPDATE SET
                           summary_en = COALESCE(EXCLUDED.summary_en, video_summaries.summary_en),
                           summary_fr = COALESCE(EXCLUDED.summary_fr, video_summaries.summary_fr)''',
                    (video_id, ticker, summary_en, summary_fr)
                )
            else:
                # For SQLite, first check if row exists
                cursor = conn.execute(
                    'SELECT summary_en, summary_fr FROM video_summaries WHERE video_id = ? AND ticker = ?',
                    (video_id, ticker)
                )
                existing = cursor.fetchone()
                if existing:
                    # Update, keeping existing values if new ones are None
                    conn.execute(
                        '''UPDATE video_summaries SET
                           summary_en = COALESCE(?, summary_en),
                           summary_fr = COALESCE(?, summary_fr)
                           WHERE video_id = ? AND ticker = ?''',
                        (summary_en, summary_fr, video_id, ticker)
                    )
                else:
                    conn.execute(
                        'INSERT INTO video_summaries (video_id, ticker, summary_en, summary_fr) VALUES (?, ?, ?, ?)',
                        (video_id, ticker, summary_en, summary_fr)
                    )

    return jsonify({'success': True})


@app.route('/api/investing/graph-download', methods=['POST'])
@login_required
def record_graph_download():
    """Record a graph download event."""
    data = request.get_json()
    graph_type = data.get('graph_type')

    if not graph_type:
        return jsonify({'error': 'graph_type is required'}), 400

    with get_db() as conn:
        conn.execute(
            'INSERT INTO graph_downloads (user_id, graph_type) VALUES (?, ?)',
            (request.user_id, graph_type)
        )
        conn.commit()

    return jsonify({'success': True})


@app.route('/api/investing/stock-view', methods=['POST'])
@login_required
def record_stock_view():
    """Record a stock view event (view count and time spent)."""
    data = request.get_json()
    ticker = data.get('ticker')
    time_spent = data.get('time_spent_seconds', 0)

    if not ticker:
        return jsonify({'error': 'ticker is required'}), 400

    from datetime import date
    today = date.today().isoformat()

    with get_db() as conn:
        # Upsert: increment view count and add time spent
        conn.execute('''
            INSERT INTO stock_views (user_id, stock_ticker, view_date, view_count, time_spent_seconds)
            VALUES (?, ?, ?, 1, ?)
            ON CONFLICT(user_id, stock_ticker, view_date) DO UPDATE SET
                view_count = stock_views.view_count + 1,
                time_spent_seconds = stock_views.time_spent_seconds + excluded.time_spent_seconds,
                last_viewed_at = CURRENT_TIMESTAMP
        ''', (request.user_id, ticker, today, time_spent))
        conn.commit()

    return jsonify({'success': True})


@app.route('/api/admin/users/<int:user_id>/stock-views', methods=['GET'])
@admin_required
def get_user_stock_views(user_id):
    """Get stock view statistics for a specific user (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT stock_ticker,
                   SUM(view_count) as total_views,
                   SUM(time_spent_seconds) as total_time_seconds
            FROM stock_views
            WHERE user_id = ?
            GROUP BY stock_ticker
            ORDER BY total_views DESC
        ''', (user_id,))
        views = [dict(row) for row in cursor.fetchall()]

    return jsonify({'views': views})


@app.route('/api/admin/stock-views', methods=['GET'])
@admin_required
def get_stock_views_stats():
    """Get aggregated stock view statistics (admin only)."""
    excluded_email = 'rose.louis.mail@gmail.com'

    with get_db() as conn:
        # Get aggregated stats by stock (exclude admin)
        cursor = conn.execute('''
            SELECT sv.stock_ticker,
                   COUNT(DISTINCT sv.user_id) as unique_users,
                   SUM(sv.view_count) as total_views,
                   SUM(sv.time_spent_seconds) as total_time_seconds
            FROM stock_views sv
            JOIN users u ON sv.user_id = u.id
            WHERE u.email != ?
            GROUP BY sv.stock_ticker
            ORDER BY total_views DESC
        ''', (excluded_email,))
        by_stock = [dict(row) for row in cursor.fetchall()]

        # Get stats by user (exclude admin)
        cursor = conn.execute('''
            SELECT u.id, u.name, u.email,
                   COUNT(DISTINCT sv.stock_ticker) as stocks_viewed,
                   SUM(sv.view_count) as total_views,
                   SUM(sv.time_spent_seconds) as total_time_seconds
            FROM stock_views sv
            JOIN users u ON sv.user_id = u.id
            WHERE u.email != ?
            GROUP BY u.id
            ORDER BY total_views DESC
        ''', (excluded_email,))
        by_user = [dict(row) for row in cursor.fetchall()]

    return jsonify({
        'by_stock': by_stock,
        'by_user': by_user
    })


@app.route('/api/admin/company-stats', methods=['GET'])
@admin_required
def get_company_stats():
    """Get company popularity stats - how many portfolios/watchlists each ticker appears in."""
    with get_db() as conn:
        # Get portfolio counts by ticker
        cursor = conn.execute('''
            SELECT pt.stock_ticker as ticker, COUNT(DISTINCT pt.user_id) as portfolio_count
            FROM portfolio_transactions pt
            GROUP BY pt.stock_ticker
        ''')
        portfolio_counts = {row['ticker']: row['portfolio_count'] for row in cursor.fetchall()}

        # Get watchlist counts by ticker
        cursor = conn.execute('''
            SELECT w.stock_ticker as ticker, COUNT(DISTINCT w.user_id) as watchlist_count
            FROM watchlist w
            GROUP BY w.stock_ticker
        ''')
        watchlist_counts = {row['ticker']: row['watchlist_count'] for row in cursor.fetchall()}

        # Combine all tickers
        all_tickers = set(portfolio_counts.keys()) | set(watchlist_counts.keys())
        companies = []
        for ticker in all_tickers:
            p_count = portfolio_counts.get(ticker, 0)
            w_count = watchlist_counts.get(ticker, 0)
            companies.append({
                'ticker': ticker,
                'portfolio_count': p_count,
                'watchlist_count': w_count,
                'total': p_count + w_count
            })

        # Sort by total (portfolio + watchlist) descending
        companies.sort(key=lambda x: x['total'], reverse=True)

    return jsonify({'companies': companies})


@app.route('/api/admin/company-stats/<ticker>', methods=['GET'])
@admin_required
def get_company_users(ticker):
    """Get users who have this ticker in their portfolio or watchlist."""
    with get_db() as conn:
        # Get users with this ticker in portfolio
        cursor = conn.execute('''
            SELECT DISTINCT u.id, u.name, u.picture
            FROM users u
            JOIN portfolio_transactions pt ON u.id = pt.user_id
            WHERE pt.stock_ticker = ?
            ORDER BY u.name
        ''', (ticker,))
        portfolio_users = [dict(row) for row in cursor.fetchall()]

        # Get users with this ticker in watchlist
        cursor = conn.execute('''
            SELECT DISTINCT u.id, u.name, u.picture
            FROM users u
            JOIN watchlist w ON u.id = w.user_id
            WHERE w.stock_ticker = ?
            ORDER BY u.name
        ''', (ticker,))
        watchlist_users = [dict(row) for row in cursor.fetchall()]

    return jsonify({
        'ticker': ticker,
        'portfolio_users': portfolio_users,
        'watchlist_users': watchlist_users
    })


@app.route('/api/admin/stock-views/<ticker>', methods=['GET'])
@admin_required
def get_stock_views_detail(ticker):
    """Get detailed view statistics for a specific stock (admin only)."""
    with get_db() as conn:
        # Get all views for this stock by user
        cursor = conn.execute('''
            SELECT u.id, u.name, u.picture,
                   sv.view_date,
                   sv.view_count,
                   sv.time_spent_seconds,
                   sv.last_viewed_at
            FROM stock_views sv
            JOIN users u ON sv.user_id = u.id
            WHERE sv.stock_ticker = ?
            ORDER BY sv.view_date DESC
        ''', (ticker,))
        views = [dict(row) for row in cursor.fetchall()]

        # Get totals
        cursor = conn.execute('''
            SELECT COUNT(DISTINCT sv.user_id) as unique_users,
                   SUM(sv.view_count) as total_views,
                   SUM(sv.time_spent_seconds) as total_time_seconds
            FROM stock_views sv
            WHERE sv.stock_ticker = ?
        ''', (ticker,))
        totals = dict(cursor.fetchone())

    return jsonify({
        'ticker': ticker,
        'views': views,
        'totals': totals
    })


@app.route('/api/admin/time-spent', methods=['GET'])
@admin_required
def get_time_spent_stats():
    """Get daily time spent stats for all users (admin only)."""
    excluded_email = 'rose.louis.mail@gmail.com'

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT a.activity_date, SUM(a.minutes) as total_minutes
            FROM user_activity a
            JOIN users u ON a.user_id = u.id
            WHERE u.email != ?
            GROUP BY a.activity_date
            ORDER BY a.activity_date ASC
        ''', (excluded_email,))
        daily_stats = [dict(row) for row in cursor.fetchall()]

    return jsonify({'daily_stats': daily_stats})


@app.route('/api/admin/time-spent/<period>', methods=['GET'])
@admin_required
def get_time_spent_details(period):
    """Get users' time spent for a specific date, week, or month (admin only).

    Period formats:
    - Date: YYYY-MM-DD
    - Week: YYYY-WXX
    - Month: YYYY-MM
    """
    excluded_email = 'rose.louis.mail@gmail.com'

    with get_db() as conn:
        if '-W' in period:
            # Week format: YYYY-WXX
            year, week = period.split('-W')
            if USE_POSTGRES:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, SUM(a.minutes) as minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE EXTRACT(YEAR FROM a.activity_date::date) = %s
                      AND EXTRACT(WEEK FROM a.activity_date::date) = %s
                      AND u.email != %s
                    GROUP BY u.id, u.name, u.picture
                    ORDER BY minutes DESC
                ''', (int(year), int(week), excluded_email))
            else:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, SUM(a.minutes) as minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE strftime('%Y', a.activity_date) = ?
                      AND CAST(strftime('%W', a.activity_date) AS INTEGER) + 1 = ?
                      AND u.email != ?
                    GROUP BY u.id
                    ORDER BY minutes DESC
                ''', (year, int(week), excluded_email))
        elif len(period) == 7:
            # Month format: YYYY-MM
            if USE_POSTGRES:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, SUM(a.minutes) as minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE to_char(a.activity_date::date, 'YYYY-MM') = %s
                      AND u.email != %s
                    GROUP BY u.id, u.name, u.picture
                    ORDER BY minutes DESC
                ''', (period, excluded_email))
            else:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, SUM(a.minutes) as minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE strftime('%Y-%m', a.activity_date) = ?
                      AND u.email != ?
                    GROUP BY u.id
                    ORDER BY minutes DESC
                ''', (period, excluded_email))
        else:
            # Date format: YYYY-MM-DD
            if USE_POSTGRES:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, a.minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE a.activity_date = %s
                      AND u.email != %s
                    ORDER BY a.minutes DESC
                ''', (period, excluded_email))
            else:
                cursor = conn.execute('''
                    SELECT u.id, u.name, u.picture, a.minutes
                    FROM user_activity a
                    JOIN users u ON a.user_id = u.id
                    WHERE a.activity_date = ?
                      AND u.email != ?
                    ORDER BY a.minutes DESC
                ''', (period, excluded_email))

        users = [dict(row) for row in cursor.fetchall()]

    return jsonify({'users': users, 'period': period})


@app.route('/api/admin/page-breakdown', methods=['GET'])
@admin_required
def get_page_breakdown():
    """Get aggregated time spent by page/section (admin only)."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT p.page, SUM(p.minutes) as total_minutes
            FROM page_activity p
            GROUP BY p.page
            ORDER BY total_minutes DESC
        ''')

        breakdown = [dict(row) for row in cursor.fetchall()]
        total = sum(item['total_minutes'] for item in breakdown)

    return jsonify({
        'breakdown': breakdown,
        'total_minutes': total
    })


@app.route('/api/admin/clear-video-cache', methods=['POST'])
@admin_required
def clear_video_cache():
    """Clear YouTube video cache, channel fetch log, and company selections (admin only)."""
    with get_db() as conn:
        conn.execute('DELETE FROM youtube_videos_cache')
        conn.execute('DELETE FROM youtube_channel_fetch_log')
        conn.execute('DELETE FROM company_video_selections')
    return jsonify({'success': True, 'message': 'Video cache, fetch log, and company selections cleared.'})


@app.route('/api/admin/backfill-demo-portfolios', methods=['POST'])
@admin_required
def backfill_demo_portfolios():
    """Add demo portfolios to users (admin only). Use force=true to add to ALL users."""
    from auth import create_demo_portfolio

    data = request.get_json() or {}
    force = data.get('force', False)

    with get_db() as conn:
        if force:
            # Add to ALL users
            cursor = conn.execute('SELECT id, email FROM users')
        else:
            # Find users without any transactions
            cursor = conn.execute('''
                SELECT u.id, u.email
                FROM users u
                LEFT JOIN portfolio_transactions pt ON u.id = pt.user_id
                GROUP BY u.id
                HAVING COUNT(pt.id) = 0
            ''')
        users = cursor.fetchall()

        backfilled_count = 0
        for user in users:
            if create_demo_portfolio(user['id'], force=force):
                backfilled_count += 1

    return jsonify({
        'success': True,
        'message': f'Demo portfolios created for {backfilled_count} users',
        'users_backfilled': backfilled_count,
        'force': force
    })


@app.route('/api/admin/remove-demo-portfolios', methods=['POST'])
@admin_required
def remove_demo_portfolios():
    """Remove demo portfolios from all users (admin only)."""
    with get_db() as conn:
        # First, delete all transactions linked to demo portfolio accounts
        cursor = conn.execute('''
            DELETE FROM portfolio_transactions
            WHERE account_id IN (
                SELECT id FROM investment_accounts WHERE name = 'Demo Portfolio'
            )
        ''')
        transactions_deleted = cursor.rowcount

        # Then delete the demo portfolio accounts
        cursor = conn.execute('''
            DELETE FROM investment_accounts WHERE name = 'Demo Portfolio'
        ''')
        accounts_deleted = cursor.rowcount

    return jsonify({
        'success': True,
        'message': f'Removed {accounts_deleted} demo portfolios ({transactions_deleted} transactions)',
        'accounts_deleted': accounts_deleted,
        'transactions_deleted': transactions_deleted
    })


@app.route('/api/admin/delete-user', methods=['POST'])
@admin_required
def delete_user():
    """Delete a user and all their data by email (admin only)."""
    data = request.get_json() or {}
    email = data.get('email')

    if not email:
        return jsonify({'error': 'Email required'}), 400

    with get_db() as conn:
        # Get user ID
        cursor = conn.execute('SELECT id FROM users WHERE email = ?', (email,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': f'User {email} not found'}), 404

        user_id = row['id']

        # Delete all related data
        conn.execute('DELETE FROM user_activity WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM page_activity WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM stock_views WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM portfolio_transactions WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM watchlist WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM investment_accounts WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM user_preferences WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM graph_downloads WHERE user_id = ?', (user_id,))
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))

    return jsonify({
        'success': True,
        'message': f'Deleted user {email} (id={user_id})'
    })


@app.route('/api/admin/cleanup-orphaned-transactions', methods=['POST'])
@admin_required
def cleanup_orphaned_transactions():
    """Delete portfolio transactions whose investment account no longer exists."""
    with get_db() as conn:
        # Find orphaned transactions (account_id not in investment_accounts)
        cursor = conn.execute('''
            SELECT pt.id, pt.user_id, pt.account_id, pt.stock_ticker, u.name as user_name
            FROM portfolio_transactions pt
            LEFT JOIN investment_accounts ia ON pt.account_id = ia.id
            LEFT JOIN users u ON pt.user_id = u.id
            WHERE ia.id IS NULL
        ''')
        orphaned = [dict(row) for row in cursor.fetchall()]

        if orphaned:
            # Delete orphaned transactions
            cursor = conn.execute('''
                DELETE FROM portfolio_transactions
                WHERE account_id NOT IN (SELECT id FROM investment_accounts)
                   OR account_id IS NULL
            ''')
            deleted_count = cursor.rowcount
        else:
            deleted_count = 0

    return jsonify({
        'success': True,
        'orphaned_found': len(orphaned),
        'deleted': deleted_count,
        'details': orphaned[:20]  # Show first 20 for reference
    })


@app.route('/api/reward/eligibility', methods=['GET'])
@login_required
def check_reward_eligibility():
    """Check if user is eligible for the 5th visit reward (5+ sessions, not yet claimed by this user)."""
    with get_db() as conn:
        # Check if this user has already claimed the reward
        cursor = conn.execute('SELECT id FROM first_visitor_reward WHERE user_id = ?', (request.user_id,))
        if cursor.fetchone():
            return jsonify({'eligible': False, 'reason': 'already_claimed'})

        # Check if current user has 5+ sessions
        cursor = conn.execute('SELECT session_count FROM users WHERE id = ?', (request.user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'eligible': False, 'reason': 'user_not_found'})

        if user['session_count'] >= 5:
            return jsonify({'eligible': True})
        else:
            return jsonify({'eligible': False, 'reason': 'not_enough_sessions', 'current_sessions': user['session_count']})


@app.route('/api/reward/claim', methods=['POST'])
@login_required
def claim_reward():
    """Claim the 5th visit reward by selecting a company for analysis."""
    from email_utils import send_reward_notification_email

    data = request.get_json()
    selected_company = data.get('company', '').strip().upper()

    if not selected_company:
        return jsonify({'error': 'Company ticker is required'}), 400

    with get_db() as conn:
        # Check if this user already claimed
        cursor = conn.execute('SELECT id FROM first_visitor_reward WHERE user_id = ?', (request.user_id,))
        if cursor.fetchone():
            return jsonify({'error': 'You have already claimed this reward'}), 409

        # Check user has 5+ sessions
        cursor = conn.execute('SELECT session_count, name, email FROM users WHERE id = ?', (request.user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        if user['session_count'] < 5:
            return jsonify({'error': 'Not enough sessions to claim reward'}), 403

        user_name = user['name'] or 'Unknown User'
        user_email = user['email']

        # Claim the reward
        conn.execute('''
            INSERT INTO first_visitor_reward (user_id, user_name, user_email, selected_company)
            VALUES (?, ?, ?, ?)
        ''', (request.user_id, user_name, user_email, selected_company))

        # Send notification email
        send_reward_notification_email(user_name, user_email, selected_company)

    return jsonify({'success': True, 'message': 'Reward claimed successfully!'})


@app.route('/api/feedback', methods=['POST'])
@login_required
def submit_feedback():
    """Submit user feedback via email, with optional screenshot attachments."""
    from email_utils import send_feedback_email

    data = request.get_json()
    message = data.get('message', '').strip()
    images = data.get('images', [])  # List of base64 data URIs

    if not message and not images:
        return jsonify({'error': 'Message or screenshot is required'}), 400

    if len(message) > 5000:
        return jsonify({'error': 'Message too long (max 5000 characters)'}), 400

    # Limit number of images
    if len(images) > 5:
        return jsonify({'error': 'Too many images (max 5)'}), 400

    # Get user info
    with get_db() as conn:
        cursor = conn.execute('SELECT name, email FROM users WHERE id = ?', (request.user_id,))
        user = cursor.fetchone()

    if not user:
        return jsonify({'error': 'User not found'}), 404

    user_name = user['name'] or 'Unknown User'
    user_email = user['email']

    success = send_feedback_email(user_name, user_email, message, images)

    if success:
        return jsonify({'success': True, 'message': 'Feedback sent successfully'})
    else:
        return jsonify({'error': 'Failed to send feedback. Please try again later.'}), 500


# =============================================================================
# DEMO ALPHAWISE API ROUTES
# These routes use separate demo_* tables for the demo app
# =============================================================================

@app.route('/api/demo/accounts', methods=['GET'])
@login_required
def get_demo_accounts():
    """Get user's demo investment accounts."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT id, name, account_type, bank, display_order, created_at
               FROM demo_investment_accounts WHERE user_id = ?
               ORDER BY display_order ASC, created_at ASC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    accounts = [{
        'id': row['id'],
        'name': row['name'],
        'account_type': row['account_type'],
        'bank': row['bank'],
        'bank_info': BANKS.get(row['bank'], {}),
        'type_info': ACCOUNT_TYPES.get(row['account_type'], {}),
    } for row in rows]
    return jsonify({'accounts': accounts})


@app.route('/api/demo/accounts', methods=['POST'])
@login_required
def create_demo_account():
    """Create a new demo investment account."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    name = data.get('name', '').strip()
    account_type = data.get('account_type', '').upper().strip()
    bank = data.get('bank', '').upper().strip()

    if not name:
        return jsonify({'error': 'Account name required'}), 400
    if account_type not in ACCOUNT_TYPES:
        return jsonify({'error': f'Invalid account type. Valid: {list(ACCOUNT_TYPES.keys())}'}), 400
    if bank not in BANKS:
        return jsonify({'error': f'Invalid bank. Valid: {list(BANKS.keys())}'}), 400

    with get_db() as conn:
        cursor = conn.execute(
            'SELECT COALESCE(MAX(display_order), -1) + 1 as next_order FROM demo_investment_accounts WHERE user_id = ?',
            (request.user_id,)
        )
        next_order = cursor.fetchone()['next_order']

        cursor = conn.execute('''
            INSERT INTO demo_investment_accounts (user_id, name, account_type, bank, display_order)
            VALUES (?, ?, ?, ?, ?)
            RETURNING id
        ''', (request.user_id, name, account_type, bank, next_order))
        account_id = cursor.fetchone()['id']

    return jsonify({
        'success': True,
        'id': account_id,
        'name': name,
        'account_type': account_type,
        'bank': bank,
        'bank_info': BANKS[bank],
        'type_info': ACCOUNT_TYPES[account_type],
    })


@app.route('/api/demo/accounts/<int:account_id>', methods=['DELETE'])
@login_required
def delete_demo_account(account_id):
    """Delete a demo investment account."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT id FROM demo_investment_accounts WHERE user_id = ? AND id = ?',
            (request.user_id, account_id)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Account not found'}), 404

        conn.execute('DELETE FROM demo_investment_accounts WHERE id = ? AND user_id = ?',
                    (account_id, request.user_id))

    return jsonify({'success': True})


@app.route('/api/demo/accounts/<int:account_id>/rename', methods=['PUT'])
@login_required
def rename_demo_account(account_id):
    """Rename a demo investment account."""
    data = request.get_json()
    new_name = data.get('name', '').strip() if data else ''

    if not new_name:
        return jsonify({'error': 'New name required'}), 400

    with get_db() as conn:
        cursor = conn.execute(
            'SELECT id FROM demo_investment_accounts WHERE id = ? AND user_id = ?',
            (account_id, request.user_id)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Account not found'}), 404

        conn.execute('UPDATE demo_investment_accounts SET name = ? WHERE id = ? AND user_id = ?',
                    (new_name, account_id, request.user_id))

    return jsonify({'success': True, 'name': new_name})


@app.route('/api/demo/accounts/reorder', methods=['PUT'])
@login_required
def reorder_demo_accounts():
    """Reorder demo investment accounts."""
    data = request.get_json()
    if not data or 'order' not in data:
        return jsonify({'error': 'Order array required'}), 400

    order = data['order']  # List of account IDs in desired order

    with get_db() as conn:
        for idx, account_id in enumerate(order):
            conn.execute(
                'UPDATE demo_investment_accounts SET display_order = ? WHERE id = ? AND user_id = ?',
                (idx, account_id, request.user_id)
            )

    return jsonify({'success': True})


@app.route('/api/demo/transactions', methods=['GET'])
@login_required
def get_demo_transactions():
    """Get user's demo transaction history."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT pt.id, pt.stock_ticker, pt.transaction_type, pt.quantity,
                      pt.transaction_date, pt.price_per_share, pt.account_id,
                      ia.name as account_name, ia.account_type, ia.bank
               FROM demo_portfolio_transactions pt
               LEFT JOIN demo_investment_accounts ia ON pt.account_id = ia.id
               WHERE pt.user_id = ?
               ORDER BY pt.transaction_date DESC, pt.id DESC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    transactions = [{
        'id': row['id'],
        'stock_ticker': row['stock_ticker'],
        'transaction_type': row['transaction_type'],
        'quantity': row['quantity'],
        'transaction_date': row['transaction_date'],
        'price_per_share': row['price_per_share'],
        'price_currency': 'EUR',  # Default to EUR for demo
        'account_id': row['account_id'],
        'account_name': row['account_name'],
        'account_type': row['account_type'],
        'bank': row['bank'],
    } for row in rows]
    return jsonify({'transactions': transactions})


@app.route('/api/demo/transactions', methods=['POST'])
@login_required
def add_demo_transaction():
    """Add a new demo transaction."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    stock_ticker = data.get('stock_ticker', '').upper().strip()
    transaction_type = data.get('transaction_type', '').upper().strip()
    quantity = data.get('quantity')
    transaction_date = data.get('transaction_date')
    account_id = data.get('account_id')
    provided_price = data.get('price_per_share')

    if not stock_ticker:
        return jsonify({'error': 'Stock ticker required'}), 400
    if transaction_type not in ['BUY', 'SELL']:
        return jsonify({'error': 'Transaction type must be BUY or SELL'}), 400
    if quantity is None or not isinstance(quantity, (int, float)) or quantity <= 0:
        return jsonify({'error': 'Valid quantity required (must be > 0)'}), 400
    if not transaction_date:
        return jsonify({'error': 'Transaction date required (YYYY-MM-DD)'}), 400

    quantity = round(float(quantity), 2)

    # Validate account_id if provided
    if account_id is not None:
        with get_db() as conn:
            cursor = conn.execute(
                'SELECT id FROM demo_investment_accounts WHERE id = ? AND user_id = ?',
                (account_id, request.user_id)
            )
            if not cursor.fetchone():
                return jsonify({'error': 'Invalid account_id'}), 400

    try:
        datetime.strptime(transaction_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    transaction_date = get_previous_weekday(transaction_date)

    # Use provided price or fetch from Yahoo
    if provided_price is not None:
        price_per_share = float(provided_price)
    else:
        try:
            price_per_share = fetch_stock_price(stock_ticker, transaction_date, use_cache=False)
            if price_per_share is not None:
                price_per_share = float(price_per_share)
        except Exception as e:
            return jsonify({'error': f'Could not fetch price for {stock_ticker} on {transaction_date}: {str(e)}'}), 400

    try:
        with get_db() as conn:
            cursor = conn.execute('''
                INSERT INTO demo_portfolio_transactions (user_id, account_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                RETURNING id
            ''', (request.user_id, account_id, stock_ticker, transaction_type, quantity, transaction_date, price_per_share))
            transaction_id = cursor.fetchone()['id']
    except Exception as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'id': transaction_id,
        'duplicate': False,
        'account_id': account_id,
        'stock_ticker': stock_ticker,
        'transaction_type': transaction_type,
        'quantity': quantity,
        'transaction_date': transaction_date,
        'price_per_share': price_per_share,
        'price_currency': 'EUR'
    })


@app.route('/api/demo/transactions/<int:transaction_id>', methods=['DELETE'])
@login_required
def delete_demo_transaction(transaction_id):
    """Delete a demo transaction."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT id FROM demo_portfolio_transactions WHERE id = ? AND user_id = ?',
            (transaction_id, request.user_id)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Transaction not found'}), 404

        conn.execute('DELETE FROM demo_portfolio_transactions WHERE id = ? AND user_id = ?',
                    (transaction_id, request.user_id))

    return jsonify({'success': True})


@app.route('/api/demo/holdings', methods=['GET'])
@login_required
def get_demo_holdings():
    """Get computed current holdings from demo transactions."""
    with get_db() as conn:
        cursor = conn.execute(
            '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
               FROM demo_portfolio_transactions WHERE user_id = ?
               ORDER BY transaction_date ASC, id ASC''',
            (request.user_id,)
        )
        rows = cursor.fetchall()

    holdings_map = {}

    for row in rows:
        ticker = row['stock_ticker']
        qty = row['quantity']
        price = row['price_per_share']
        date = row['transaction_date']
        tx_type = row['transaction_type']

        if ticker not in holdings_map:
            holdings_map[ticker] = {'quantity': 0, 'lots': []}

        if tx_type == 'BUY':
            holdings_map[ticker]['quantity'] += qty
            holdings_map[ticker]['lots'].append({'qty': qty, 'price': price, 'date': date})
        elif tx_type == 'SELL':
            holdings_map[ticker]['quantity'] -= qty
            remaining_sell = qty
            while remaining_sell > 0 and holdings_map[ticker]['lots']:
                lot = holdings_map[ticker]['lots'][0]
                if lot['qty'] <= remaining_sell:
                    remaining_sell -= lot['qty']
                    holdings_map[ticker]['lots'].pop(0)
                else:
                    lot['qty'] -= remaining_sell
                    remaining_sell = 0

    holdings = []
    for ticker, data in holdings_map.items():
        if data['quantity'] > 0:
            total_cost = sum(lot['qty'] * lot['price'] for lot in data['lots'])
            total_qty = sum(lot['qty'] for lot in data['lots'])
            avg_cost = total_cost / total_qty if total_qty > 0 else 0

            holdings.append({
                'stock_ticker': ticker,
                'quantity': data['quantity'],
                'cost_basis': round(avg_cost, 2)
            })

    return jsonify({'holdings': holdings})


def compute_demo_holdings_from_transactions(user_id, account_ids=None):
    """Helper to compute demo holdings using FIFO."""
    from investing_utils import fetch_eurusd_rate

    with get_db() as conn:
        if account_ids and len(account_ids) > 0:
            placeholders = ','.join('?' for _ in account_ids)
            cursor = conn.execute(
                f'''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
                   FROM demo_portfolio_transactions WHERE user_id = ? AND account_id IN ({placeholders})
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id, *account_ids)
            )
        else:
            cursor = conn.execute(
                '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
                   FROM demo_portfolio_transactions WHERE user_id = ? AND account_id IS NOT NULL
                   ORDER BY transaction_date ASC, id ASC''',
                (user_id,)
            )
        rows = cursor.fetchall()

    holdings_map = {}

    for row in rows:
        ticker = row['stock_ticker']
        qty = row['quantity']
        price = row['price_per_share']
        date = row['transaction_date']
        tx_type = row['transaction_type']

        if ticker not in holdings_map:
            holdings_map[ticker] = {'quantity': 0, 'lots': []}

        if tx_type == 'BUY':
            holdings_map[ticker]['quantity'] += qty
            # For demo, assume EUR prices
            eur_price = price
            holdings_map[ticker]['lots'].append({'qty': qty, 'eur_price': eur_price, 'date': date})
        elif tx_type == 'SELL':
            holdings_map[ticker]['quantity'] -= qty
            remaining_sell = qty
            while remaining_sell > 0 and holdings_map[ticker]['lots']:
                lot = holdings_map[ticker]['lots'][0]
                if lot['qty'] <= remaining_sell:
                    remaining_sell -= lot['qty']
                    holdings_map[ticker]['lots'].pop(0)
                else:
                    lot['qty'] -= remaining_sell
                    remaining_sell = 0

    holdings = []
    for ticker, data in holdings_map.items():
        if data['quantity'] > 0:
            total_eur_cost = sum(lot['qty'] * lot['eur_price'] for lot in data['lots'])
            total_qty = sum(lot['qty'] for lot in data['lots'])
            avg_eur_cost = total_eur_cost / total_qty if total_qty > 0 else 0

            holdings.append({
                'stock_ticker': ticker,
                'quantity': data['quantity'],
                'cost_basis_eur': round(avg_eur_cost, 2),
            })

    return holdings


@app.route('/api/demo/portfolio/composition', methods=['GET'])
@login_required
def get_demo_portfolio_composition():
    """Get demo portfolio composition with current values."""
    from investing_utils import get_stock_price_and_change, get_current_eurusd_rate

    account_ids_param = request.args.get('account_ids', '')
    account_ids = [int(x) for x in account_ids_param.split(',') if x.strip().isdigit()] if account_ids_param else None

    holdings = compute_demo_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({'composition': [], 'total_value_eur': 0, 'total_cost_eur': 0})

    eurusd_rate = get_current_eurusd_rate()
    composition = []
    total_value_eur = 0
    total_cost_eur = 0

    for h in holdings:
        ticker = h['stock_ticker']
        qty = h['quantity']
        cost_basis_eur = h['cost_basis_eur']

        price_data = get_stock_price_and_change(ticker)
        current_price = price_data.get('price', 0)
        change_1d = price_data.get('change_1d', 0)
        currency = price_data.get('currency', 'USD')

        if currency == 'USD':
            current_price_eur = current_price / eurusd_rate if eurusd_rate else current_price
        else:
            current_price_eur = current_price

        current_value_eur = qty * current_price_eur
        total_cost = qty * cost_basis_eur
        gain_loss_eur = current_value_eur - total_cost
        gain_loss_pct = (gain_loss_eur / total_cost * 100) if total_cost > 0 else 0

        composition.append({
            'ticker': ticker,
            'quantity': qty,
            'current_price_eur': round(current_price_eur, 2),
            'current_value_eur': round(current_value_eur, 2),
            'cost_basis_eur': round(cost_basis_eur, 2),
            'total_cost_eur': round(total_cost, 2),
            'gain_loss_eur': round(gain_loss_eur, 2),
            'gain_loss_pct': round(gain_loss_pct, 2),
            'change_1d': round(change_1d, 2) if change_1d else 0,
        })

        total_value_eur += current_value_eur
        total_cost_eur += total_cost

    composition.sort(key=lambda x: x['current_value_eur'], reverse=True)

    return jsonify({
        'composition': composition,
        'total_value_eur': round(total_value_eur, 2),
        'total_cost_eur': round(total_cost_eur, 2),
        'total_gain_loss_eur': round(total_value_eur - total_cost_eur, 2),
        'total_gain_loss_pct': round((total_value_eur - total_cost_eur) / total_cost_eur * 100, 2) if total_cost_eur > 0 else 0,
    })


@app.route('/api/demo/portfolio/performance-period', methods=['GET'])
@login_required
def get_demo_portfolio_performance_period():
    """Get demo portfolio historical performance for a time period."""
    from investing_utils import fetch_eurusd_rate, get_historical_price

    period = request.args.get('period', '1M')
    account_ids_param = request.args.get('account_ids', '')
    account_ids = [int(x) for x in account_ids_param.split(',') if x.strip().isdigit()] if account_ids_param else None

    # Calculate date range
    today = datetime.now().date()
    if period == '1W':
        start_date = today - timedelta(days=7)
    elif period == '1M':
        start_date = today - timedelta(days=30)
    elif period == '3M':
        start_date = today - timedelta(days=90)
    elif period == '6M':
        start_date = today - timedelta(days=180)
    elif period == '1Y':
        start_date = today - timedelta(days=365)
    elif period == 'YTD':
        start_date = datetime(today.year, 1, 1).date()
    elif period == 'ALL':
        start_date = today - timedelta(days=365*5)
    else:
        start_date = today - timedelta(days=30)

    # Get all transactions for the user within accounts
    with get_db() as conn:
        if account_ids and len(account_ids) > 0:
            placeholders = ','.join('?' for _ in account_ids)
            cursor = conn.execute(
                f'''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
                   FROM demo_portfolio_transactions WHERE user_id = ? AND account_id IN ({placeholders})
                   ORDER BY transaction_date ASC, id ASC''',
                (request.user_id, *account_ids)
            )
        else:
            cursor = conn.execute(
                '''SELECT stock_ticker, transaction_type, quantity, transaction_date, price_per_share
                   FROM demo_portfolio_transactions WHERE user_id = ? AND account_id IS NOT NULL
                   ORDER BY transaction_date ASC, id ASC''',
                (request.user_id,)
            )
        transactions = cursor.fetchall()

    if not transactions:
        return jsonify({'data': [], 'period': period})

    # Build holdings timeline
    eurusd_rate = get_current_eurusd_rate()
    date_range = []
    current_date = start_date
    while current_date <= today:
        if current_date.weekday() < 5:  # Weekdays only
            date_range.append(current_date.isoformat())
        current_date += timedelta(days=1)

    performance_data = []
    for date_str in date_range:
        # Calculate holdings at this date
        holdings_at_date = {}
        total_invested = 0

        for tx in transactions:
            if tx['transaction_date'] <= date_str:
                ticker = tx['stock_ticker']
                if ticker not in holdings_at_date:
                    holdings_at_date[ticker] = {'quantity': 0, 'cost': 0}

                if tx['transaction_type'] == 'BUY':
                    holdings_at_date[ticker]['quantity'] += tx['quantity']
                    holdings_at_date[ticker]['cost'] += tx['quantity'] * tx['price_per_share']
                elif tx['transaction_type'] == 'SELL':
                    holdings_at_date[ticker]['quantity'] -= tx['quantity']

        # Calculate total value at this date
        total_value = 0
        for ticker, data in holdings_at_date.items():
            if data['quantity'] > 0:
                total_invested += data['cost']
                price = get_historical_price(ticker, date_str)
                if price:
                    # Assume EUR for simplicity in demo
                    total_value += data['quantity'] * price

        if total_value > 0:
            performance_data.append({
                'date': date_str,
                'value': round(total_value, 2),
                'invested': round(total_invested, 2),
            })

    return jsonify({'data': performance_data, 'period': period})


@app.route('/api/demo/portfolio/top-movers', methods=['GET'])
@login_required
def get_demo_top_movers():
    """Get top gainers and losers from demo portfolio."""
    from investing_utils import get_stock_price_and_change, get_current_eurusd_rate

    account_ids_param = request.args.get('account_ids', '')
    account_ids = [int(x) for x in account_ids_param.split(',') if x.strip().isdigit()] if account_ids_param else None

    holdings = compute_demo_holdings_from_transactions(request.user_id, account_ids)

    if not holdings:
        return jsonify({'gainers': [], 'losers': []})

    eurusd_rate = get_current_eurusd_rate()
    movers = []

    for h in holdings:
        ticker = h['stock_ticker']
        qty = h['quantity']

        price_data = get_stock_price_and_change(ticker)
        current_price = price_data.get('price', 0)
        change_1d = price_data.get('change_1d', 0)
        currency = price_data.get('currency', 'USD')

        if currency == 'USD':
            current_price_eur = current_price / eurusd_rate if eurusd_rate else current_price
        else:
            current_price_eur = current_price

        current_value_eur = qty * current_price_eur
        day_change_eur = current_value_eur * (change_1d / 100) if change_1d else 0

        movers.append({
            'ticker': ticker,
            'change_pct': round(change_1d, 2) if change_1d else 0,
            'change_eur': round(day_change_eur, 2),
            'value_eur': round(current_value_eur, 2),
        })

    movers.sort(key=lambda x: x['change_pct'], reverse=True)

    gainers = [m for m in movers if m['change_pct'] > 0][:5]
    losers = [m for m in movers if m['change_pct'] < 0][-5:][::-1]

    return jsonify({'gainers': gainers, 'losers': losers})


# Banks and account types endpoints (reuse from investing)
@app.route('/api/demo/banks', methods=['GET'])
def get_demo_banks():
    """Get available banks."""
    return jsonify({'banks': BANKS})


@app.route('/api/demo/account-types', methods=['GET'])
def get_demo_account_types():
    """Get available account types."""
    return jsonify({'account_types': ACCOUNT_TYPES})


# =============================================================================
# ALPHAWISE MODEL PORTFOLIO API ROUTES
# =============================================================================

# Default stocks for the model portfolio (20 large US/EU stocks at 5% each)
ALPHAWISE_DEFAULT_STOCKS = [
    # US Large Caps (10 stocks)
    ('AAPL', 5.0),   # Apple
    ('MSFT', 5.0),   # Microsoft
    ('GOOGL', 5.0),  # Alphabet
    ('AMZN', 5.0),   # Amazon
    ('NVDA', 5.0),   # NVIDIA
    ('META', 5.0),   # Meta
    ('TSLA', 5.0),   # Tesla
    ('JPM', 5.0),    # JPMorgan
    ('V', 5.0),      # Visa
    ('JNJ', 5.0),    # Johnson & Johnson
    # EU Large Caps (10 stocks)
    ('ASML', 5.0),   # ASML
    ('NVO', 5.0),    # Novo Nordisk
    ('SAP', 5.0),    # SAP
    ('MC.PA', 5.0),  # LVMH
    ('OR.PA', 5.0),  # L'Oreal
    ('SIE.DE', 5.0), # Siemens
    ('AZN', 5.0),    # AstraZeneca
    ('SHEL', 5.0),   # Shell
    ('TTE', 5.0),    # TotalEnergies
    ('SAN', 5.0),    # Santander
]


def init_alphawise_model_portfolio():
    """Initialize the model portfolio with default stocks if empty."""
    with get_db() as conn:
        # Check if table exists and has data
        try:
            cursor = conn.execute('SELECT COUNT(*) FROM alphawise_model_portfolio')
            count = cursor.fetchone()[0]
            if count == 0:
                # Insert default stocks
                for ticker, allocation in ALPHAWISE_DEFAULT_STOCKS:
                    conn.execute(
                        'INSERT INTO alphawise_model_portfolio (stock_ticker, allocation_pct) VALUES (?, ?)',
                        (ticker, allocation)
                    )
                logger.info(f"Initialized AlphaWise model portfolio with {len(ALPHAWISE_DEFAULT_STOCKS)} stocks")
        except Exception as e:
            # Table might not exist yet
            logger.warning(f"Could not initialize model portfolio: {e}")


# Initialize model portfolio on startup
init_alphawise_model_portfolio()


@app.route('/api/demo/model-portfolio', methods=['GET'])
def get_model_portfolio():
    """Get the AlphaWise model portfolio composition with current prices."""
    from investing_utils import get_stock_price_and_change, get_current_eurusd_rate

    with get_db() as conn:
        cursor = conn.execute('''
            SELECT stock_ticker, allocation_pct
            FROM alphawise_model_portfolio
            ORDER BY allocation_pct DESC, stock_ticker ASC
        ''')
        stocks = [{'ticker': row['stock_ticker'], 'allocation_pct': row['allocation_pct']} for row in cursor.fetchall()]

    if not stocks:
        return jsonify({'holdings': [], 'total_allocation': 0})

    # Fetch current prices for all stocks
    eurusd_rate = get_current_eurusd_rate()
    holdings = []

    # Color palette for pie chart (same as PortfolioComposition)
    colors = [
        '#22c55e', '#3b82f6', '#f97316', '#a855f7', '#ec4899',
        '#14b8a6', '#eab308', '#ef4444', '#6366f1', '#84cc16',
        '#06b6d4', '#f43f5e', '#8b5cf6', '#10b981', '#f59e0b',
        '#6b7280', '#d946ef', '#0ea5e9', '#78716c', '#fb923c'
    ]

    for i, stock in enumerate(stocks):
        ticker = stock['ticker']
        allocation = stock['allocation_pct']

        price_data = get_stock_price_and_change(ticker)
        current_price = price_data.get('price', 0) if price_data else 0
        change_1d = price_data.get('change_1d') if price_data else None

        holdings.append({
            'ticker': ticker,
            'weight': round(allocation, 1),
            'current_price': round(current_price, 2),
            'change_1d': round(change_1d, 2) if change_1d is not None else None,
            'color': colors[i % len(colors)],
        })

    total_allocation = sum(h['weight'] for h in holdings)

    return jsonify({
        'holdings': holdings,
        'total_allocation': round(total_allocation, 1),
        'eurusd_rate': eurusd_rate,
    })


@app.route('/api/demo/model-portfolio/performance', methods=['GET'])
def get_model_portfolio_performance():
    """Get model portfolio performance since 01/01/2023 vs S&P 500.
    Assumes 100% capital allocation at start with current portfolio weights.
    """
    import yfinance as yf
    import pandas as pd
    from investing_utils import get_yfinance_ticker

    # Get model portfolio composition
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT stock_ticker, allocation_pct
            FROM alphawise_model_portfolio
            ORDER BY allocation_pct DESC
        ''')
        stocks = [{'ticker': row['stock_ticker'], 'weight': row['allocation_pct'] / 100.0} for row in cursor.fetchall()]

    if not stocks:
        return jsonify({'data': [], 'message': 'No stocks in model portfolio'})

    # Date range: 01/01/2023 to today
    start_date = '2023-01-01'
    end_date = datetime.now().strftime('%Y-%m-%d')

    # Get S&P 500 benchmark (SPY)
    benchmark_ticker = 'SPY'

    try:
        # Fetch all stock data
        portfolio_tickers = [get_yfinance_ticker(s['ticker']) for s in stocks]
        all_tickers = portfolio_tickers + [benchmark_ticker]

        # Remove duplicates while preserving order
        seen = set()
        unique_tickers = []
        for t in all_tickers:
            if t not in seen:
                seen.add(t)
                unique_tickers.append(t)
        all_tickers = unique_tickers

        logger.info(f"Fetching data for tickers: {all_tickers}")
        data = yf.download(all_tickers, start=start_date, end=end_date, progress=False, auto_adjust=True)

        if data.empty:
            return jsonify({'data': [], 'message': 'No price data available'})

        # Get closing prices - handle both single and multi-ticker formats
        if len(all_tickers) == 1:
            close_prices = data[['Close']].copy()
            close_prices.columns = [all_tickers[0]]
        else:
            # Multi-ticker: data has MultiIndex columns like ('Close', 'AAPL')
            close_prices = data['Close'].copy()

        logger.info(f"Close prices columns: {list(close_prices.columns)}")

        # Calculate portfolio performance
        # Normalize to 100 at start
        first_valid_idx = close_prices.first_valid_index()
        if first_valid_idx is None:
            return jsonify({'data': [], 'message': 'No valid price data'})

        # Get first valid prices for normalization
        first_prices = close_prices.loc[first_valid_idx]
        logger.info(f"First valid date: {first_valid_idx}")
        logger.info(f"First prices: {first_prices.to_dict()}")

        # Calculate daily portfolio value (weighted sum of normalized prices)
        portfolio_values = []
        benchmark_values = []
        dates = []

        for date_idx, row in close_prices.iterrows():
            date_str = date_idx.strftime('%Y-%m-%d')

            # Calculate weighted portfolio value
            portfolio_value = 0
            valid_weight = 0
            for stock in stocks:
                yf_ticker = get_yfinance_ticker(stock['ticker'])
                if yf_ticker in row.index and pd.notna(row[yf_ticker]) and yf_ticker in first_prices.index and pd.notna(first_prices[yf_ticker]):
                    normalized = (row[yf_ticker] / first_prices[yf_ticker]) * 100
                    portfolio_value += normalized * stock['weight']
                    valid_weight += stock['weight']

            # Normalize by valid weight to handle missing data
            if valid_weight > 0:
                portfolio_value = portfolio_value / valid_weight

            # Benchmark value - check using row.index for proper Series lookup
            benchmark_value = None
            if benchmark_ticker in row.index and pd.notna(row[benchmark_ticker]) and benchmark_ticker in first_prices.index and pd.notna(first_prices[benchmark_ticker]):
                benchmark_value = (row[benchmark_ticker] / first_prices[benchmark_ticker]) * 100

            if portfolio_value > 0:
                dates.append(date_str)
                portfolio_values.append(round(portfolio_value, 2))
                benchmark_values.append(round(benchmark_value, 2) if benchmark_value else None)

        # Build response data
        result_data = []
        for i in range(len(dates)):
            result_data.append({
                'date': dates[i],
                'portfolio': portfolio_values[i],
                'benchmark': benchmark_values[i],
            })

        # Calculate summary stats
        if len(portfolio_values) > 0:
            portfolio_return = portfolio_values[-1] - 100
            benchmark_return = (benchmark_values[-1] - 100) if benchmark_values[-1] else None
        else:
            portfolio_return = 0
            benchmark_return = None

        return jsonify({
            'data': result_data,
            'summary': {
                'portfolio_return': round(portfolio_return, 2),
                'benchmark_return': round(benchmark_return, 2) if benchmark_return else None,
                'start_date': start_date,
                'end_date': end_date,
            }
        })

    except Exception as e:
        print(f"Error fetching model portfolio performance: {e}")
        return jsonify({'data': [], 'error': str(e)}), 500


@app.route('/api/demo/admin/model-portfolio', methods=['GET'])
@admin_required
def get_model_portfolio_admin():
    """Get the model portfolio for admin editing."""
    with get_db() as conn:
        cursor = conn.execute('''
            SELECT id, stock_ticker, allocation_pct, created_at, updated_at
            FROM alphawise_model_portfolio
            ORDER BY allocation_pct DESC, stock_ticker ASC
        ''')
        stocks = [{
            'id': row['id'],
            'ticker': row['stock_ticker'],
            'allocation_pct': row['allocation_pct'],
            'created_at': row['created_at'],
            'updated_at': row['updated_at'],
        } for row in cursor.fetchall()]

    total_allocation = sum(s['allocation_pct'] for s in stocks)

    return jsonify({
        'stocks': stocks,
        'total_allocation': round(total_allocation, 1),
    })


@app.route('/api/demo/admin/model-portfolio', methods=['POST'])
@admin_required
def add_model_portfolio_stock():
    """Add a stock to the model portfolio."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    ticker = data.get('ticker', '').upper().strip()
    allocation = data.get('allocation_pct', 5.0)

    if not ticker:
        return jsonify({'error': 'Ticker is required'}), 400

    try:
        allocation = float(allocation)
        if allocation <= 0 or allocation > 100:
            return jsonify({'error': 'Allocation must be between 0 and 100'}), 400
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid allocation'}), 400

    with get_db() as conn:
        # Check if ticker already exists
        cursor = conn.execute(
            'SELECT id FROM alphawise_model_portfolio WHERE stock_ticker = ?',
            (ticker,)
        )
        if cursor.fetchone():
            return jsonify({'error': 'Ticker already in portfolio'}), 400

        cursor = conn.execute(
            'INSERT INTO alphawise_model_portfolio (stock_ticker, allocation_pct) VALUES (?, ?)',
            (ticker, allocation)
        )
        new_id = cursor.lastrowid

    return jsonify({
        'success': True,
        'id': new_id,
        'ticker': ticker,
        'allocation_pct': allocation,
    })


@app.route('/api/demo/admin/model-portfolio/<int:stock_id>', methods=['PUT'])
@admin_required
def update_model_portfolio_stock(stock_id):
    """Update a stock's allocation in the model portfolio."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    allocation = data.get('allocation_pct')

    try:
        allocation = float(allocation)
        if allocation <= 0 or allocation > 100:
            return jsonify({'error': 'Allocation must be between 0 and 100'}), 400
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid allocation'}), 400

    with get_db() as conn:
        cursor = conn.execute(
            'SELECT id FROM alphawise_model_portfolio WHERE id = ?',
            (stock_id,)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Stock not found'}), 404

        if USE_POSTGRES:
            conn.execute(
                'UPDATE alphawise_model_portfolio SET allocation_pct = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (allocation, stock_id)
            )
        else:
            conn.execute(
                'UPDATE alphawise_model_portfolio SET allocation_pct = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (allocation, stock_id)
            )

    return jsonify({'success': True})


@app.route('/api/demo/admin/model-portfolio/<int:stock_id>', methods=['DELETE'])
@admin_required
def delete_model_portfolio_stock(stock_id):
    """Remove a stock from the model portfolio."""
    with get_db() as conn:
        cursor = conn.execute(
            'SELECT id FROM alphawise_model_portfolio WHERE id = ?',
            (stock_id,)
        )
        if not cursor.fetchone():
            return jsonify({'error': 'Stock not found'}), 404

        conn.execute('DELETE FROM alphawise_model_portfolio WHERE id = ?', (stock_id,))

    return jsonify({'success': True})


@app.route('/api/demo/admin/model-portfolio/reset', methods=['POST'])
@admin_required
def reset_model_portfolio():
    """Reset the model portfolio to default stocks."""
    with get_db() as conn:
        conn.execute('DELETE FROM alphawise_model_portfolio')
        for ticker, allocation in ALPHAWISE_DEFAULT_STOCKS:
            conn.execute(
                'INSERT INTO alphawise_model_portfolio (stock_ticker, allocation_pct) VALUES (?, ?)',
                (ticker, allocation)
            )

    return jsonify({
        'success': True,
        'message': f'Reset to {len(ALPHAWISE_DEFAULT_STOCKS)} default stocks'
    })


if __name__ == '__main__':
    app.run(debug=True, port=5001)